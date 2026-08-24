# -*- coding: utf-8 -*-
# Copyright (c) 2026 윤주호. All rights reserved.
# 무단 복제·배포·수정을 금합니다.
"""빈 양식(template_form.xlsx)에 채워서 완성 성적서(xlsx)를 만든다.
   항목 순서(item_order)대로 9행부터 채움 — 양식이 통일돼 있으므로 자재가 몇 종이든 동일 로직."""
import os, re, shutil, subprocess, sys, tempfile, time
from datetime import date
import database as db
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

BASE_DIR = os.path.dirname(__file__)
TEMPLATE = os.path.join(BASE_DIR, "template_form.xlsx")
STANDARD_TEMPLATE = os.path.join(BASE_DIR, "standard_template.xlsx")  # 기준서(SAM 양식)
LOGO_PATH = os.path.join(BASE_DIR, "logo.png")  # 회사 로고 — 성적서·기준서 좌상단에 동일하게 삽입
OUT_DIR = os.path.join(db.DATA_DIR, "성적서 발행")
os.makedirs(OUT_DIR, exist_ok=True)

# 로고 크기·위치 — 엑셀 이미지 크기 설정 기준
# 성적서: 요청받은 그대로(너비 8.52cm 고정, 높이는 새 로고의 실제 비율에 맞춰 재계산 — 안 찌그러지게)
LOGO_OFFSET_CM = 0.5  # A4 좌상단에서 아래로 0.5cm, 오른쪽으로 0.5cm
CERT_LOGO_WIDTH_CM = 8.52
# 기준서: 상단 행(로고 바로 아래 "제품 구성품 검사 주요사항" 제목)과 안 겹치게 축소(옵션 A)
STD_LOGO_HEIGHT_CM = 1.1

STD_FIRST_ROW = 10  # 기준서 항목표 첫 행(순번 A)
STD_LAST_ROW = 29   # 기준서 항목표 마지막 행(순번 T, 최대 20항목)


def compute_drawing_no(material_no):
    """자재번호에서 도면번호 계산: 앞에 'A' 붙이고 'P'를 '-'로 치환. 예: 602106P246 -> A602106-246"""
    return "A" + (material_no or "").replace("P", "-")


def _logo_aspect_ratio():
    """로고 원본 파일의 가로:세로 비율(가로/세로). 못 읽으면 None."""
    try:
        from PIL import Image as PILImage
        with PILImage.open(LOGO_PATH) as im:
            return im.width / im.height
    except Exception:
        return None


def _insert_logo(ws, height_cm=None, width_cm=None):
    """
    회사 로고를 시트 좌상단에 삽입 — 셀 경계가 아니라 페이지 물리적 위치(A4 기준
    위에서 0.5cm, 왼쪽에서 0.5cm) 그대로 배치해야 해서 OneCellAnchor를 직접 구성한다.
    height_cm/width_cm 중 하나만 넘기면 로고 원본 비율 그대로 나머지를 계산해서 찌그러짐 방지.
    """
    if not os.path.exists(LOGO_PATH):
        return
    ratio = _logo_aspect_ratio() or (CERT_LOGO_WIDTH_CM / 1.89)  # 못 읽으면 예전 비율로 대체

    if width_cm is not None and height_cm is None:
        height_cm = width_cm / ratio
    elif height_cm is not None and width_cm is None:
        width_cm = height_cm * ratio
    elif width_cm is None and height_cm is None:
        width_cm, height_cm = CERT_LOGO_WIDTH_CM, CERT_LOGO_WIDTH_CM / ratio

    img = XLImage(LOGO_PATH)
    size = XDRPositiveSize2D(cm_to_EMU(width_cm), cm_to_EMU(height_cm))
    marker = AnchorMarker(col=0, colOff=cm_to_EMU(LOGO_OFFSET_CM),
                          row=0, rowOff=cm_to_EMU(LOGO_OFFSET_CM))
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)


def _find_soffice():
    """PATH에 없으면 Windows의 흔한 설치 위치를 직접 탐색."""
    found = shutil.which("soffice") or shutil.which("soffice.exe")
    if found:
        return found
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


SOFFICE = _find_soffice()
if SOFFICE is None:
    print("[경고] LibreOffice(soffice)를 찾을 수 없습니다. "
          "설치 경로가 다르다면 report_builder.py의 candidates 목록에 직접 추가하세요.",
          file=sys.stderr)

MEAS_COLS = ["E", "F", "G", "H", "I", "J"]
FIRST_ROW = 9   # 항목 A가 들어가는 행
LAST_ROW = 27   # 항목 S가 들어가는 마지막 행(템플릿이 지원하는 최대 항목수)
ITEM_COLS = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "N", "O", "P", "Q"]


def _safe_folder_name(s):
    """폴더명에 못 쓰는 문자 제거/치환."""
    s = (s or "미상").strip()
    return re.sub(r'[\\/:*?"<>|]', "-", s) or "미상"


def _safe_filename_part(s):
    """
    파일명에 못 쓰는 문자(\\ / : * ? " < > | 및 제어문자)는 치환하지 않고 그냥 제거.
    나머지는 그대로 둔다(한글·공백·괄호·쉼표 등은 파일명에 문제없음).

    엑셀에서 가져온 값에는 전각 콜론(：)이나 줄바꿈이 섞여 들어오는 경우가 있어서
    같은 규칙(제거)을 전각 문자에도 적용하고, 줄바꿈·연속 공백은 공백 하나로 줄인다.
    """
    s = (s or "").strip()
    s = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "", s)
    s = re.sub(r"[：＊？＂＜＞｜／＼]", "", s)   # 전각 형태도 같은 규칙으로 제거
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def report_output_dir():
    """
    최상위 '성적서 발행' 폴더 밑에 '생성날짜(오늘)' 폴더 하나만 둔다.
    같은 날 여러 업체·자재가 생성돼도 전부 이 폴더 안에 모이고, 파일명으로 구분한다.
    """
    today = date.today().strftime("%Y-%m-%d")
    path = os.path.join(OUT_DIR, today)
    os.makedirs(path, exist_ok=True)
    return path


def build_report_filename(supplier, material_no, product_name):
    """'260821_ACE_600005P086_둥근머리 볼트(M4*16L,STS304)' 형식의 파일명(확장자 제외)."""
    date_prefix = date.today().strftime("%y%m%d")
    supplier_part = _safe_filename_part(supplier) or "미상"
    material_part = _safe_filename_part(material_no) or "미상"
    product_part = _safe_filename_part(product_name)
    name = f"{date_prefix}_{supplier_part}_{material_part}"
    if product_part:
        name += f"_{product_part}"
    return name


def _dedupe_path(path):
    """
    같은 날 같은 자재를 여러 번 검사하면 파일명이 겹칠 수 있어서(예전엔 발주번호별
    폴더로 자동 구분됐지만 이제 날짜 폴더 하나로 합쳐짐) — 이미 있으면 (2), (3)...을 붙인다.
    """
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    n = 2
    while True:
        candidate = f"{root}({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


# Excel 이미지 크기 설정 기준(cm) → 픽셀 변환 (96 DPI: 1cm ≈ 37.795px)
CM_TO_PX = 96 / 2.54
SIGNATURE_WIDTH_CM = 5.7
SIGNATURE_HEIGHT_CM = 2.3
SIGNATURE_WIDTH_PX = round(SIGNATURE_WIDTH_CM * CM_TO_PX)   # 215px
SIGNATURE_HEIGHT_PX = round(SIGNATURE_HEIGHT_CM * CM_TO_PX)  # 87px


def is_critical_aql(aql):
    """AQL이 0.65인 항목은 중요항목으로 취급해 항목기호 앞에 '*'를 자동으로 붙인다."""
    if aql is None:
        return False
    try:
        return abs(float(aql) - 0.65) < 1e-9
    except (TypeError, ValueError):
        return False


def item_label(item_name, aql):
    """중요항목은 '*A' 형태로, 아니면 그냥 'A'로.

    중요항목 판단 기준 두 가지:
      1) AQL이 0.65 (원칙 — 시스템이 자동으로 붙여준다)
      2) 원본 성적서에 이미 '*'가 찍혀 있어 item_name에 그대로 들어온 경우

    ※ 이 함수는 반드시 멱등이어야 한다. 예전엔 item_name이 '*H'인 채로 들어오면
      앞에 '*'를 한 번 더 붙여서 성적서에 '**H'로 찍히는 버그가 있었다(364건).
    """
    raw = item_name or ""
    name = raw.strip().lstrip("*").strip()
    critical = is_critical_aql(aql) or raw.strip().startswith("*")
    return f"*{name}" if critical else name


def format_aql(aql):
    """AQL을 사람이 보는 표기로 바꾼다. DB에는 '퍼센트10'으로 저장되지만 화면·성적서에는 '10%'로 나가야 함.
    '퍼센트10' -> '10%',  '퍼센트10.0' -> '10%',  '전수' -> '전수',  4 -> '4'
    (성적서·기준서·웹화면이 전부 이 함수 하나를 쓰도록 모아둠 — 예전엔 세 군데에 복사돼 있었다)
    """
    if aql is None:
        return ""
    text = str(aql).strip()
    if text.startswith("퍼센트"):
        num = text[3:]
        try:
            num = f"{float(num):g}"   # "10.0" -> "10", "10.5"는 그대로
        except ValueError:
            pass
        return f"{num}%"
    return text


def _format_gauge_expiry(iso_date_str):
    """'2027-01-31' -> '27.01.31' (YY.MM.DD). 형식이 다르면 원본 그대로 표시."""
    if not iso_date_str:
        return ""
    try:
        d = date.fromisoformat(iso_date_str)
        return f"{d.year % 100:02d}.{d.month:02d}.{d.day:02d}"
    except ValueError:
        return iso_date_str


def _build_remark_richtext(remarks):
    """
    remarks: {"inspector": "이름: 코멘트" 또는 None, "manager": ..., "approver": ...}
    검사자=검정, 중간관리자=진한녹색, 최종결정권자=진한보라 로 색을 나눠 한 셀에 표시.
    아직 안 적힌 칸도 빈 줄로 항상 같이 인쇄해서(수기로 나중에 적을 수 있게) 형식을 고정한다.
    """
    black = InlineFont(rFont="맑은 고딕", color="000000")
    dark_green = InlineFont(rFont="맑은 고딕", color="006400")
    dark_purple = InlineFont(rFont="맑은 고딕", color="4B0082")

    lines = [
        ("inspector", black, "검사자"),
        ("manager", dark_green, "중간 관리자 :"),
        ("approver", dark_purple, "부장 :"),
    ]

    blocks = ["비고 : "]
    for key, font, label in lines:
        text = (remarks or {}).get(key) or ""
        blocks.append("\n")
        blocks.append(TextBlock(font, f"{label} {text}".rstrip()))

    return CellRichText(*blocks)


def _fill_sheet(ws, material_no, product_name, header, results, overall,
                 approver=None, signature_path=None,
                 per_cycle_label=None, total_time_label=None, remarks=None, approval_type=None):
    """
    시트 하나(ws)에 성적서 내용을 채운다. build_report(단일 자재)와
    build_group_report(조립품, 시트 여러 개)가 공용으로 쓰는 핵심 로직.
    signature_path 삽입은 그룹일 때 마지막 시트에만 하도록 호출부에서 조절한다(place_signature).
    반환: signature_error 또는 None
    """
    ws["A4"] = f"품명 및 규격：{product_name}"
    ws["K4"] = f"자재번호：{material_no}"
    ws["A3"] = f"검사 날짜 : {header.get('inspect_date','')}"
    ws["A5"] = f"납품 업체：{header.get('vendor','')}"
    ws["F5"] = f"발주 번호：{header.get('po_no','')}"
    ws["A6"] = f"입고 날짜 / 로트번호 : {header.get('lot','')}"
    qty_raw = header.get("qty")
    try:
        ws["F6"] = int(qty_raw)
    except (TypeError, ValueError):
        ws["F6"] = ""
    ws["N30"] = f"검사자：{header.get('inspector','')}"
    if per_cycle_label:
        ws["K5"] = f"개당 측정 시간 : {per_cycle_label}"
    if total_time_label:
        ws["K6"] = f"총 측정 시간 : {total_time_label}"

    _insert_logo(ws, width_cm=CERT_LOGO_WIDTH_CM)  # 성적서: 너비 8.52cm 고정, 높이는 비율대로

    # 인쇄 시 가로·세로 폭이 항상 1페이지 안에 들어가도록 강제 (원본 템플릿의 고정 배율 54%는
    # 열이 많으면 폭이 넘쳐서 계측기유효기간(Q열)이 다음 페이지로 밀려나는 문제가 있었음)
    # 주의: ws.page_setup.fitToPage 세터는 openpyxl에서 _parent 연결이 끊긴 상태로 읽힐 때
    # AttributeError가 나서, sheet_properties.pageSetUpPr을 직접 건드려 우회함.
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    red = Font(name="맑은 고딕", color="FF0000", bold=True)
    blue = Font(name="맑은 고딕", color="0000FF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    center_shrink = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)

    # 공유 템플릿 파일 자체에 예전 자재의 값이 여러 행에 고정 텍스트로 박혀있을 수 있으므로,
    # 이번 자재가 실제로 쓰는 행을 채우기 전에 항목표 전체 범위(A~S, 최대 19행)를 먼저 비운다.
    for row in range(FIRST_ROW, LAST_ROW + 1):
        for col in ITEM_COLS:
            ws[f"{col}{row}"] = None
            ws[f"{col}{row}"].font = Font(name="맑은 고딕")

    for idx, r in enumerate(results):
        row = FIRST_ROW + idx
        if r.get("item_name") is not None:
            ws[f"A{row}"] = item_label(r["item_name"], r.get("aql"))  # 중요항목(AQL 0.65)은 자동으로 "*" 표시
        ws[f"A{row}"].alignment = center  # 번호(항목기호)
        if r.get("spec_display") is not None:
            ws[f"B{row}"] = r["spec_display"]
        # AQL 칸은 두 줄로 — 본사 양식이라 AQL 표기는 유지하되, 실제 판정 근거를 같이 적는다.
        #   AQL 4.0
        #   샘플 6개/Ac1 이내 합격   (또는 Ac=0이면 예전처럼 "무결점")
        # 실제 계측은 항상 최대 6개까지만 하지만(inspect_form.html 참고), 합격 허용 불량개수(Ac)는
        # AQL로 계산된 진짜 표본수(sample_qty)에 해당하는 KS Q ISO 2859-1 표준표 값을 쓴다
        # (aql_ac_allowance(), 사용자 확정 — 6개는 참고 표본, Ac는 표준표 기준).
        if r.get("aql") is not None:
            aql_text = format_aql(r["aql"])
            sample_qty = r.get("sample_qty")
            ac_allowance = r.get("ac_allowance")
            if sample_qty:
                measured_n = min(sample_qty, 6)
                if ac_allowance:
                    aql_text = f"{aql_text}\n샘플 {measured_n}개/Ac{ac_allowance} 이내 합격"
                else:
                    aql_text = f"{aql_text}\n샘플 {measured_n}개/무결점"
            ws[f"C{row}"] = aql_text
        # ※ wrap_text 없으면 LibreOffice가 PDF 변환할 때 줄바꿈을 무시하고 한 줄로 뭉갠다
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center",
                                            wrap_text=True, shrink_to_fit=True)
        if r.get("sample_qty") is not None:
            ws[f"D{row}"] = r["sample_qty"]
        ws[f"D{row}"].alignment = center  # 샘플 수량
        if r.get("inspect_method") is not None:
            method_cell = ws[f"P{row}"]
            method_cell.value = r["inspect_method"]
            method_cell.alignment = center_shrink

        lower = r.get("lower_limit")
        upper = r.get("upper_limit")
        for i, col in enumerate(MEAS_COLS):
            val = r["values"][i] if i < len(r["values"]) else None
            cell = ws[f"{col}{row}"]
            cell.value = val
            out_of_spec = False
            if isinstance(val, (int, float)) and (lower is not None or upper is not None):
                if lower is not None and val < lower:
                    out_of_spec = True
                if upper is not None and val > upper:
                    out_of_spec = True
            cell.font = Font(name="맑은 고딕", size=15, color="FF0000" if out_of_spec else "000000")

        if r["max"] is not None:
            ws[f"K{row}"] = r["max"]
        ws[f"K{row}"].alignment = center
        if r["min"] is not None:
            ws[f"N{row}"] = r["min"]
        ws[f"N{row}"].alignment = center
        cell = ws[f"O{row}"]
        cell.value = r["verdict"]
        cell.font = red if r["verdict"] == "불합격" else blue
        cell.alignment = center
        expiry = _format_gauge_expiry(r.get("gauge_expiry"))
        if expiry:
            ws[f"Q{row}"] = expiry
        ws[f"Q{row}"].alignment = center

    if approval_type == "special":
        mark = "□ 합격      □ 불합격      ■ 특채"
        mark_color = "FF8C00"
    else:
        # 합격 또는 검토필요 후 정상 승인 → ■합격 처리 (승인자의 최종 결정이 합격임)
        is_pass = overall in ("합격", "검토필요")
        mark = "■ 합격      □ 불합격      □ 특채" if is_pass else "□ 합격      ■ 불합격      □ 특채"
        mark_color = "0000FF" if is_pass else "FF0000"
    result_cell = ws["A29"]
    result_cell.value = f"검사 결과：   {mark}"
    result_cell.font = Font(name="맑은 고딕", bold=True, color=mark_color)

    remark_cell = ws["A28"]
    remark_cell.value = _build_remark_richtext(remarks)
    remark_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws["A30"] = f"검토：{approver or ''}"
    signature_error = None
    if signature_path:
        if not os.path.exists(signature_path):
            signature_error = f"서명 파일을 찾을 수 없어: {signature_path}"
        else:
            try:
                img = XLImage(signature_path)
                img.width = SIGNATURE_WIDTH_PX
                img.height = SIGNATURE_HEIGHT_PX
                ws.add_image(img, "C30")
            except Exception as e:
                signature_error = f"서명 이미지 삽입 실패: {e}"
    return signature_error


def _fill_standard_sheet(ws, material_no, product_name, drawing_version, revision_date,
                          edition, unit, results):
    """
    기준서(SAM 양식) 시트 하나에 내용을 채운다. standard_template.xlsx 구조 그대로 준수:
    A5/C5=품명규격, A7/C7=자재코드, F7=단위, A8/C8=도면번호, D8=도면버전,
    F5=날짜(제O판), F8=페이지번호, 항목표는 10행부터(순번A/검사항목B/검사장비및방법E/AQL허용수준G).
    """
    drawing_no = compute_drawing_no(material_no)

    ws["C5"] = product_name or ""
    ws["C7"] = material_no
    ws["F7"] = f"단위 ：{unit or 'mm'}"
    ws["C8"] = drawing_no
    ws["D8"] = f"도면 버전 ：{drawing_version or '1'}"
    ws["F8"] = "페이지 번호 ： 1  /  1"

    date_label = revision_date or ""
    try:
        y, m, d = date_label.split("-")
        date_label = f"{y}년{int(m)}월 {int(d)}일"
    except (ValueError, AttributeError):
        pass
    ws["F5"] = f"날짜: {date_label} (제{edition or 1}판)"

    _insert_logo(ws, height_cm=STD_LOGO_HEIGHT_CM)  # 기준서: 제목줄과 안 겹치게 높이 1.1cm로 축소(비율 유지)

    center = Alignment(horizontal="center", vertical="center")

    # 공유 템플릿에 예전 예시 자재(602506P005)의 값이 그대로 남아있으므로,
    # 실제 항목을 쓰기 전에 항목표 전체 범위(10~29행, A~T)를 먼저 비운다.
    for row in range(STD_FIRST_ROW, STD_LAST_ROW + 1):
        for col in ("A", "B", "E", "G"):
            ws[f"{col}{row}"] = None

    for idx, r in enumerate(results):
        row = STD_FIRST_ROW + idx
        if row > STD_LAST_ROW:
            break  # 기준서 양식이 지원하는 최대 20항목(A~T)을 넘으면 더 못 채움
        aql = r.get("aql")
        label_cell = ws[f"A{row}"]
        label_cell.value = item_label(r.get("item_name"), aql)
        label_cell.alignment = center

        ws[f"B{row}"] = r.get("spec_display") or ""

        method_cell = ws[f"E{row}"]
        method_cell.value = r.get("inspect_method") or ""
        method_cell.alignment = center

        aql_cell = ws[f"G{row}"]
        aql_cell.value = format_aql(aql)
        aql_cell.alignment = center


def build_report(material_no, product_name, header, results, overall,
                  approver=None, signature_path=None,
                  per_cycle_label=None, total_time_label=None, remarks=None, approval_type=None,
                  standard_info=None):
    """
    header: {"vendor":..,"po_no":..,"lot":..,"inspect_date":..,"inspector":..,"qty":..}
    results: judge.judge_all()의 결과 리스트
    approver: 승인자 이름 (텍스트로 기재)
    signature_path: 서명 이미지 파일 경로 (PNG). 있으면 검토란에 삽입.
    per_cycle_label / total_time_label: "N시간 N분 N초" 형식의 개당/총 측정 시간 문자열
    remarks: {"inspector":"이름: 코멘트", "manager":..., "approver":...} — 비고란에 역할별 색상으로 기재
    approval_type: 'special'이면 검사 결과 체크박스를 □특채로 표시(합/불 대신)
    standard_info: {"drawing_version":..,"revision_date":..,"edition":..,"unit":..} — 있으면
                    성적서 시트 뒤에 기준서(SAM 양식) 페이지를 이어붙임
    반환: (xlsx_path, pdf_path 또는 None, pdf_error 또는 None, signature_error 또는 None)
    """
    out_dir = report_output_dir()
    base_name = build_report_filename(header.get("vendor"), material_no, product_name)
    xlsx_path = os.path.join(out_dir, f"{base_name}.xlsx")
    xlsx_path = _dedupe_path(xlsx_path)
    shutil.copy(TEMPLATE, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[0]
    ws.title = material_no[:31]

    signature_error = _fill_sheet(
        ws, material_no, product_name, header, results, overall,
        approver=approver, signature_path=signature_path,
        per_cycle_label=per_cycle_label, total_time_label=total_time_label,
        remarks=remarks, approval_type=approval_type,
    )

    if standard_info is not None and os.path.exists(STANDARD_TEMPLATE):
        _append_standard_sheet(wb, material_no, product_name, standard_info, results)

    wb.save(xlsx_path)

    pdf_path, pdf_error = _to_pdf(xlsx_path, out_dir)
    return xlsx_path, pdf_path, pdf_error, signature_error


def _append_standard_sheet(wb, material_no, product_name, standard_info, results):
    """기준서 템플릿 시트를 성적서 워크북 뒤에 새 시트로 복사해와서 채운다."""
    std_wb = openpyxl.load_workbook(STANDARD_TEMPLATE)
    std_template_ws = std_wb.worksheets[0]

    new_ws = wb.create_sheet(title=f"기준서_{material_no}"[:31])
    # 셀 값·서식·병합·열너비까지 그대로 복제
    for row in std_template_ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()
    for mc in std_template_ws.merged_cells.ranges:
        new_ws.merge_cells(str(mc))
    for col_letter, dim in std_template_ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in std_template_ws.row_dimensions.items():
        new_ws.row_dimensions[row_idx].height = dim.height
    new_ws.page_setup.orientation = std_template_ws.page_setup.orientation
    new_ws.page_setup.paperSize = std_template_ws.page_setup.paperSize
    new_ws.page_setup.fitToWidth = 1
    new_ws.page_setup.fitToHeight = 1
    new_ws.sheet_properties.pageSetUpPr.fitToPage = True

    _fill_standard_sheet(
        new_ws, material_no, product_name,
        standard_info.get("drawing_version"), standard_info.get("revision_date"),
        standard_info.get("edition"), standard_info.get("unit"),
        results,
    )


def build_group_report(group_no, group_name, header, parts, overall,
                        approver=None, signature_path=None,
                        per_cycle_label=None, total_time_label=None, remarks=None, approval_type=None,
                        standard_info_map=None):
    """
    조립품(부품 여러 개로 분해해서 검사하는 자재) 전용 — 부품마다 시트를 하나씩 만들어
    통합 워크북 1개로 출력한다. PDF 변환 시 시트별로 페이지가 나뉜다.
    parts: [{"material_no":.., "product_name":.., "results":[...]}, ...] — 부품 등록 순서대로
    나머지 인자는 build_report와 동일. 서명 이미지는 마지막 시트에만 삽입한다.
    standard_info_map: {부품 material_no: {"drawing_version":..,"revision_date":..,"edition":..,"unit":..}}
                        — 있으면 각 부품 성적서 시트 바로 뒤에 그 부품의 기준서 페이지를 이어붙임
    반환: (xlsx_path, pdf_path 또는 None, pdf_error 또는 None, signature_error 또는 None)
    """
    out_dir = report_output_dir()
    base_name = build_report_filename(header.get("vendor"), group_no, group_name)
    xlsx_path = os.path.join(out_dir, f"{base_name}.xlsx")
    xlsx_path = _dedupe_path(xlsx_path)

    wb = openpyxl.load_workbook(TEMPLATE)
    template_ws = wb.worksheets[0]

    signature_error = None
    template_print_area = template_ws.print_area
    template_area_range = template_print_area.split("!")[-1] if "!" in (template_print_area or "") else template_print_area

    # 부품별 시트를 채우기 시작하기 "전에" 필요한 시트를 전부 미리 복제해둔다.
    # copy_worksheet를 루프 도중(idx>=1 시점)에 하면, 그때는 이미 template_ws(=idx 0의 시트)가
    # idx 0의 실제 값(항목기호 A열 등)으로 채워진 뒤라서 그 "오염된" 상태가 그대로 복제된다.
    # A열은 ITEM_COLS 클리어 대상이 아니라 값이 없는 행에 남아있던 값이 지워지지 않으므로,
    # 뒤 부품의 항목 수가 앞 부품보다 적으면 앞 부품의 항목기호가 뒤 부품 시트에 유령처럼 남는
    # 버그가 있었다 — 항상 깨끗한 template_ws에서 복제하도록 순서를 바꿈.
    sheets = [template_ws] + [wb.copy_worksheet(template_ws) for _ in range(len(parts) - 1)]

    for idx, part in enumerate(parts):
        ws = sheets[idx]
        if idx > 0:
            # openpyxl의 copy_worksheet는 인쇄영역·페이지설정(배율 등)을 복사하지 않으므로 직접 지정
            ws.page_setup.scale = template_ws.page_setup.scale
            ws.page_setup.orientation = template_ws.page_setup.orientation
            ws.page_setup.paperSize = template_ws.page_setup.paperSize
        ws.title = part["material_no"][:31]
        # 시트 이름이 바뀌었으니 인쇄영역도 새 이름 기준으로 다시 지정(openpyxl이 현재 제목으로 자동 접두)
        ws.print_area = template_area_range

        is_last = (idx == len(parts) - 1)
        err = _fill_sheet(
            ws, part["material_no"], part.get("product_name") or "", header,
            part["results"], overall,
            approver=approver, signature_path=(signature_path if is_last else None),
            per_cycle_label=per_cycle_label, total_time_label=total_time_label,
            remarks=(remarks if is_last else None), approval_type=approval_type,
        )
        if err:
            signature_error = err

        if standard_info_map and part["material_no"] in standard_info_map and os.path.exists(STANDARD_TEMPLATE):
            _append_standard_sheet(wb, part["material_no"], part.get("product_name") or "",
                                   standard_info_map[part["material_no"]], part["results"])

    wb.save(xlsx_path)

    pdf_path, pdf_error = _to_pdf(xlsx_path, out_dir)
    return xlsx_path, pdf_path, pdf_error, signature_error


def _to_pdf(xlsx_path, out_dir):
    """
    LibreOffice로 xlsx -> PDF 변환. (기존 코드는 xlsx 재계산용 변환을 한 번 더 호출했는데,
    LibreOffice headless를 연속 호출하면 프로필 잠금 충돌로 두 번째 호출이 조용히 실패하는
    경우가 있어 — PDF 변환 한 번만 수행하도록 정리함. PDF로 변환하는 과정에서
    LibreOffice가 수식도 함께 계산하므로 별도 재계산 단계는 불필요.)
    반환: (pdf_path 또는 None, 에러메시지 또는 None)
    """
    if SOFFICE is None:
        return None, "LibreOffice(soffice)를 찾을 수 없어. 설치돼 있는지, PATH에 잡혀 있는지 확인해줘."

    # 매 실행마다 독립된 사용자 프로필 사용 — 동시/연속 호출 시 프로필 잠금 충돌 방지
    profile_dir = tempfile.mkdtemp(prefix="lo_profile_")
    profile_uri = "file:///" + profile_dir.replace("\\", "/")

    try:
        proc = subprocess.run(
            [SOFFICE, "--headless", "--norestore",
             f"-env:UserInstallation={profile_uri}",
             "--convert-to", "pdf", "--outdir", out_dir, xlsx_path],
            capture_output=True, timeout=90
        )
    except Exception as e:
        return None, f"LibreOffice 실행 중 오류: {e}"
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)

    pdf_path = os.path.splitext(xlsx_path)[0] + ".pdf"
    if os.path.exists(pdf_path):
        return pdf_path, None

    stderr_msg = (proc.stderr or b"").decode("utf-8", errors="ignore").strip()
    stdout_msg = (proc.stdout or b"").decode("utf-8", errors="ignore").strip()
    detail = stderr_msg or stdout_msg or f"returncode={proc.returncode}"
    return None, f"PDF 변환 실패: {detail}"


def merge_report_with_drawing(report_pdf, drawing_pdf, output_path,
                               incl_drawing=True, incl_standard=True, incl_report=True):
    """도면→기준서→성적서 순으로 PDF 병합.

    report_pdf 구조: 성적서(첫 번째 페이지) + 기준서(나머지 페이지).
    - incl_drawing: 도면 PDF 앞에 붙임
    - incl_standard: report_pdf의 기준서 페이지(2페이지 이후) 포함
    - incl_report: report_pdf의 성적서 페이지(첫 페이지) 포함

    반환: (최종 PDF 경로, 에러메시지 or None)
    """
    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        return None, "pypdf 미설치 — 'pip install pypdf' 실행 필요"

    writer = PdfWriter()

    # ① 도면
    if incl_drawing and drawing_pdf and os.path.exists(drawing_pdf):
        try:
            for page in PdfReader(drawing_pdf).pages:
                writer.add_page(page)
        except Exception as e:
            return None, f"도면 PDF 읽기 실패: {e}"

    # ② 기준서 (report_pdf 2페이지 이후)
    report_pages = []
    if report_pdf and os.path.exists(report_pdf):
        try:
            report_pages = list(PdfReader(report_pdf).pages)
        except Exception as e:
            return None, f"성적서 PDF 읽기 실패: {e}"

    if incl_standard and len(report_pages) > 1:
        for page in report_pages[1:]:
            writer.add_page(page)

    # ③ 성적서 (report_pdf 첫 페이지)
    if incl_report and report_pages:
        writer.add_page(report_pages[0])

    if len(writer.pages) == 0:
        return None, "선택된 항목이 없어 PDF를 생성할 수 없어."

    with open(output_path, "wb") as f:
        writer.write(f)
    return output_path, None


def append_pdf(base_pdf, extra_pdf, output_path):
    """base_pdf 뒤에 extra_pdf를 붙여 output_path에 저장. 반환: (경로, 에러)"""
    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        return None, "pypdf 미설치"
    writer = PdfWriter()
    for p in PdfReader(base_pdf).pages:
        writer.add_page(p)
    for p in PdfReader(extra_pdf).pages:
        writer.add_page(p)
    with open(output_path, "wb") as f:
        writer.write(f)
    return output_path, None
