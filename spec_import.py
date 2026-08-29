# -*- coding: utf-8 -*-
"""
빈 성적서 양식(자재별로 항목/규격만 채워진 파일)을 여러 개 한 번에 업로드해서
규격표(specs 테이블)에 자동 등록하는 파서.

셀 위치를 고정으로 믿지 않고, 라벨 텍스트를 스캔해서 위치를 찾는다:
  - 자재번호/품명: 앞부분 몇 행을 스캔해서 "자재번호"/"품명" 텍스트가 있는 셀을 찾고
    같은 셀의 콜론 뒤 또는 바로 오른쪽 셀 값을 사용
  - 항목표: "검사 항목"과 "AQL"이 같이 있는 헤더 행을 찾고, 그 헤더 텍스트로
    번호/검사항목/AQL/검사방법 열 위치를 판단 (셀 위치가 파일마다 달라도 대응)
  - 한 파일 안에 시트가 여러 개면(자재마다 시트가 다름) 시트마다 각각 처리
"""
import re
import openpyxl

# 숫자 뒤에 붙는 단위(㎜, V, Ω, ㎋, ㎷, T ...) — 연산자·괄호가 아닌 아무 문자 1~3개로 본다.
# (Ω만 해도 U+03A9 / U+2126 두 종류가 데이터에 섞여 있어서 문자를 일일이 나열하는 건 못 믿는다)
_UNIT = r"(?:[^\d\s±~()%+\-]{1,3})?"
# 표기 맨 앞에 붙는 지름 기호 — "Ø44.9 - 0.1", "⌀133 + 0.1" 처럼 숫자 앞에 오는 경우
_DIA = r"(?:[Ø⌀ØΦφø∅]\s*)?"

# "108.5 ± 0.8" 같은 대칭 공차 (뒤에 %가 붙으면 기준값의 몇 %인지로 해석)
# 숫자와 ± 사이에 단위가 끼는 경우도 인식: "3T ± 0.15", "4Ω ±10%", "12mm ± 0.5"
_RE_PLUS_MINUS = re.compile(rf"([\d.]+)\s*{_UNIT}\s*±\s*([\d.]+)\s*(%?)")
# "63 - 0.2"(하한 방향) / "115 + 2"(상한 방향) 같은 한쪽 공차 — 전체 문자열이 이 패턴 하나여야 함
_RE_ONE_SIDED = re.compile(rf"^\s*{_DIA}([\d.]+)\s*([+-])\s*([\d.]+)\s*$")
# "(17.1~20.9)" 같은 괄호 안 명시적 범위 — 있으면 이걸 최우선으로 사용
# 숫자마다 단위가 붙는 "(3.00V~5.00V)", "(450㎷~650㎷)" 형태도 인식
_RE_RANGE = re.compile(rf"\(?\s*([\d.]+)\s*{_UNIT}\s*~\s*([\d.]+)\s*{_UNIT}\s*\)?")
# "72 (-0.05~-0.1)" / "109 (-0.2~0.1)" / "37 +0.1~0.05" 같은 경우 —
# 앞의 숫자(72)가 기준값이고 뒤쪽 두 숫자는 그 기준에서 뺄/더할 공차 오프셋.
# 괄호는 있어도 없어도 되고, 오프셋 중 "최소 하나"에 부호가 붙어 있으면 오프셋으로 본다
# (부호가 하나도 없으면 "(17.1~20.9)" 같은 절대범위와 구분이 안 되므로 아래 _RE_RANGE로 넘김).
#   ※ 예전엔 두 오프셋 "둘 다" 부호를 요구해서 "109 (-0.2~0.1)"이 이 패턴에 안 걸리고
#      _RE_RANGE에 잡혀 [0.1, 0.2] 로 저장됐다 → 실제 측정값 109가 항상 불합격되던 버그.
_RE_OFFSET_RANGE = re.compile(
    rf"^\s*{_DIA}(-?[\d.]+)\s*\(?\s*([+-]?[\d.]+)\s*~\s*([+-]?[\d.]+)\s*\)?\s*$"
)
# "5㎛ 이상" / "10 이상" 같은 최소값만 있는 단측 표기 — 상한 없음
_RE_AT_LEAST = re.compile(r"([\d.]+)\s*(?:[^\d\s]*)\s*이상")
# "10 이하" 같은 최대값만 있는 단측 표기 — 하한 없음
_RE_AT_MOST = re.compile(r"([\d.]+)\s*(?:[^\d\s]*)\s*이하")
# "4T" 같은 두께(Thickness) 단독 표기 — 자체 공차가 안 적혀 있으면 KS B ISO 2768-1
# 일반공차(표1/C급)를 자동 적용한다(_general_tolerance_for, 2026-08-30 사용자 확정 —
# 첨부 이미지 기준표 그대로).
_RE_THICKNESS_ONLY = re.compile(r"^\s*([\d.]+)\s*T\s*$")

# KS B ISO 2768-1 (표1 / C급) 일반 공차 — (구간 상한, 허용편차) 순서, 구간은 초과~이하.
# 첫 구간만 0.5 이상(그 미만 두께는 이 표의 적용 범위 밖이라 자동 계산하지 않는다).
_GENERAL_TOLERANCE_TABLE = [
    (3,    0.2),
    (6,    0.3),
    (30,   0.5),
    (120,  0.8),
    (400,  1.2),
    (1000, 2.0),
    (2000, 3.0),
    (4000, 4.0),
]


def _general_tolerance_for(nominal):
    """KS B ISO 2768-1 표1(C급) 기준, 치수 구간에 맞는 허용 편차(±)를 돌려준다.
    0.5 미만이거나 4000 초과면 이 표의 적용 범위 밖이라 None."""
    if nominal is None or nominal < 0.5 or nominal > 4000:
        return None
    for upper, tol in _GENERAL_TOLERANCE_TABLE:
        if nominal <= upper:
            return tol
    return None

MAX_LABEL_SEARCH_ROW = 12     # 자재번호/품명 라벨을 찾을 때 앞에서 몇 행까지 볼지
MAX_HEADER_SEARCH_ROW = 15    # 항목표 헤더 행을 찾을 때 앞에서 몇 행까지 볼지
MAX_DATA_ROWS = 40            # 항목표를 최대 몇 행까지 읽을지 (무한루프 방지)


def _clean(v):
    if v is None:
        return ""
    return v.replace("\n", "").strip() if isinstance(v, str) else str(v).strip()


def _extract_after_colon(text):
    """'자재번호：602106P103' -> '602106P103'. 전각/반각 콜론 둘 다 지원."""
    text = _clean(text)
    for sep in ("：", ":"):
        if sep in text:
            return text.split(sep, 1)[1].strip()
    return text


def _find_label_value(ws, label_keywords, max_row=MAX_LABEL_SEARCH_ROW, max_col=20):
    """
    앞부분 셀들을 스캔해서 label_keywords(예: ["자재번호"]) 중 하나라도 포함된 셀을 찾고,
    그 셀 텍스트에 콜론 뒤 값이 있으면 그 값을, 없으면 같은 행 바로 오른쪽 셀 값을 반환.
    못 찾으면 "".
    """
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if not v or not isinstance(v, str):
                continue
            text = _clean(v)
            if any(kw in text for kw in label_keywords):
                after = _extract_after_colon(text)
                if after and after != text:
                    return after
                # 콜론이 없는 라벨이면 오른쪽 셀 값을 시도
                right = ws.cell(row=r, column=c + 1).value
                if right:
                    return _clean(right)
    return ""


def _find_header_row(ws, max_row=MAX_HEADER_SEARCH_ROW, max_col=20):
    """
    "검사 항목"(또는 "항목")과 "AQL"이 같은 행에 있는 헤더 행을 찾는다.
    반환: (header_row_idx 또는 None, {col_idx: 정제된 헤더텍스트})
    """
    for r in range(1, max_row + 1):
        row_texts = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                row_texts[c] = _clean(v)
        texts = row_texts.values()
        has_item_header = any(("검사" in t and "항목" in t) or t == "항목" for t in texts)
        has_aql_header = any("AQL" in t.upper() for t in texts)
        if has_item_header and has_aql_header:
            return r, row_texts
    return None, {}


def _map_columns(header_texts):
    """헤더 텍스트 딕셔너리에서 번호/검사항목/AQL/검사방법 열 인덱스를 찾는다."""
    cols = {"no": None, "spec": None, "aql": None, "method": None}
    for c, t in header_texts.items():
        if t == "번호" and cols["no"] is None:
            cols["no"] = c
        elif ("검사" in t and "항목" in t) or t == "항목":
            cols["spec"] = c
        elif "AQL" in t.upper():
            cols["aql"] = c
        elif "검사" in t and "방법" in t:
            cols["method"] = c
    return cols


# 엑셀에서 넘어오는 유사 문자들 — 그대로 두면 숫자 패턴이 안 걸린다
_DASH_LOOKALIKES = "‐‑‒–—―−－"   # ‐‑‒–—―−－
_TILDE_LOOKALIKES = "～〜"                                        # ～〜
_PLUSMINUS_LOOKALIKES = "＋"                                          # ＋


def normalize_spec_text(text):
    """규격 표기 텍스트를 파싱하기 좋은 형태로 정규화한다.

    실제로 데이터에 섞여 있던 것들:
      - "282 – 0.5"  : 하이픈이 아니라 EN DASH(–) → 숫자 패턴이 안 걸려 규격 미인식
      - "_x000D_"    : 엑셀이 셀 안 줄바꿈을 저장할 때 남기는 잔재 문자열
      - 전각 물결(～)/전각 플러스(＋), 줄바꿈, 연속 공백
    """
    s = str(text or "")
    s = s.replace("_x000D_", " ").replace("_x000d_", " ")
    for ch in _DASH_LOOKALIKES:
        s = s.replace(ch, "-")
    for ch in _TILDE_LOOKALIKES:
        s = s.replace(ch, "~")
    for ch in _PLUSMINUS_LOOKALIKES:
        s = s.replace(ch, "+")
    return re.sub(r"\s+", " ", s).strip()


def _parse_tolerance(spec_display, is_visual):
    """
    규격 표기 텍스트에서 하한/상한을 뽑아낸다.
    반환: (lower, upper, needs_review: bool, review_note: str 또는 None)
    """
    if is_visual:
        return None, None, False, None

    text = normalize_spec_text(spec_display)

    m = _RE_OFFSET_RANGE.match(text)
    # 오프셋 중 최소 하나에 부호가 있어야 "기준값 + 공차"로 인정 (아니면 절대범위로 넘김)
    if m and (m.group(2).lstrip()[:1] in "+-" or m.group(3).lstrip()[:1] in "+-"):
        nominal, off1, off2 = float(m.group(1)), float(m.group(2)), float(m.group(3))
        lo, hi = nominal + min(off1, off2), nominal + max(off1, off2)
        return lo, hi, False, None

    m = _RE_RANGE.search(text)
    if m:
        lo, hi = float(m.group(1)), float(m.group(2))
        return min(lo, hi), max(lo, hi), False, None

    m = _RE_PLUS_MINUS.search(text)
    if m:
        nominal, tol = float(m.group(1)), float(m.group(2))
        if m.group(3) == "%":
            # "265 ±5%" → 265의 5% = 13.25 가 공차 (예전엔 %를 무시하고 ±5로 읽었음)
            tol = nominal * tol / 100.0
        return nominal - tol, nominal + tol, False, None

    m = _RE_ONE_SIDED.match(text)
    if m:
        nominal, sign, tol = float(m.group(1)), m.group(2), float(m.group(3))
        if sign == "+":
            return nominal, nominal + tol, False, None   # "115 + 2" → 115 ~ 117
        return nominal - tol, nominal, False, None        # "63 - 0.2" → 62.8 ~ 63

    m = _RE_AT_LEAST.search(text)
    if m:
        # "5㎛ 이상" — 최소값만 있고 상한은 없음(무제한). 자동 인식은 하되 하한만 채움
        return float(m.group(1)), None, False, None

    m = _RE_AT_MOST.search(text)
    if m:
        # "10 이하" — 최대값만 있고 하한은 없음(무제한)
        return None, float(m.group(1)), False, None

    m = _RE_THICKNESS_ONLY.match(text)
    if m:
        # "4T" — 두께(T) 단독 표기. 자체 공차가 없으면 KS B ISO 2768-1 일반공차(표1/C급)를
        # 치수 구간에 맞춰 자동 적용한다(사용자 확정, 2026-08-30).
        nominal = float(m.group(1))
        tol = _general_tolerance_for(nominal)
        if tol is not None:
            return nominal - tol, nominal + tol, False, None
        return None, None, True, "두께(T) 표기로 인식됐지만 일반공차표 적용 범위(0.5~4000) 밖이라 하한/상한을 직접 입력해야 해"

    return None, None, True, None


def _parse_sheet(ws, source_name):
    """
    시트 하나를 파싱한다.
    반환: (material_no, material_name, items, item_warnings, fail_reason)
    - 성공: fail_reason은 None, material_no/items가 채워짐 (item_warnings는 항목별 확인요청, 있을 수도 없을 수도)
    - 실패: fail_reason에 사람이 읽을 실패 사유 문자열, material_no는 ""일 수도 값이 있을 수도(자재번호는 찾았지만 항목표를 못 읽은 경우 등)
    """
    sheet_label = f"{source_name} / 시트 '{ws.title}'"

    material_no = _find_label_value(ws, ["자재번호", "자재 번호"])
    material_name = _find_label_value(ws, ["품명", "품명 및 규격"])

    if not material_no:
        return "", material_name, [], [], f"[{sheet_label}] 자재번호를 찾을 수 있는 셀이 없어."

    header_row, header_texts = _find_header_row(ws)
    if header_row is None:
        return material_no, material_name, [], [], \
            f"[{sheet_label} / {material_no}] '검사 항목'/'AQL' 헤더 행을 못 찾았어 — 셀 구조를 확인해줘."

    cols = _map_columns(header_texts)
    if cols["spec"] is None or cols["aql"] is None:
        return material_no, material_name, [], [], \
            f"[{sheet_label} / {material_no}] 헤더는 찾았지만 '검사항목' 또는 'AQL' 열 위치를 특정 못 했어."

    items = []
    item_warnings = []
    order = 1
    empty_streak = 0
    for r in range(header_row + 1, header_row + 1 + MAX_DATA_ROWS):
        spec_val = ws.cell(row=r, column=cols["spec"]).value
        spec_display = _clean(spec_val)
        if not spec_display:
            empty_streak += 1
            if empty_streak >= 2:   # 빈 행이 연속 2개면 표가 끝난 것으로 판단
                break
            continue
        empty_streak = 0

        item_name = _clean(ws.cell(row=r, column=cols["no"]).value) if cols["no"] else str(order)
        aql = ws.cell(row=r, column=cols["aql"]).value
        inspect_method = _clean(ws.cell(row=r, column=cols["method"]).value) if cols["method"] else ""

        is_visual = ("외관" in spec_display) or (inspect_method == "육안")
        judge_type = "visual" if is_visual else "numeric"

        lower, upper, needs_review, review_note = _parse_tolerance(spec_display, is_visual)
        if needs_review:
            reason = review_note or f"규격 표기 '{spec_display}'에서 하한/상한을 자동으로 못 읽었어"
            item_warnings.append(
                f"[{sheet_label} / {material_no} / 항목 {item_name}] "
                f"{reason} — 등록은 됐지만 규격 상세 화면에서 직접 확인·수정해줘."
            )

        items.append({
            "item_name": item_name,
            "spec_display": spec_display,
            "judge_type": judge_type,
            "lower_limit": lower,
            "upper_limit": upper,
            "inspect_method": inspect_method,
            "aql": aql,
            "item_order": order,
        })
        order += 1

    if not items:
        return material_no, material_name, [], [], \
            f"[{sheet_label} / {material_no}] 헤더는 찾았는데 항목 데이터가 하나도 없어."

    return material_no, material_name, items, item_warnings, None


def parse_spec_file(file_stream_or_path, source_name=""):
    """
    xlsx 파일 하나(시트 여러 개 가능)를 읽어서 시트별 결과 리스트를 반환.
    반환: [(material_no, material_name, items, item_warnings, fail_reason), ...] — 시트 개수만큼.
    fail_reason이 None이면 성공(등록 대상), 아니면 실패(등록 대상 아님).
    """
    wb = openpyxl.load_workbook(file_stream_or_path, data_only=True)
    results = []
    for ws in wb.worksheets:
        results.append(_parse_sheet(ws, source_name))
    return results
