# -*- coding: utf-8 -*-
# Copyright (c) 2026 윤주호. All rights reserved.
# 무단 복제·배포·수정을 금합니다.
"""IQC 성적서 시스템 — Flask 골격 (2단계: DB + 판정 로직 + 기본 화면)"""

# ──────────────────────────────────────────────
# 라이선스 체크 — 절대 이 블록을 제거하거나 수정하지 말 것
# Copyright (c) 2026 윤주호. 무단 사용 금지.
# ──────────────────────────────────────────────
def _check_license():
    import urllib.request as _ur
    import datetime as _d
    import sys as _sys

    # ① 만료일 체크 (이 날짜 이후 자동 종료)
    expiry = _d.date(2027, 12, 31)
    if _d.date.today() > expiry:
        print("\n[오류] 소프트웨어 사용 기간이 만료되었습니다. 개발자(윤주호)에게 문의하세요.\n")
        _sys.exit(1)

    # ② 원격 활성화 체크 (개발자가 원격으로 즉시 비활성화 가능)
    _REMOTE_URL = "https://raw.githubusercontent.com/thebrilliantpillar-beep/QC-SYSTEM/main/status.txt"
    try:
        res = _ur.urlopen(_REMOTE_URL, timeout=5)
        status = res.read().strip()
        if status != b"OK":
            print("\n[오류] 라이선스가 비활성화되었습니다. 개발자(윤주호)에게 문의하세요.\n")
            _sys.exit(1)
    except Exception:
        # 인터넷 연결 없으면 원격 체크 건너뜀 (NAS 내부망 환경 고려)
        pass

_check_license()
# ──────────────────────────────────────────────

import os, base64, io, time, shutil, re, zipfile, tempfile
from datetime import datetime as _dt
from functools import wraps
from datetime import timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, session, g, send_file, send_from_directory, jsonify
import database as db
import report_builder
import spec_import as spec_import_module

app = Flask(__name__)

# ──────────────────────────────────────────────
# 유틸리티 함수
# ──────────────────────────────────────────────
# AQL 표기('퍼센트10' → '10%')는 성적서·기준서와 반드시 같아야 하므로 report_builder 것을 그대로 쓴다
format_aql_display = report_builder.format_aql

_WEEKDAYS_KO = ["월", "화", "수", "목", "금", "토", "일"]


# (형식, 그 형식이 차지하는 글자 수) — 자릿수로 구분해야 오해석을 막을 수 있다
_DATE_FORMATS = (
    ("%Y-%m-%d", 10), ("%Y/%m/%d", 10), ("%Y.%m.%d", 10),
    ("%y-%m-%d", 8),  ("%y/%m/%d", 8),  ("%y.%m.%d", 8),
    ("%Y%m%d", 8),    ("%y%m%d", 6),
)


def _parse_any_date(text):
    """어떤 형식으로 들어왔든 날짜 부분만 뽑아낸다. 못 읽으면 None.

    구분자 없는 숫자 형식은 **길이가 정확히 맞을 때만** 인정한다.
    안 그러면 '260821'(=2026-08-21)을 %Y%m%d로 읽어서 2608년 2월 1일이 돼버린다.
    """
    text = str(text or "").strip()
    if not text:
        return None
    head = text.split()[0]          # '2026-08-23 14:30' 처럼 시각이 붙어 있으면 날짜만
    for fmt, width in _DATE_FORMATS:
        if len(head) < width:
            continue
        if fmt in ("%Y%m%d", "%y%m%d") and len(head) != width:
            continue
        try:
            return _dt.strptime(head[:width], fmt).date()
        except ValueError:
            continue
    return None


def format_date_korean(date_str):
    """검사자가 어떤 형식으로 넣었든 화면 표기는 '2026-08-23 (일)' 로 통일한다.
    해석 못 하는 값은 원문 그대로 돌려준다(값을 잃지 않기 위해)."""
    d = _parse_any_date(date_str)
    if d is None:
        return str(date_str or "").strip()
    return f"{d:%Y-%m-%d} ({_WEEKDAYS_KO[d.weekday()]})"


def format_datetime_korean(value):
    """승인일시처럼 시각이 같이 있는 값 → '2026-08-23 (일) 14:30'.
    시각이 없으면 날짜만 돌려준다."""
    text = str(value or "").strip()
    d = _parse_any_date(text)
    if d is None:
        return text
    out = f"{d:%Y-%m-%d} ({_WEEKDAYS_KO[d.weekday()]})"
    m = re.search(r"(\d{1,2}):(\d{2})", text)
    if m:
        out += f" {int(m.group(1)):02d}:{m.group(2)}"
    return out


# Jinja2 필터 등록
app.jinja_env.filters['aql_display'] = format_aql_display
app.jinja_env.filters['date_korean'] = format_date_korean
app.jinja_env.filters['datetime_korean'] = format_datetime_korean
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-later")
app.permanent_session_lifetime = timedelta(hours=24)  # 하루 한 번 로그인하면 그 뒤로 계속 유지 (admin 제외)

SIGNATURE_DIR = os.path.join(db.DATA_DIR, "signatures")
os.makedirs(SIGNATURE_DIR, exist_ok=True)

# 도면 PDF 폴더 — 경로 변경 시 이 한 줄만 수정하면 됨
# 실 데이터(운영자가 채워넣는 파일)라 서명/NCR사진과 동일하게 DATA_DIR 기준으로 둔다.
# 로컬에선 DATA_DIR이 기본적으로 앱 폴더라 기존 경로(자동출력/도면)와 동일하게 유지된다.
DRAWING_DIR = os.path.join(db.DATA_DIR, "자동출력", "도면")
os.makedirs(DRAWING_DIR, exist_ok=True)

BACKUP_DIR = os.path.join(db.DATA_DIR, "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)
BACKUP_README = os.path.join(BACKUP_DIR, "README.txt")


@app.route("/static/signatures/<path:filename>")
def signature_file(filename):
    # 서명 이미지는 static/ 안이 아니라 DATA_DIR(영구 디스크)에 저장되므로
    # Flask 기본 static 라우트 대신 이 라우트가 대신 서빙한다 (URL은 그대로 유지)
    return send_from_directory(SIGNATURE_DIR, filename)


@app.route("/static/ncr_photos/<path:filename>")
def ncr_photo_file(filename):
    return send_from_directory(NCR_PHOTO_DIR, filename)

DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin1234"
ADMIN_IDLE_TIMEOUT_SEC = 600  # "users" 권한 보유 계정만 10분 무동작 시 자동 로그아웃

# 개별 권한 종류 — admin이 계정마다 하나하나 부여/회수
# 권한을 기능별로 완전 세분화 — 계정 상세페이지에서 그룹 단위로 체크해서 부여한다.
PERM_GROUPS = [
    ("입고", [
        ("intake", "입고 리스트 등록·관리"),
    ]),
    ("검사", [
        ("inspect_input",    "검사 입력"),
        ("inspect_edit_all", "타인이 입력한 성적서도 수정"),
        ("inspect_history",  "검사 이력 열람"),
        ("history_delete",   "검사 이력 삭제"),
        ("defect_history",   "불량 이력 열람"),
        ("ncr",              "부적합 통보서 작성"),
        ("ncr_confirm",      "부적합 통보서 확인·발송"),
        ("return",           "반품 처리"),
    ]),
    ("승인", [
        ("approve",        "승인·반려·특채·불합격 확정"),
        ("approve_revoke", "결정 회수"),
    ]),
    ("출력", [
        ("output", "성적서 출력"),
        ("custom_template", "커스텀 성적서 양식 제작·지정"),
    ]),
    ("자재", [
        ("material_view",   "자재 열람"),
        ("material_edit",   "자재 등록·수정·삭제(개별등록 포함)"),
        ("material_import", "자재 일괄 등록"),
    ]),
    ("마스터", [
        ("gauge",    "계측기 관리"),
        ("supplier", "업체 관리"),
    ]),
    ("관리", [
        ("users", "계정 관리 (10분 자동로그아웃 대상)"),
        ("smtp",  "이메일(SMTP) 설정"),
        ("logs",  "활동 로그 열람"),
    ]),
]
PERM_LABELS = {code: label for _grp, items in PERM_GROUPS for code, label in items}
ALL_PERMS = list(PERM_LABELS.keys())

# 예전 8개 권한 → 세분화 권한 매핑 (최초 1회 자동 마이그레이션)
PERM_MIGRATION = {
    "intake":      ["intake"],
    "spec":        ["material_view", "material_edit", "material_import", "gauge", "supplier"],
    "inspect":     ["inspect_input", "inspect_history", "defect_history", "ncr"],
    "inspect_all": ["inspect_edit_all"],
    "approve":     ["approve", "approve_revoke", "ncr_confirm", "return"],
    "output":      ["output"],
    "users":       ["users", "smtp", "history_delete"],
    "logs":        ["logs"],
}


def ensure_perm_migration():
    """계정 권한을 옛 8개 체계에서 세분화 체계로 1회 변환한다(버전 플래그로 멱등 보장)."""
    if db.get_setting("perm_schema_version", "1") == "2":
        return
    for u in db.list_users():
        old_tokens = [p for p in (u["permissions"] or "").split(",") if p]
        merged, seen = [], set()
        for tok in old_tokens:
            for np in PERM_MIGRATION.get(tok, [tok]):
                if np in ALL_PERMS and np not in seen:
                    seen.add(np)
                    merged.append(np)
        db.update_user_permissions(u["id"], ",".join(merged))
    db.set_setting("perm_schema_version", "2")


def ensure_inspect_method_fill_20260825():
    """검사방식(inspect_method) 미입력 항목 일괄 정리 (2026-08-25, 1회 실행, 버전 플래그로 멱등 보장).
    숫자측정 항목이 아닌(육안으로 판단 가능한) 미입력 항목은 전부 '육안'으로 채우고,
    600006P276의 *A 항목(치수 75±0.8)만 예외로 '버니어캘리퍼스'를 채운다 — 로컬 DB에서
    사용자가 직접 확인하고 승인한 규칙을 그대로 운영 DB에도 적용."""
    if db.get_setting("inspect_method_fill_20260825", "0") == "1":
        return
    conn = db.get_conn()
    conn.execute("""
        UPDATE specs SET inspect_method='버니어캘리퍼스'
        WHERE material_no='600006P276' AND item_name='*A'
          AND (inspect_method IS NULL OR TRIM(inspect_method)='')
    """)
    cur = conn.execute("""
        UPDATE specs SET inspect_method='육안'
        WHERE judge_type IN ('ok_ng', 'visual')
          AND (inspect_method IS NULL OR TRIM(inspect_method)='')
    """)
    filled = cur.rowcount
    conn.commit()
    conn.close()
    db.set_setting("inspect_method_fill_20260825", "1")
    db.log_activity(None, "system", "system", "검사방식 미입력 일괄 수정 (배포 마이그레이션)",
                     "specs", None, f"{filled}건을 육안으로 자동 설정 (numeric 미입력 항목은 개별 확인 후 수동 처리)")


def record_change(action, target_type=None, target_id=None, detail=None):
    """
    데이터가 새로 생기거나 수정될 때마다 호출:
    1) activity_log 테이블에 기록 (users 권한 보유자만 /logs에서 조회 가능)
    2) iqc.db 전체를 backups/ 밑에 타임스탬프 붙여 복사
    3) backups/README.txt에 사람이 읽을 수 있는 상세 로그 한 줄 추가
    """
    user_id = g.user["id"] if g.user else None
    username = g.user["username"] if g.user else "system"
    perms_summary = (g.user["permissions"] if g.user else None) or ""

    log_id = db.log_activity(user_id, username, perms_summary, action, target_type, target_id, detail)

    now = _dt.now()
    ts = now.strftime("%Y%m%d_%H%M%S")
    backup_name = f"iqc_{ts}_{log_id}.db"
    backup_path = os.path.join(BACKUP_DIR, backup_name)
    try:
        shutil.copy(db.DB_PATH, backup_path)
    except Exception:
        backup_name = "(백업 실패)"

    line = (f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] "
            f"사용자: {username} | "
            f"액션: {action} | 대상: {target_type or '-'} {target_id or ''} | "
            f"상세: {detail or '-'} | 백업파일: {backup_name}\n")
    try:
        with open(BACKUP_README, "a", encoding="utf-8") as f:
            f.write(line)
    except Exception:
        pass

    prune_backups()


# 백업 보관 정책 — 변경이 잦아서 그냥 두면 폴더가 무한정 커진다(수동 정리는 안 하게 됨).
BACKUP_KEEP_RECENT = 30    # 최근 N개는 무조건 보관
BACKUP_KEEP_DAILY_DAYS = 90  # 그보다 오래된 건 하루에 1개(그날 마지막 것)만 남기고, 이 기간이 지나면 삭제


def prune_backups():
    """백업 폴더 자동 정리:
      - 최근 30개는 그대로 둔다
      - 그 이전 것은 '하루 1개'만 남긴다 (그날 가장 마지막 백업)
      - 90일보다 오래된 것은 지운다
    파일명 규칙 iqc_YYYYMMDD_HHMMSS_{log_id}.db 를 전제로 한다.
    """
    try:
        files = []
        for name in os.listdir(BACKUP_DIR):
            if not (name.startswith("iqc_") and name.endswith(".db")):
                continue   # 사람이 손으로 만든 백업(iqc_before_... 등)은 규칙이 달라도 아래에서 걸러짐
            parts = name[:-3].split("_")
            if len(parts) < 3 or len(parts[1]) != 8 or not parts[1].isdigit():
                continue   # 날짜 형식이 아니면 자동 정리 대상에서 제외 (수동 백업 보호)
            files.append((parts[1], name))

        if len(files) <= BACKUP_KEEP_RECENT:
            return

        files.sort(key=lambda x: x[1], reverse=True)   # 파일명이 곧 시간순
        recent = files[:BACKUP_KEEP_RECENT]
        older = files[BACKUP_KEEP_RECENT:]

        cutoff = (_dt.now() - timedelta(days=BACKUP_KEEP_DAILY_DAYS)).strftime("%Y%m%d")
        keep_per_day = {}
        for day, name in older:
            if day < cutoff:
                continue                       # 너무 오래됨 → 보관 대상 아님
            keep_per_day.setdefault(day, name)  # 정렬이 최신순이라 첫 번째가 그날 마지막 백업

        keep = {n for _, n in recent} | set(keep_per_day.values())
        for _, name in older:
            if name in keep:
                continue
            try:
                os.remove(os.path.join(BACKUP_DIR, name))
            except OSError:
                pass
    except Exception:
        pass   # 백업 정리 실패가 본 기능을 막으면 안 됨


@app.context_processor
def inject_perm_labels():
    user_perms = _user_perms(g.user) if hasattr(g, "user") else set()
    # 네비게이션 배지 카운트
    nav_counts = {}
    if hasattr(g, "user") and g.user:
        nav_counts["inspect"] = len(db.list_intake(status="대기"))
        nav_counts["approve"] = len(db.list_inspections(status="pending"))
        nav_counts["output"]  = len(db.list_pending_output_inspections())
        # 소분류 뱃지
        try:
            followup = db.get_defect_followup()
            # "부적합 통보서" 배지 = /ncr 목록에서 실제로 액션이 필요한 것만.
            #   - ncr_review(draft, 확인 대기)만 여기 넣는다
            #   - ncr_write(작성 필요)는 아직 NCR 레코드가 안 만들어진 상태라 /ncr 목록엔 안 뜬다
            #     → 배지에 넣으면 클릭했을 때 목록이 비어서 사용자가 혼란스러워짐
            #     이 케이스는 아래 defect_followup(불량 이력) 배지에 이미 포함돼 있어서
            #     검사자가 성적서 상세로 들어가서 「통보서 작성」 버튼을 누르는 흐름으로 처리한다.
            nav_counts["ncr_draft"] = len(followup["ncr_review"])
            # "불량 이력" 배지 = 후속조치 전체 대기 건수(재검사 대기 + 통보서 작성/확인/발송 대기)
            nav_counts["defect_followup"] = (len(followup["recheck"]) + len(followup["ncr_write"])
                                              + len(followup["ncr_review"]) + len(followup["ncr_send"]))
        except Exception:
            nav_counts["ncr_draft"] = 0
            nav_counts["defect_followup"] = 0
        try:
            nav_counts["return"] = len([r for r in db.list_returns() if (r["status"] or "") in ("draft", "pending")])
        except Exception:
            nav_counts["return"] = 0
    # admin(마스터 계정) 여부 — 템플릿에서 관리자 전용 UI 노출 조건으로 사용
    is_admin = bool(g.user) and (g.user["username"] or "").strip().lower() == "admin"
    return {"perm_labels": PERM_LABELS, "user_perms": user_perms,
            "status_display": status_display, "sample_size": sample_size,
            "nav_counts": nav_counts, "is_admin": is_admin}


def _admin_only():
    """admin 계정 전용 라우트 가드. admin이 아니면 홈으로 되돌린다."""
    if not g.user or (g.user["username"] or "").strip().lower() != "admin":
        flash("이 작업은 admin 계정만 할 수 있어.")
        return redirect(url_for("home"))
    return None


def _approval_status_label(status, overall_result, approval_type):
    """status_display()와 같은 규칙을 raw 값(row 객체 아님)으로 받아 라벨만 돌려준다 —
    검사이력 외에도 NCR·반품처럼 원본 성적서를 조인해서 쓰는 목록의 검색 필터에서
    재사용하기 위해 분리."""
    if status == "pending":
        if overall_result == "검토필요":
            return "검토필요"
        return "대기중"
    if status == "rejected":
        return "반려"
    if status == "superseded":
        return "재검사 진행됨"
    if status == "approved":
        if approval_type == "special":
            return "특채 승인"
        if approval_type == "failed":
            return "불합격 확정"
        return "합격 승인"
    return status or ""


APPROVAL_STATUS_LABELS = ["대기중", "검토필요", "반려", "재검사 진행됨", "특채 승인", "불합격 확정", "합격 승인"]
OVERALL_RESULT_OPTIONS = ["합격", "검토필요", "불합격", "규격미입력"]


def status_display(insp):
    """
    성적서 한 건의 (한글 상태 라벨, badge css 클래스) 반환.
    승인된 건은 합격/불합격/특채 여부까지 구분해서 보여준다.
    """
    label = _approval_status_label(insp["status"], insp["overall_result"], insp["approval_type"])
    cls_map = {"대기중": "pending", "검토필요": "review", "반려": "fail", "재검사 진행됨": "muted",
               "특채 승인": "special", "불합격 확정": "fail", "합격 승인": "pass"}
    return (label, cls_map.get(label, "pending"))


PAGE_SIZE = 50


def _paginate(items, page_arg="page", per_page=PAGE_SIZE):
    """긴 목록을 페이지 단위로 자른다. 화면(_pager.html)이 요구하는 dict 하나로 통일해서
    돌려준다 — 각 목록 라우트마다 페이지 계산 로직을 복붙하지 않게(2026-09-02).

    쿼리스트링의 ?page=N을 읽고, 범위를 벗어나면 1~last로 클램프한다. items가 리스트가
    아니어도(제너레이터 등) 안전하게 처리하려면 호출부에서 list()로 감싸서 넘길 것."""
    total = len(items)
    last = max(1, -(-total // per_page))
    try:
        page = int(request.args.get(page_arg, "1"))
    except ValueError:
        page = 1
    page = max(1, min(page, last))
    start = (page - 1) * per_page
    end = start + per_page
    return {
        "items": items[start:end],
        "page": page,
        "last": last,
        "total": total,
        "per_page": per_page,
        "has_prev": page > 1,
        "has_next": page < last,
        "page_arg": page_arg,
        "start_index": start + 1 if total else 0,
        "end_index": min(end, total),
    }


def _list_search_params():
    """4개 목록 화면(검사이력/불량이력/부적합통보서/반품처리)이 공통으로 쓰는 검색 조건을
    쿼리스트링에서 뽑아 dict로 정리."""
    a = request.args
    return {
        "q_inspector": (a.get("q_inspector") or "").strip(),
        "q_supplier": (a.get("q_supplier") or "").strip(),
        "q_product": (a.get("q_product") or "").strip(),
        "q_material": (a.get("q_material") or "").strip(),
        "f_result": (a.get("f_result") or "").strip(),
        "f_status": (a.get("f_status") or "").strip(),
        "insp_start": (a.get("insp_start") or "").strip(),
        "insp_end": (a.get("insp_end") or "").strip(),
        "recv_start": (a.get("recv_start") or "").strip(),
        "recv_end": (a.get("recv_end") or "").strip(),
    }


def _row_passes_search(f, inspector=None, supplier=None, product=None, material=None,
                        result=None, status=None, insp_date=None, recv_date=None):
    """공통 검색 필터 한 건 판정. 필드가 그 화면에 아예 없으면(None) 그 조건은 건너뛰고,
    있는데 값이 비어 있으면(빈 문자열) 정상적으로 걸러진다."""
    if f["q_inspector"] and inspector is not None and f["q_inspector"] not in inspector:
        return False
    if f["q_supplier"] and supplier is not None and f["q_supplier"] not in supplier:
        return False
    if f["q_product"] and product is not None and f["q_product"] not in product:
        return False
    if f["q_material"] and material is not None and f["q_material"] not in material:
        return False
    if f["f_result"] and result is not None and f["f_result"] != result:
        return False
    if f["f_status"] and status is not None and f["f_status"] != status:
        return False

    if (f["insp_start"] or f["insp_end"]) and insp_date is not None:
        d = _parse_any_date(insp_date)
        if not d:
            return False
        if f["insp_start"]:
            sd = _parse_any_date(f["insp_start"])
            if sd and d < sd:
                return False
        if f["insp_end"]:
            ed = _parse_any_date(f["insp_end"])
            if ed and d > ed:
                return False

    if (f["recv_start"] or f["recv_end"]) and recv_date is not None:
        d = _parse_any_date(recv_date)
        if not d:
            return False
        if f["recv_start"]:
            sd = _parse_any_date(f["recv_start"])
            if sd and d < sd:
                return False
        if f["recv_end"]:
            ed = _parse_any_date(f["recv_end"])
            if ed and d > ed:
                return False

    return True


def ensure_default_admin():
    """계정이 하나도 없으면(최초 실행) 전체 권한을 가진 기본 계정을 하나 만들어둔다."""
    if db.count_users() == 0:
        db.create_user(DEFAULT_ADMIN_USERNAME, DEFAULT_ADMIN_PASSWORD, "관리자", ",".join(ALL_PERMS))


# ---------- 로그인/권한 ----------

def _user_perms(user_row):
    if user_row is None:
        return set()
    # admin 계정(username='admin')은 모든 권한을 자동 보유.
    # 어느 리스트 페이지에서든 수정·삭제·회수·재발행 등을 할 수 있도록 하는 마스터 계정.
    # 활동로그는 아예 삭제 라우트/UI가 없으므로(감사기록 보호) 여기서도 자동으로 지켜진다.
    if (user_row["username"] or "").strip().lower() == "admin":
        return set(ALL_PERMS)
    return set(p for p in (user_row["permissions"] or "").split(",") if p)


@app.before_request
def load_logged_in_user():
    user_id = session.get("user_id")
    g.user = db.get_user(user_id) if user_id else None

    if g.user is not None and "users" in _user_perms(g.user):
        # "users"(계정관리) 권한 보유 계정만 10분 무동작 시 자동 로그아웃
        last_seen = session.get("last_seen")
        now = time.time()
        if last_seen and (now - last_seen) > ADMIN_IDLE_TIMEOUT_SEC:
            session.clear()
            g.user = None
        else:
            session["last_seen"] = now


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def perm_required(*required_perms):
    """나열된 권한 중 하나라도 있으면 통과. 아무것도 안 넘기면(perm_required()) 'users' 권한 보유자만 통과."""
    checks = required_perms or ("users",)

    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if g.user is None:
                return redirect(url_for("login", next=request.path))
            user_perms = _user_perms(g.user)
            if not any(p in user_perms for p in checks):
                flash("이 화면에 접근할 권한이 없어.")
                return redirect(url_for("home"))
            return view(*args, **kwargs)
        return wrapped
    return decorator


# 이전 role_required 이름을 쓰던 코드 호환용 별칭 (아래에서 전부 perm_required로 교체할 예정)
role_required = perm_required
admin_required = perm_required()  # "users" 권한 보유자만


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET" and g.user is not None:
        # 이미 로그인된 상태로 /login에 들어오면(북마크, 뒤로가기 등) 로그인 폼 위에
        # 상단 메뉴까지 같이 뜨는 이상한 화면이 됐다 — 바로 홈으로 보낸다.
        return redirect(url_for("home"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = db.get_user_by_username(username)
        if user is None or user["password"] != password:
            flash("아이디 또는 비밀번호가 틀렸어.")
            return redirect(url_for("login"))

        session.permanent = True
        session["user_id"] = user["id"]
        session["last_seen"] = time.time()
        next_url = request.args.get("next") or url_for("home")
        return redirect(next_url)

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/session/ping", methods=["POST"])
def session_ping():
    """디자이너처럼 서버 요청 없이 오래 머무는 화면이 세션을 유지·확인하는 용도.
    로그인 화면(HTML)으로 리다이렉트하지 않고 항상 JSON을 돌려준다 →
    화면이 '응답 해석 실패' 대신 만료를 정확히 알 수 있다.
    before_request가 이미 last_seen을 갱신(또는 만료 시 session.clear)했으므로
    여기서는 그 결과만 알려준다."""
    from flask import jsonify
    if g.user is None:
        return jsonify({"ok": False, "expired": True}), 401
    return jsonify({"ok": True})


@app.route("/my-password", methods=["GET", "POST"])
@login_required
def my_password():
    """본인 비밀번호 직접 변경 (임시 계정 발급 후 본인이 바꾸는 용도)."""
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "").strip()
        if g.user["password"] != current:
            flash("현재 비밀번호가 틀렸어.")
        elif not new_password:
            flash("새 비밀번호를 입력해줘.")
        else:
            db.update_user_password(g.user["id"], new_password)
            flash("비밀번호가 변경됐어.")
            record_change("본인 비밀번호 변경", "user", g.user["username"])
        return redirect(url_for("my_password"))

    return render_template("my_password.html")


# ---------- DB 백업 다운로드 ----------

@app.route("/download-db")
@perm_required("users")
def download_db():
    today = _dt.now().strftime("%Y%m%d_%H%M%S")
    filename = f"iqc_backup_{today}.db"
    record_change("DB 백업 다운로드", "system", 0, f"다운로드: {filename}")
    return send_file(db.DB_PATH, as_attachment=True, download_name=filename)


# ---------- 데이터 정정(관리자 전용) ----------

@app.route("/admin/inspection-item/<int:item_id>/rename", methods=["POST"])
@perm_required("users")
def admin_rename_inspection_item(item_id):
    """항목기호 중복 등록 사고 정정용 — inspection_items 한 행의 item_name만 고친다.
    정상 사용 흐름에는 없는 버튼이라 UI 없이 직접 호출하는 관리자 전용 라우트."""
    expected_old = (request.form.get("expected_old") or "").strip()
    new_name = (request.form.get("new_name") or "").strip()
    if not new_name:
        return jsonify({"ok": False, "error": "new_name이 비어 있어."}), 400
    conn = db.get_conn()
    row = conn.execute("SELECT * FROM inspection_items WHERE id=?", (item_id,)).fetchone()
    conn.close()
    if row is None:
        return jsonify({"ok": False, "error": "해당 항목을 찾을 수 없어."}), 404
    if expected_old and row["item_name"] != expected_old:
        return jsonify({"ok": False, "error": f"현재 item_name이 예상과 달라: {row['item_name']}"}), 409
    db.rename_inspection_item(item_id, new_name)
    record_change("성적서 항목기호 정정", "inspection_item", item_id,
                  f"inspection {row['inspection_id']} / {row['item_name']} → {new_name} (항목기호 중복 사고 수습)")
    return jsonify({"ok": True})


@app.route("/admin/sync-material-names", methods=["POST"])
@perm_required("users")
def admin_sync_material_names():
    """specs.material_name(등록 당시 복사본)이 materials.material_name(정본)과
    어긋나 있던 데이터를 정리 — 검사이력/승인 화면에 자재명이 안 보이던 사고 수습.
    자세한 설명은 db.sync_material_names_from_master() 참고."""
    result = db.sync_material_names_from_master()
    record_change("자재명 동기화 정정", "system", 0,
                  f"specs {result['specs_fixed']}건, inspections {result['inspections_fixed']}건 (정본 기준 정정)")
    return jsonify({"ok": True, **result})


# ---------- 도면 ----------

def find_drawing_pdf(material_no):
    """자재번호에 맞는 도면 PDF 경로 반환. 없으면 None.
    우선순위: ① materials.drawing_file 수동 지정 → ② 자재번호.pdf 자동 매칭"""
    material = db.get_material(material_no)
    if material and material["drawing_file"]:
        p = os.path.join(DRAWING_DIR, material["drawing_file"])
        if os.path.exists(p):
            return p

    # 자동 매칭: 자재번호.pdf
    auto = os.path.join(DRAWING_DIR, f"{material_no}.pdf")
    if os.path.exists(auto):
        return auto
    return None


def materials_with_drawings(material_nos):
    """주어진 자재번호들 중 도면 파일이 실제로 있는 것만 집합으로 반환.
    목록 화면에서 도면보기 버튼을 도면 있는 행에만 보여주기 위해 씀."""
    material_nos = set(material_nos)
    if not material_nos:
        return set()
    try:
        existing_files = set(os.listdir(DRAWING_DIR))
    except OSError:
        existing_files = set()
    result = set()
    for mno in material_nos:
        material = db.get_material(mno)
        if material and material["drawing_file"] and material["drawing_file"] in existing_files:
            result.add(mno)
        elif f"{mno}.pdf" in existing_files:
            result.add(mno)
    return result


@app.route("/drawing/<material_no>")
@perm_required("inspect_input", "inspect_history", "approve", "output", "material_view")
def serve_drawing(material_no):
    """도면 PDF를 팝업 iframe에서 볼 수 있도록 서빙."""
    path = find_drawing_pdf(material_no)
    if path is None:
        return "도면 파일을 찾을 수 없어.", 404
    return send_file(path, mimetype="application/pdf")


@app.route("/drawing-list")
@perm_required("material_view")
def drawing_list_json():
    """도면 폴더 안의 PDF 파일 목록 반환 (도면 지정 드롭다운용)."""
    import json
    if not os.path.isdir(DRAWING_DIR):
        return json.dumps([])
    files = sorted(f for f in os.listdir(DRAWING_DIR) if f.lower().endswith(".pdf"))
    return app.response_class(json.dumps(files), mimetype="application/json")


@app.route("/spec/<material_no>/drawing", methods=["POST"])
@perm_required("material_edit")
def spec_drawing_assign(material_no):
    """수동 도면 파일 지정."""
    drawing_file = request.form.get("drawing_file", "").strip() or None
    db.update_drawing_file(material_no, drawing_file)
    record_change("도면 파일 지정", "material", material_no,
                  f"drawing_file={drawing_file or '(해제)'}")
    flash("도면 파일 지정이 저장됐어.")
    return redirect(url_for("spec_detail", material_no=material_no))


@app.route("/spec/<material_no>/drawing/upload", methods=["POST"])
@perm_required("material_edit")
def spec_drawing_upload(material_no):
    """도면 PDF 파일을 직접 업로드. 자재번호.pdf 로 저장해서 자동 매칭되게 한다."""
    file = request.files.get("drawing_pdf")
    if not file or file.filename == "":
        flash("업로드할 도면 파일을 선택해줘.")
        return redirect(url_for("spec_detail", material_no=material_no))
    ext = os.path.splitext(file.filename)[1].lower()
    if ext != ".pdf":
        flash("PDF 파일만 업로드할 수 있어.")
        return redirect(url_for("spec_detail", material_no=material_no))

    db.upsert_material(material_no)
    fname = f"{material_no}.pdf"
    file.save(os.path.join(DRAWING_DIR, fname))
    # 수동 지정이 남아있으면 방금 올린 파일 대신 옛 파일이 계속 쓰이니 해제해서
    # 자동 매칭(자재번호.pdf = 방금 올린 파일)이 바로 적용되게 한다.
    db.update_drawing_file(material_no, None)
    record_change("도면 파일 업로드", "material", material_no, f"업로드: {fname}")
    flash("도면 파일이 업로드됐어.")
    return redirect(url_for("spec_detail", material_no=material_no))


# ---------- 계정 관리 (관리자 전용) ----------

@app.route("/users", methods=["GET", "POST"])
@perm_required("users")
def user_management():
    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            display_name = request.form.get("display_name", "").strip()
            perms = [p for p in request.form.getlist("permissions") if p in ALL_PERMS]
            if not username or not password:
                flash("아이디와 비밀번호를 입력해줘.")
            elif db.get_user_by_username(username):
                flash("이미 있는 아이디야.")
            else:
                db.create_user(username, password, display_name, ",".join(perms))
                perm_names = ", ".join(PERM_LABELS.get(p, p) for p in perms) or "없음"
                record_change("계정 생성", "user", username, f"권한: {perm_names}")
                new_user = db.get_user_by_username(username)
                flash(f"계정 '{username}' 만들어졌어. 아래에서 권한을 설정해줘. (임시 비밀번호는 첫 로그인 후 본인이 바꾸게 안내)")
                if new_user:
                    return redirect(url_for("user_detail", user_id=new_user["id"]))

        elif action == "delete":
            user_id = int(request.form.get("user_id"))
            if user_id == g.user["id"]:
                flash("본인 계정은 삭제할 수 없어.")
            else:
                target_user = db.get_user(user_id)
                db.delete_user(user_id)
                flash("계정 삭제됐어.")
                record_change("계정 삭제", "user", target_user["username"] if target_user else user_id)

        elif action == "reset_password":
            user_id = int(request.form.get("user_id"))
            new_password = request.form.get("new_password", "").strip()
            if not new_password:
                flash("새 비밀번호를 입력해줘.")
            else:
                db.update_user_password(user_id, new_password)
                flash("비밀번호가 변경됐어.")
                target_user = db.get_user(user_id)
                record_change("비밀번호 초기화(관리자)", "user", target_user["username"] if target_user else user_id)

        elif action == "update_profile":
            user_id = int(request.form.get("user_id"))
            username = request.form.get("username", "").strip()
            display_name = request.form.get("display_name", "").strip()
            if not username:
                flash("아이디는 비워둘 수 없어.")
            else:
                ok, error = db.update_user_profile(user_id, username, display_name)
                if not ok:
                    flash(error)
                else:
                    flash("계정 정보가 수정됐어.")
                    record_change("계정 정보 수정", "user", username, f"이름: {display_name}")

        elif action == "update_permissions":
            user_id = int(request.form.get("user_id"))
            perms = [p for p in request.form.getlist("permissions") if p in ALL_PERMS]
            target_user = db.get_user(user_id)
            if target_user is None:
                flash("존재하지 않는 계정이야.")
            elif user_id == g.user["id"] and "users" not in perms:
                flash("본인의 계정 관리 권한은 스스로 회수할 수 없어(잠길 수 있어서). 다른 관리자 계정으로 바꿔줘.")
            else:
                db.update_user_permissions(user_id, ",".join(perms))
                flash(f"'{target_user['username']}' 계정 권한이 변경됐어.")
                perm_names = ", ".join(PERM_LABELS.get(p, p) for p in perms) or "없음"
                record_change("계정 권한 변경", "user", target_user["username"], f"권한: {perm_names}")

                # 최종결정권자 체크 — 권한을 먼저 저장한 뒤에 처리해야 '승인 권한' 검사가 맞는다
                want_final = request.form.get("is_final_approver") == "1"
                was_final = bool(target_user["is_final_approver"])
                if want_final != was_final:
                    ok, err = db.set_final_approver(user_id, want_final)
                    if not ok:
                        flash(err)
                    else:
                        label = "지정" if want_final else "해제"
                        flash(f"'{target_user['username']}' 최종결정권자 {label}됐어.")
                        record_change(f"최종결정권자 {label}", "user", target_user["username"], "")
                elif was_final and "approve" not in perms:
                    # 승인 권한이 빠지면 최종결정권자 자격도 같이 내려간다
                    db.set_final_approver(user_id, False)
                    flash(f"'{target_user['username']}'의 승인 권한이 빠져서 최종결정권자 지정도 해제됐어.")
                    record_change("최종결정권자 해제", "user", target_user["username"], "승인 권한 회수에 따른 자동 해제")

        # 상세페이지에서 온 요청이면 그 페이지로 돌아간다
        if request.form.get("return_to") == "detail" and request.form.get("user_id"):
            try:
                return redirect(url_for("user_detail", user_id=int(request.form.get("user_id"))))
            except (ValueError, TypeError):
                pass
        return redirect(url_for("user_management"))

    users = db.list_users()
    return render_template("users.html", users=users, current_user_id=g.user["id"],
                           perm_labels=PERM_LABELS, all_perms=ALL_PERMS)


@app.route("/users/<int:user_id>")
@perm_required("users")
def user_detail(user_id):
    """계정 하나의 상세·권한 설정 페이지."""
    target = db.get_user(user_id)
    if target is None:
        flash("존재하지 않는 계정이야.")
        return redirect(url_for("user_management"))
    u_perms = set((target["permissions"] or "").split(","))
    finals = db.list_final_approvers()
    return render_template("user_detail.html", u=target, u_perms=u_perms,
                           perm_groups=PERM_GROUPS, perm_labels=PERM_LABELS,
                           current_user_id=g.user["id"],
                           final_approvers=finals,
                           max_final_approvers=db.MAX_FINAL_APPROVERS)


@app.route("/logs")
@perm_required("logs")
def activity_logs():
    logs = db.list_activity_logs(limit=300)
    return render_template("logs.html", logs=logs)



# AQL별 샘플수량 구간표 — 성적서_양식.xlsx의 IFS 수식을 그대로 옮긴 것.
# (수량 상한, 샘플수) 튜플 리스트. 새 AQL값이 나오면 이 표에 추가해야 함(임의 계산 금지).
AQL_TABLES = {
    4: [(8, 2), (15, 2), (25, 3), (50, 5), (90, 5), (150, 8), (280, 13), (500, 20),
        (1200, 32), (3200, 50), (10000, 88), (35000, 125), (150000, 200), (500000, 315),
        (float("inf"), 500)],
    0.65: [(8, 3), (15, 5), (25, 8), (50, 13), (90, 20), (150, 32), (280, 50), (500, 88),
           (1200, 125), (3200, 200), (10000, 315), (35000, 500), (150000, 800), (500000, 1250),
           (float("inf"), 2000)],
    1.5: [(8, 2), (15, 3), (25, 5), (50, 8), (90, 13), (150, 20), (280, 32), (500, 50),
          (1200, 88), (3200, 125), (10000, 200), (35000, 315), (150000, 500), (500000, 800),
          (float("inf"), 1250)],
}


def sample_size(aql, quantity):
    """
    AQL과 입고수량으로 샘플수량 계산.
    - aql이 숫자(4/1.5/0.65 등)면 기존 구간표 사용
    - aql == "전수" 면 입고수량 전부
    - aql이 "퍼센트N" (예: "퍼센트10") 이면 입고수량의 N%를 반올림한 정수 (최소 1개)
    표에 없는 숫자 AQL이면 None(화면에서 확인 요청).
    """
    if quantity is None or aql is None:
        return None
    try:
        quantity = int(quantity)
    except (TypeError, ValueError):
        return None

    aql_str = str(aql).strip()
    if aql_str == "전수":
        return quantity
    if aql_str.startswith("퍼센트"):
        try:
            percent = float(aql_str[3:])
        except ValueError:
            return None
        # "반올림해서 정수" — 사사오입(0.5 이상 올림) 방식으로 계산
        raw = quantity * percent / 100
        result = int(raw + 0.5) if raw >= 0 else int(raw - 0.5)
        return max(1, min(quantity, result))

    try:
        aql_num = float(aql)
        if aql_num == int(aql_num):
            aql_num = int(aql_num)
    except (TypeError, ValueError):
        return None
    table = AQL_TABLES.get(aql_num)
    if not table:
        return None
    for limit, size in table:
        if quantity <= limit:
            return min(quantity, size)
    return quantity


# KS Q ISO 2859-1(=ANSI/ASQ Z1.4) 1회 샘플링 방식, 보통검사(Normal), Ac(합격판정개수).
# 세원플라텍 공식 AQL표(사용자 제공, aql_dump2.txt로 셀단위 검증)에서 추출.
# 표에 화살표(↓/↑)로 비어있는 칸은 legend 규칙대로 캐스케이드 해서 이미 실수치로 풀어놓은 것 —
# 화살표를 여기서 다시 계산하지 말 것.
#
# key = AQL_TABLES와 완전히 동일한 샘플수(n) 값(코드문자 J의 표준 n=80 대신 앱은 88을 쓰므로
# 88 자리에 J행의 Ac를 그대로 넣어뒀다 — 88 > 80이라 더 엄격한 쪽으로 안전).
# value = Ac(합격판정개수). Re는 이 표 전 구간에서 항상 Ac+1이라 별도로 안 둠.
AQL_AC_TABLE = {
    0.65: {2: 0, 3: 0, 5: 0, 8: 0, 13: 0, 20: 0, 32: 1, 50: 1, 88: 1,
           125: 2, 200: 3, 315: 5, 500: 7, 800: 10, 1250: 14, 2000: 21},
    1.5:  {2: 0, 3: 0, 5: 0, 8: 0, 13: 1, 20: 1, 32: 1, 50: 2, 88: 3,
           125: 5, 200: 7, 315: 10, 500: 14, 800: 21, 1250: 21, 2000: 21},
    4:    {2: 0, 3: 0, 5: 1, 8: 1, 13: 1, 20: 2, 32: 3, 50: 5, 88: 7,
           125: 10, 200: 14, 315: 21, 500: 21, 800: 21, 1250: 21, 2000: 21},
}


def aql_ac_allowance(aql, sample_qty):
    """AQL과 샘플수량으로 '합격 허용 불량개수'(Ac)를 돌려준다.

    실측은 항상 최대 6개까지만 하지만(inspect_form.html 참고), 판정 기준(Ac)은
    AQL로 계산된 진짜 표본수(sample_qty)에 해당하는 표준표 값을 그대로 쓴다 — 표본수 자체를
    6개로 줄인 게 아니라 "6개 중 Ac개까지는 봐준다"는 참고용 적용이라는 뜻(사용자 확정).

    sample_qty가 AQL_AC_TABLE에 없는 값이면(전수/퍼센트N AQL, 또는 입고수량이 표본수보다
    작아서 sample_size()가 입고수량 그대로 돌려준 경우 등) 기존처럼 Ac=0(무결점)으로 안전하게
    처리한다.
    """
    try:
        aql_num = float(aql)
        if aql_num == int(aql_num):
            aql_num = int(aql_num)
    except (TypeError, ValueError):
        return 0
    table = AQL_AC_TABLE.get(aql_num)
    if not table or sample_qty is None:
        return 0
    return table.get(sample_qty, 0)


def format_duration(total_sec):
    """N시간 N분 N초 형식. 시간/분이 0이면 생략, 다 0이면 '0초'."""
    total_sec = int(total_sec or 0)
    h, rem = divmod(total_sec, 3600)
    m, s = divmod(rem, 60)
    parts = []
    if h:
        parts.append(f"{h}시간")
    if m:
        parts.append(f"{m}분")
    if s or not parts:
        parts.append(f"{s}초")
    return " ".join(parts)


EXTRA_SAMPLE_SEC = 10  # 기준 AQL 초과 샘플 1개당 추가 시간(고정값, 단위: 초)


def build_specs_with_sample(specs, quantity):
    """규격 목록에 화면·계산용 파생값을 붙여서 돌려준다.

    - sample_qty   : 항목의 AQL과 입고수량으로 계산한 샘플수량
    - ac_allowance : 위 샘플수량 기준 합격판정개수(Ac). re_allowance는 항상 Ac+1(호출부에서 계산)
    - no_limit     : 숫자 판정인데 하한·상한이 둘 다 비어 있는 항목
                     (judge_numeric()이 이런 항목을 무조건 불합격 처리하므로 화면에서 경고해야 함)

    검사 입력 / 성적서 상세 / 재검사 / 시간계산이 전부 같은 계산을 쓰도록 한 곳에 모아둔 헬퍼.
    (예전엔 같은 리스트 컴프리헨션이 8군데에 복사돼 있어서 한 곳만 고치면 나머지가 어긋났다)
    """
    out = []
    for s in specs:
        d = dict(s)
        d["sample_qty"] = sample_size(d.get("aql"), quantity)
        d["ac_allowance"] = aql_ac_allowance(d.get("aql"), d["sample_qty"]) if d.get("aql") is not None else None
        d["no_limit"] = (d.get("judge_type") in ("numeric", "numeric_pair")
                         and d.get("lower_limit") is None
                         and d.get("upper_limit") is None)
        out.append(d)
    return out


def _norm_method(s):
    """계측기 이름/측정 방식 매칭용 정규화 — 공백을 모두 없애고 소문자로.
    (사용자 확정: '도장두께측정기'와 '도장 두께 측정기'는 같은 것으로 본다)"""
    import re
    return re.sub(r"\s+", "", (s or "")).strip().lower()


def gauge_method_options():
    """자재 등록 화면의 '측정 방식' 드롭다운에 넣을 후보 목록.
    등록된 계측기 이름 + 시각검사(육안/외관) — 규격 등록 시 이 중에서 고르되
    직접 입력도 가능(datalist)."""
    names = [g["name"] for g in db.list_gauges() if (g["name"] or "").strip()]
    for v in ("육안", "외관"):
        if v not in names:
            names.append(v)
    # 중복 제거(정규화 기준) 후 정렬
    seen, out = set(), []
    for n in names:
        k = _norm_method(n)
        if k and k not in seen:
            seen.add(k)
            out.append(n)
    return sorted(out)


def build_sel_gauge(specs, gauges, prefill=None):
    """항목별로 '어느 계측기가 선택돼야 하는지'(계측기 이름)를 정해서 dict로 돌려준다.
    - 이미 입력된 값(prefill의 gauge_name)이 있으면 그게 우선
    - 없으면 항목의 측정 방식(inspect_method)을 등록된 계측기와 공백무시로 매칭해 자동 선택
    - 같은 이름(정규화)이 여러 계측기에 있으면 애매하므로 자동선택하지 않는다
    검사폼에서 유효기간이 자동으로 채워지게 하는 핵심."""
    idx, dup = {}, set()
    for gm in gauges:
        k = _norm_method(gm["name"])
        if not k:
            continue
        if k in idx:
            dup.add(k)
        else:
            idx[k] = gm["name"]
    for k in dup:
        idx.pop(k, None)

    sel = {}
    for s in specs:
        sid = s["id"]
        pf_name = (prefill or {}).get(f"item_{sid}_gauge_name") if prefill else None
        pf_exp = (prefill or {}).get(f"item_{sid}_gauge_expiry") if prefill else None
        if pf_name:                      # 이미 고른 계측기가 있으면 그대로 둠
            sel[sid] = pf_name
            continue
        if pf_exp:                       # 직접 입력한 유효기간이 있으면 자동선택 안 함
            continue
        if s.get("judge_type") not in ("numeric", "numeric_pair", "ok_ng"):
            continue
        method = s.get("inspect_method") or ""
        if "육안" in method:
            continue
        name = idx.get(_norm_method(method))
        if name:
            sel[sid] = name
    return sel


def _unique_aql_sample_qty(specs_with_sample):
    """AQL 그룹당 한 번만 세어 {AQL(숫자 또는 '전수'/'퍼센트N' 문자열): {"qty":샘플수량, "ac":Ac}} 딕셔너리 반환."""
    seen_aql = {}
    for s in specs_with_sample:
        aql = s.get("aql")
        qty = s.get("sample_qty")
        if aql is None or qty is None:
            continue
        aql_str = str(aql).strip()
        if aql_str == "전수" or aql_str.startswith("퍼센트"):
            key = aql_str
        else:
            try:
                key = float(aql)
                if key == int(key):
                    key = int(key)
            except (TypeError, ValueError):
                continue
        if key not in seen_aql:
            seen_aql[key] = {"qty": qty, "ac": s.get("ac_allowance")}
    return seen_aql


def compute_total_time_sec(specs_with_sample, per_cycle_sec):
    """
    총 측정 시간 계산.
    - 전수/퍼센티지 전용: 샘플수 × 개당시간 (단순 곱셈)
    - 수치 AQL 혼합: 가장 높은 AQL(샘플 가장 적음)을 동적 기준으로 삼아,
      나머지 그룹의 초과 샘플만 EXTRA_SAMPLE_SEC(10초)로 추가.
      기준 AQL이 4 없어도 현재 스펙에서 가장 높은 수치를 자동으로 선택.
    """
    if not per_cycle_sec:
        return 0
    seen_aql = _unique_aql_sample_qty(specs_with_sample)
    if not seen_aql:
        return 0

    special = {k: v["qty"] for k, v in seen_aql.items() if isinstance(k, str)}
    numeric = {k: v["qty"] for k, v in seen_aql.items() if not isinstance(k, str)}

    # 전수/퍼센티지만 있는 경우 → 샘플수 × 개당시간 (단순 계산)
    if not numeric:
        return max(special.values()) * per_cycle_sec

    # 수치 AQL이 있는 경우 → 가장 높은 AQL(샘플 가장 적음)을 기준으로
    base_aql = max(numeric.keys())   # 4 > 1.5 > 0.65 — 숫자 클수록 샘플 적음
    base_qty = numeric[base_aql]
    total = base_qty * per_cycle_sec
    for aql_key, v in seen_aql.items():
        if aql_key == base_aql:
            continue
        extra = max(0, v["qty"] - base_qty)
        total += extra * EXTRA_SAMPLE_SEC
    return total


def total_time_label_for(header):
    """목록 화면(승인 목록/검사 이력)에 표시할 '총 측정 시간' — 성적서 상세와 동일한 계산."""
    per_cycle_sec = header["actual_time_sec"] or 0
    if not per_cycle_sec:
        return "-"
    specs, _, _ = _get_specs_for_material(header["material_no"])
    if not specs:
        return "-"
    specs_with_sample = build_specs_with_sample(specs, header["quantity"])
    return format_duration(compute_total_time_sec(specs_with_sample, per_cycle_sec))


@app.route("/")
@login_required
def home():
    pending_inspect = db.list_intake(status="대기")
    pending_approve = db.list_inspections(status="pending")
    pending_output  = db.list_pending_output_inspections()
    recent          = db.list_inspections()[:10]
    today_stats     = db.get_today_stats()
    gauge_warnings        = db.get_gauge_expiry_warnings(days=15)
    gauge_master_warnings = db.get_gauge_master_warnings(days=30)
    repeat_defects        = db.get_repeat_defects(min_count=3)
    return render_template("home.html",
        pending_inspect=pending_inspect,
        pending_approve=pending_approve,
        pending_output=pending_output,
        recent=recent,
        today_stats=today_stats,
        gauge_warnings=gauge_warnings,
        gauge_master_warnings=gauge_master_warnings,
        repeat_defects=repeat_defects,
    )


@app.route("/today")
@perm_required("inspect_input", "inspect_history", "approve", "intake")
def daily_status():
    """금일 현황 — 홈 요약보다 자세한 하루치 상황판.
    아침에 열어서 '오늘 뭐부터 해야 하나'를 바로 알 수 있게 만든 화면."""
    day = request.args.get("date", "").strip()
    d = _parse_any_date(day) if day else None
    return render_template("daily_status.html",
                           s=db.daily_status(d.isoformat() if d else None),
                           today=_dt.now().strftime("%Y-%m-%d"))


# ---------- 입고 리스트 (엑셀 붙여넣기 등록) ----------

def _merge_duplicate_intake_rows(rows):
    """
    등록 화면(붙여넣기)에서 같은 자재번호가 여러 줄로 들어오면 발주번호만 다른
    한 배치 입고로 보고 한 줄로 합친다 — 입고수량은 더하고, 발주번호는 줄바꿈으로 이어붙임.
    (사용자 확정: 이번 등록 화면 안에서만 합친다. 이미 대기 중인 기존 건과는 합치지 않음.)
    """
    merged = {}
    order = []
    for r in rows:
        key = r["material_no"]
        if key not in merged:
            merged[key] = dict(r)
            order.append(key)
            continue
        existing = merged[key]
        try:
            existing_qty = int(existing["quantity"]) if existing["quantity"] not in (None, "") else 0
        except (TypeError, ValueError):
            existing_qty = None
        try:
            new_qty = int(r["quantity"]) if r["quantity"] not in (None, "") else 0
        except (TypeError, ValueError):
            new_qty = None
        if existing_qty is None or new_qty is None:
            # 숫자로 못 바꾸는 값이 섞여 있으면 합산 대신 원문을 이어붙임(값 유실 방지)
            parts = [str(p) for p in (existing["quantity"], r["quantity"]) if p not in (None, "")]
            existing["quantity"] = ", ".join(parts) if parts else None
        else:
            existing["quantity"] = existing_qty + new_qty
        po_parts = [p for p in (existing.get("po_number") or "").split("\n") if p]
        new_po = (r.get("po_number") or "").strip()
        if new_po and new_po not in po_parts:
            po_parts.append(new_po)
        existing["po_number"] = "\n".join(po_parts)
        if not existing.get("product_name") and r.get("product_name"):
            existing["product_name"] = r["product_name"]
        if not existing.get("assembly_no") and r.get("assembly_no"):
            existing["assembly_no"] = r["assembly_no"]
    return [merged[k] for k in order]


@app.route("/intake", methods=["GET", "POST"])
@perm_required("intake")
def intake():
    if request.method == "POST":
        import json
        raw_json = request.form.get("rows_json", "")
        try:
            grid_rows = json.loads(raw_json) if raw_json else []
        except (ValueError, TypeError):
            grid_rows = []

        # 조립품 파츠 자동 전개 여부 — 파츠 하나만 스페어로 받는 경우엔 꺼야 한다
        expand_assembly = request.form.get("expand_assembly") == "1"

        rows = []
        errors = []
        # 스프레드시트 열 순서: 입고날짜, 납품업체, 발주번호, 제품명, 자재번호, 입고수량
        for line_no, cols in enumerate(grid_rows, start=1):
            cols = [(c or "").strip() for c in cols]
            while len(cols) < 6:
                cols.append("")
            receive_date, supplier, po_number, product_name, material_no, quantity = cols[:6]
            if not any([receive_date, supplier, po_number, product_name, material_no, quantity]):
                continue  # 완전히 빈 행은 건너뜀
            if not material_no:
                errors.append(f"{line_no}번째 행: 자재번호가 비어 있어 → 건너뜀")
                continue

            # 조립품 자동 전개: 파츠번호 입력 -> 그 파츠가 속한 조립품 찾기 -> 파츠 전체를 펼침
            ma_info = db.get_ma_by_component(material_no) if expand_assembly else None
            if ma_info:
                # MA 입고: 파츠별로 입고 행 생성 (각 파츠의 제품명은 materials 테이블에서 자동 조회)
                for component_no in ma_info["components"]:
                    # 각 파츠의 제품명 조회
                    component_material = db.get_material(component_no)
                    component_name = component_material["material_name"] if component_material else product_name

                    rows.append({
                        "material_no": component_no,
                        "quantity": quantity or None,
                        "supplier": supplier,
                        "receive_date": receive_date,
                        "po_number": po_number,
                        "product_name": component_name,
                        "assembly_no": ma_info["ma_master"],
                    })
            else:
                # 일반 자재
                rows.append({
                    "material_no": material_no,
                    "quantity": quantity or None,
                    "supplier": supplier,
                    "receive_date": receive_date,
                    "po_number": po_number,
                    "product_name": product_name,
                })

        if errors:
            for e in errors:
                flash(e)
        rows = _merge_duplicate_intake_rows(rows)
        if rows:
            db.add_intake_bulk(rows)
            flash(f"{len(rows)}건 등록 완료")
            material_list = ", ".join(r["material_no"] for r in rows[:10])
            if len(rows) > 10:
                material_list += " 외"
            record_change("입고 리스트 등록", "intake", None, f"{len(rows)}건 ({material_list})")
        elif not errors:
            flash("등록할 내용이 없어.")
        return redirect(url_for("intake"))

    pending = db.list_intake(status="대기")
    done = db.list_intake(status="검사완료")
    mats = db.get_materials()
    registered = {m["material_no"] for m in mats}
    name_map = {m["material_no"]: m["material_name"] for m in mats}

    # 탭 검색어 + 페이지네이션 — 대기/완료 각자 별개(?tab=/?q=/?pending_page=/?done_page=)
    active_tab = (request.args.get("tab") or "pending").lower()
    if active_tab not in ("pending", "done"):
        active_tab = "pending"
    q = (request.args.get("q") or "").strip().lower()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    d_from = _parse_any_date(date_from) if date_from else None
    d_to = _parse_any_date(date_to) if date_to else None

    def _match(rows):
        out = []
        for r in rows:
            if q:
                name = name_map.get(r["material_no"]) or r["product_name"] or ""
                hay = " ".join(str(x or "") for x in (
                    r["material_no"], name, r["supplier"], r["po_number"])).lower()
                if q not in hay:
                    continue
            if d_from or d_to:
                rd = _parse_any_date(r["receive_date"])
                if not rd:
                    continue
                if d_from and rd < d_from:
                    continue
                if d_to and rd > d_to:
                    continue
            out.append(r)
        return out

    pending_filtered = _match(pending)
    done_filtered = _match(done)
    pending_pager = _paginate(pending_filtered, page_arg="pending_page")
    done_pager = _paginate(done_filtered, page_arg="done_page")

    return render_template("intake.html",
                           pending_pager=pending_pager, done_pager=done_pager,
                           pending_total_all=len(pending), done_total_all=len(done),
                           active_tab=active_tab, q=q,
                           date_from=date_from, date_to=date_to,
                           registered=registered, group_nos=set(), name_map=name_map)


# ---------- 규격 관리 ----------

@app.route("/spec")
@perm_required("material_view")
def spec_list():
    query = request.args.get("q", "").strip()
    search_by = request.args.get("by", "all")
    # 검사방식 미입력(method_empty)은 검색어가 없어도 조회한다.
    if query or search_by == "method_empty":
        materials = db.search_materials(query, search_by)
    else:
        materials = db.get_materials()
    materials = list(materials)
    pager = _paginate(materials)
    drawing_materials = materials_with_drawings(m["material_no"] for m in pager["items"])
    return render_template("spec.html", materials=pager["items"], query=query, search_by=search_by,
                            drawing_materials=drawing_materials, pager=pager)


@app.route("/spec/quick_add", methods=["GET", "POST"])
@perm_required("material_edit")
def spec_quick_add():
    """
    자재번호+제품명을 스프레드시트형으로 여러 줄 한번에 등록.
    "측정조건 입력" 팝업에서 항목까지 같이 입력했으면 그 항목들도 함께 등록됨.
    """
    if request.method == "POST":
        import json
        raw_rows_json = request.form.get("rows_json", "")
        raw_items_json = request.form.get("items_json", "")
        try:
            grid_rows = json.loads(raw_rows_json) if raw_rows_json else []
        except (ValueError, TypeError):
            grid_rows = []
        try:
            items_per_row = json.loads(raw_items_json) if raw_items_json else []
        except (ValueError, TypeError):
            items_per_row = []

        rows = []
        rows_items = []  # rows와 같은 순서로 매칭되는 항목 리스트
        for idx, cols in enumerate(grid_rows):
            cols = [(c or "").strip() for c in cols]
            while len(cols) < 2:
                cols.append("")
            material_no, material_name = cols[:2]
            if not material_no:
                continue
            rows.append({"material_no": material_no, "material_name": material_name})
            rows_items.append(items_per_row[idx] if idx < len(items_per_row) else [])

        item_count_total = 0
        if rows:
            db.upsert_materials_bulk(rows)
            for r, raw_items in zip(rows, rows_items):
                parsed_items = []
                for order, raw_it in enumerate(raw_items, start=1):
                    if not isinstance(raw_it, dict):
                        continue
                    spec_display = (raw_it.get("spec_display") or "").strip()
                    if not spec_display:
                        continue
                    item_name = (raw_it.get("item_name") or "").strip() or f"항목{order}"
                    judge_type = raw_it.get("judge_type") or "numeric"
                    inspect_method = (raw_it.get("inspect_method") or "").strip()

                    if judge_type in ("visual", "ok_ng"):
                        lower_limit = upper_limit = None
                    else:
                        try:
                            lower_limit = float(raw_it.get("lower_limit")) if raw_it.get("lower_limit") not in (None, "") else None
                        except (TypeError, ValueError):
                            lower_limit = None
                        try:
                            upper_limit = float(raw_it.get("upper_limit")) if raw_it.get("upper_limit") not in (None, "") else None
                        except (TypeError, ValueError):
                            upper_limit = None

                    aql_select = raw_it.get("aql_select") or "4"
                    if aql_select == "전수":
                        aql = "전수"
                    elif aql_select == "퍼센트":
                        try:
                            percent_num = float(raw_it.get("aql_percent") or 0)
                            aql = f"퍼센트{percent_num:g}"
                        except ValueError:
                            aql = "퍼센트0"
                    else:
                        try:
                            aql_f = float(aql_select)
                            aql = int(aql_f) if aql_f == int(aql_f) else aql_f
                        except ValueError:
                            aql = aql_select

                    try:
                        item_order = int(raw_it.get("item_order") or order)
                    except (TypeError, ValueError):
                        item_order = order

                    parsed_items.append({
                        "item_name": item_name, "spec_display": spec_display, "judge_type": judge_type,
                        "lower_limit": lower_limit, "upper_limit": upper_limit,
                        "inspect_method": inspect_method, "aql": aql, "item_order": item_order,
                    })

                if parsed_items:
                    db.replace_specs_for_material(r["material_no"], r["material_name"], parsed_items)
                    item_count_total += len(parsed_items)

            flash(f"{len(rows)}개 자재 등록 완료" +
                  (f" (측정조건 총 {item_count_total}개 항목 함께 등록됨)" if item_count_total else
                   " — 이제 각 자재의 규격 상세 화면에서 항목을 추가해줘."))
            material_list = ", ".join(r["material_no"] for r in rows[:10])
            if len(rows) > 10:
                material_list += " 외"
            record_change("규격 개별 등록", "spec", None,
                          f"{len(rows)}건 ({material_list}), 항목 {item_count_total}개")
        else:
            flash("등록할 내용이 없어.")
        return redirect(url_for("spec_quick_add"))

    return render_template("spec_quick_add.html", method_options=gauge_method_options())


@app.route("/spec/delete_bulk", methods=["POST"])
@perm_required("material_edit")
def spec_delete_bulk():
    material_nos = request.form.getlist("material_nos")
    if not material_nos:
        flash("삭제할 자재를 선택해줘.")
        return redirect(url_for("spec_list"))
    db.delete_materials_bulk(material_nos)
    flash(f"{len(material_nos)}개 자재의 규격을 삭제했어.")
    record_change("규격 자재 일괄 삭제", "spec", None, ", ".join(material_nos[:10]))
    return redirect(url_for("spec_list"))


def _resolve_material_name(material_no, specs=None):
    """자재명은 materials 테이블이 정본이다 — specs.material_name은 등록 당시 값을
    그대로 복사해둔 사본이라 나중에 자재명을 고쳐도 안 따라가서 어긋날 수 있다
    (실제로 30개 자재에서 비어있거나 다른 값으로 발견됨, 2026-09-01).
    검사이력·승인 화면에 이름이 빈 채로 찍히던 원인이 이 우선순위가 거꾸로였기 때문 —
    materials를 최우선으로 보고, 거기 없을 때만 specs 사본을 대신 쓴다."""
    material = db.get_material(material_no)
    if material and material["material_name"]:
        return material["material_name"]
    if specs is None:
        specs = db.get_specs_by_material(material_no)
    if specs and specs[0]["material_name"]:
        return specs[0]["material_name"]
    return None


@app.route("/spec/<material_no>")
@perm_required("material_view")
def spec_detail(material_no):
    specs = db.get_specs_by_material(material_no)
    material = db.get_material(material_no)
    material_name = _resolve_material_name(material_no, specs)
    drawing_no = report_builder.compute_drawing_no(material_no)
    drawing_pdf = find_drawing_pdf(material_no)
    drawing_auto = os.path.join(DRAWING_DIR, f"{material_no}.pdf")
    drawing_has_auto = os.path.exists(drawing_auto)
    full_inspect_config = db.get_full_inspect_config(material_no)
    return render_template("spec_detail.html", material_no=material_no, specs=specs,
                           material_name=material_name, material=material, drawing_no=drawing_no,
                           drawing_pdf=drawing_pdf, drawing_has_auto=drawing_has_auto,
                           full_inspect_config=full_inspect_config,
                           method_options=gauge_method_options())


@app.route("/spec/<material_no>/full-inspect-config", methods=["POST"])
@perm_required("material_edit")
def full_inspect_config_save(material_no):
    """전수검사 활성화 토글 저장 (form POST). 열 구성은 specs 테이블에서 자동으로 가져옴."""
    action = request.form.get("action", "")
    if action == "disable":
        db.set_full_inspect_config(material_no, None)
        flash("전수검사 설정을 해제했어.")
    else:
        note = request.form.get("note", "").strip()
        db.set_full_inspect_config(material_no, {"enabled": True, "note": note})
        flash("전수검사 설정을 저장했어.")
    record_change("전수검사 설정", "material", material_no, "")
    return redirect(url_for("spec_detail", material_no=material_no))


@app.route("/spec/<material_no>/standard_info", methods=["POST"])
@perm_required("material_edit")
def spec_standard_info_update(material_no):
    drawing_version = request.form.get("drawing_version", "1").strip() or "1"
    revision_date = request.form.get("revision_date", "").strip()
    edition_raw = request.form.get("edition", "1").strip()
    unit = request.form.get("unit", "mm").strip() or "mm"
    try:
        edition = int(edition_raw)
    except ValueError:
        edition = 1

    db.update_material_standard_info(material_no, drawing_version, revision_date, edition, unit)
    flash("기준서 정보가 저장됐어.")
    record_change("기준서 정보 수정", "material", material_no,
                  f"버전{drawing_version}/{revision_date}/제{edition}판/{unit}")
    return redirect(url_for("spec_detail", material_no=material_no))


@app.route("/spec/<material_no>/rename", methods=["POST"])
@perm_required("material_edit")
def spec_material_rename(material_no):
    new_material_no = request.form.get("new_material_no", "").strip()
    new_material_name = request.form.get("new_material_name", "").strip()
    if not new_material_no:
        flash("자재번호는 비워둘 수 없어.")
        return redirect(url_for("spec_detail", material_no=material_no))

    ok, error = db.rename_material(material_no, new_material_no, new_material_name)
    if not ok:
        flash(error)
        return redirect(url_for("spec_detail", material_no=material_no))

    flash(f"자재번호가 '{new_material_no}'(으)로 변경됐어. (과거 성적서는 옛 번호로 그대로 남아있어)")
    record_change("자재번호/자재명 변경", "material", new_material_no,
                  f"{material_no} → {new_material_no} ({new_material_name})")
    return redirect(url_for("spec_detail", material_no=new_material_no))


def _parse_spec_form(form):
    """규격 수정/추가 폼에서 값을 뽑아 db 저장용 튜플로 정리."""
    item_name = form.get("item_name", "").strip()
    spec_display = form.get("spec_display", "").strip()
    judge_type = form.get("judge_type", "numeric")
    inspect_method = form.get("inspect_method", "").strip()
    aql_raw = form.get("aql", "").strip()
    order_raw = form.get("item_order", "").strip()

    lower_raw = form.get("lower_limit", "").strip()
    upper_raw = form.get("upper_limit", "").strip()
    if judge_type in ("visual", "ok_ng"):
        lower_limit = upper_limit = None
    else:
        try:
            lower_limit = float(lower_raw) if lower_raw else None
        except ValueError:
            lower_limit = None
        try:
            upper_limit = float(upper_raw) if upper_raw else None
        except ValueError:
            upper_limit = None

    aql_select = form.get("aql_select", "").strip()
    percent_raw = form.get("aql_percent", "").strip()
    if aql_select == "전수":
        aql = "전수"
    elif aql_select == "퍼센트":
        try:
            percent_num = float(percent_raw)
            aql = f"퍼센트{percent_num:g}"
        except ValueError:
            aql = "퍼센트0"
    elif aql_select:
        try:
            aql = float(aql_select)
            if aql == int(aql):
                aql = int(aql)
        except ValueError:
            aql = aql_select
    else:
        # 구버전 폼 호환(aql_select가 없는 경우) — 기존처럼 aql 필드 그대로 파싱
        try:
            aql = float(aql_raw) if aql_raw else None
            if aql is not None and aql == int(aql):
                aql = int(aql)
        except ValueError:
            aql = aql_raw or None

    try:
        item_order = int(order_raw) if order_raw else 0
    except ValueError:
        item_order = 0

    return item_name, spec_display, judge_type, lower_limit, upper_limit, inspect_method, aql, item_order


@app.route("/spec/<material_no>/item/<int:spec_id>/update", methods=["POST"])
@perm_required("material_edit")
def spec_item_update(material_no, spec_id):
    item_name, spec_display, judge_type, lower_limit, upper_limit, inspect_method, aql, item_order = \
        _parse_spec_form(request.form)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    # 항목기호가 같은 자재의 다른 항목과 겹치면 조회 시 조인이 두 배로 불어나서
    # 항목이 겹쳐 보이거나 판정이 엉뚱하게 섞이는 사고가 실제로 있었다(2026-08-30 발견).
    existing = db.get_specs_by_material(material_no)
    if any((s["item_name"] or "") == item_name and s["id"] != spec_id for s in existing):
        msg = f"항목기호 '{item_name}'는 이미 이 자재의 다른 항목에서 쓰고 있어. 다른 기호를 써줘."
        if is_ajax:
            return jsonify({"ok": False, "error": msg}), 400
        flash(msg)
        return redirect(url_for("spec_detail", material_no=material_no))
    db.update_spec_item(spec_id, item_name, spec_display, judge_type,
                        lower_limit, upper_limit, inspect_method, aql, item_order)
    record_change("규격 항목 수정", "spec_item", spec_id, f"{material_no} / 항목 {item_name}: {spec_display}")
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True})
    flash(f"항목 {item_name or spec_id} 수정됐어.")
    return redirect(url_for("spec_detail", material_no=material_no))


@app.route("/spec/<material_no>/item/<int:spec_id>/delete", methods=["POST"])
@perm_required("material_edit")
def spec_item_delete(material_no, spec_id):
    item = db.get_spec_item(spec_id)
    db.delete_spec_item(spec_id)
    detail = f"{material_no} / 항목 {item['item_name']}: {item['spec_display']}" if item else material_no
    record_change("규격 항목 삭제", "spec_item", spec_id, detail)
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True})
    flash("항목 삭제됐어.")
    return redirect(url_for("spec_detail", material_no=material_no))


@app.route("/spec/<material_no>/items/delete_bulk", methods=["POST"])
@perm_required("material_edit")
def spec_item_delete_bulk(material_no):
    spec_ids = [int(v) for v in request.form.getlist("spec_ids")]
    if not spec_ids:
        flash("삭제할 항목을 선택해줘.")
        return redirect(url_for("spec_detail", material_no=material_no))
    db.delete_spec_items_bulk(spec_ids)
    flash(f"{len(spec_ids)}개 항목 삭제됐어.")
    record_change("규격 항목 일괄 삭제", "spec_item", None, f"{material_no} / {len(spec_ids)}개")
    return redirect(url_for("spec_detail", material_no=material_no))


@app.route("/spec/<material_no>/item/add", methods=["POST"])
@perm_required("material_edit")
def spec_item_add(material_no):
    item_name, spec_display, judge_type, lower_limit, upper_limit, inspect_method, aql, item_order = \
        _parse_spec_form(request.form)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if not spec_display:
        if is_ajax:
            return jsonify({"ok": False, "error": "규격 표기를 입력해줘."}), 400
        flash("규격 표기를 입력해줘.")
        return redirect(url_for("spec_detail", material_no=material_no))

    existing = db.get_specs_by_material(material_no)
    material_name = existing[0]["material_name"] if existing else ""
    if not item_order:
        item_order = (max((s["item_order"] or 0) for s in existing) + 1) if existing else 1

    final_item_name = item_name or f"항목{item_order}"
    # 항목기호(item_name)가 자재 안에서 중복되면 성적서 조회 시 SQL 조인이 두 배로 불어나서
    # 항목이 겹쳐 보이거나 판정이 엉뚱하게 섞이는 사고가 실제로 있었다(2026-08-30 발견).
    if any((s["item_name"] or "") == final_item_name for s in existing):
        msg = f"항목기호 '{final_item_name}'는 이미 이 자재에 있어. 다른 기호를 써줘."
        if is_ajax:
            return jsonify({"ok": False, "error": msg}), 400
        flash(msg)
        return redirect(url_for("spec_detail", material_no=material_no))
    spec_id = db.add_spec(material_no, material_name, final_item_name, spec_display,
                judge_type, lower_limit, upper_limit, inspect_method, aql, item_order)
    record_change("규격 항목 추가", "spec_item", None, f"{material_no} / 항목 {item_name or item_order}: {spec_display}")
    if is_ajax:
        return jsonify({"ok": True, "spec": {
            "id": spec_id, "item_order": item_order, "item_name": final_item_name,
            "spec_display": spec_display, "judge_type": judge_type,
            "lower_limit": lower_limit, "upper_limit": upper_limit,
            "inspect_method": inspect_method, "aql": aql,
        }})
    flash("새 항목이 추가됐어.")
    return redirect(url_for("spec_detail", material_no=material_no))


@app.route("/spec/import", methods=["GET", "POST"])
@perm_required("material_import")
def spec_import():
    """빈 성적서 양식(자재별 규격만 채워진 파일) 여러 개를 한 번에 업로드해서 규격표에 등록."""
    if request.method == "POST":
        files = request.files.getlist("spec_files")
        files = [f for f in files if f and f.filename]
        if not files:
            flash("업로드할 파일을 선택해줘.")
            return redirect(url_for("spec_import"))

        results = []   # 등록 성공: [{"material_no","material_name","item_count","filename"}]
        failures = []  # 등록 실패: [{"filename","reason"}]
        item_warnings = []  # 등록은 됐지만 항목별로 확인이 필요한 것들

        for f in files:
            if not f.filename.lower().endswith((".xlsx", ".xlsm")):
                failures.append({"filename": f.filename, "reason": ".xlsx/.xlsm 파일이 아니야."})
                continue
            try:
                sheet_results = spec_import_module.parse_spec_file(
                    io.BytesIO(f.read()), source_name=f.filename
                )
            except Exception as e:
                failures.append({"filename": f.filename, "reason": f"파일을 읽는 중 오류: {e}"})
                continue

            for material_no, material_name, items, warnings, fail_reason in sheet_results:
                if fail_reason:
                    failures.append({"filename": f.filename, "reason": fail_reason})
                    continue

                item_warnings.extend(warnings)
                db.replace_specs_for_material(material_no, material_name, items)
                results.append({
                    "material_no": material_no,
                    "material_name": material_name,
                    "item_count": len(items),
                    "filename": f.filename,
                })

        if results:
            material_list = ", ".join(r["material_no"] for r in results[:10])
            if len(results) > 10:
                material_list += " 외"
            record_change("자재 일괄 등록", "spec", None, f"{len(results)}개 자재 ({material_list})")

        return render_template("spec_import_result.html",
                               results=results, failures=failures, warnings=item_warnings)

    return render_template("spec_import.html")


# ---------- 검사 입력 ----------

@app.route("/inspect/new")
@perm_required("inspect_input")
def inspect_select():
    """입고 리스트 중 미검사(대기) 건만 표로 보여줌"""
    query = request.args.get("q", "").strip()
    pending = db.search_intake(query, status="대기") if query else db.list_intake(status="대기")
    mats = db.get_materials()
    registered = {m["material_no"] for m in mats}
    name_map = {m["material_no"]: m["material_name"] for m in mats}
    intake_ids = [r["id"] for r in pending]
    progress_map = db.get_progress_by_intake_ids(intake_ids)
    drawing_materials = materials_with_drawings(r["material_no"] for r in pending)
    return render_template("inspect_select.html", pending=pending, query=query,
                           registered=registered, group_nos=set(),
                           progress_map=progress_map, name_map=name_map,
                           drawing_materials=drawing_materials)


@app.route("/inspect/delete_bulk", methods=["POST"])
@perm_required("inspect_input", "intake")
def inspect_delete_bulk():
    """검사 대기 목록에서 선택한 건들을 삭제(아직 검사 안 한 건만)."""
    intake_ids = [int(v) for v in request.form.getlist("intake_ids")]
    if not intake_ids:
        flash("삭제할 항목을 선택해줘.")
        return redirect(url_for("inspect_select"))
    deleted = db.delete_intake_bulk(intake_ids)
    flash(f"{deleted}건 삭제됐어." if deleted else "삭제된 항목이 없어(이미 검사된 건은 지울 수 없어).")
    record_change("검사 대기 목록 삭제", "intake", None, f"{deleted}건")
    return redirect(url_for("inspect_select"))


def _get_specs_for_material(material_no):
    """자재의 규격을 반환. (자재 그룹 기능은 제거됨 — 항상 단일 자재로 처리)
    반환 형태는 기존 호출부 호환을 위해 (specs_list, is_group=False, group_name=None) 유지.
    """
    return db.get_specs_by_material(material_no), False, None


@app.route("/inspect/start/<int:intake_id>", methods=["GET", "POST"])
@perm_required("inspect_input")
def inspect_form(intake_id):
    intake_row = db.get_intake(intake_id)
    if intake_row is None:
        flash("존재하지 않는 입고 건이야.")
        return redirect(url_for("inspect_select"))

    # 같은 입고 건에 성적서가 두 개 생기면 대시보드에서 같은 로트 수량이 중복 집계된다.
    # (등록 버튼 연타·새로고침으로 1초 간격 3건이 생긴 사례가 있었음)
    # 고치려면 성적서 상세의 '수정'이나 '재검사'를 써야 한다.
    existing = db.active_inspection_for_intake(intake_id)
    if existing is not None:
        flash("이 입고 건은 이미 검사가 등록돼 있어. 내용을 고치려면 성적서에서 '수정'을 눌러줘.")
        return redirect(url_for("inspection_detail", inspection_id=existing["id"]))

    material_no = intake_row["material_no"]
    specs, is_group, group_name = _get_specs_for_material(material_no)
    if not specs:
        flash(f"{material_no}에 등록된 자재 규격이 없어. 먼저 자재를 등록해줘.")
        return redirect(url_for("spec_list"))

    # 항목별 샘플수량 미리 계산 (입고수량 기준) — 조립품이면 부품마다 자기 AQL로 각자 계산
    # no_limit: 숫자 판정인데 하한·상한이 둘 다 비어 있는 항목.
    #   judge_numeric()이 이런 항목을 무조건 불합격으로 처리하므로, 측정 전에 검사자에게 알려준다.
    specs_with_sample = build_specs_with_sample(specs, intake_row["quantity"])

    full_inspect_config = db.get_full_inspect_config(material_no)

    if request.method == "POST":
        header = {
            "material_no": material_no,
            "material_name": group_name if is_group else _resolve_material_name(material_no, specs),
            "supplier": intake_row["supplier"],
            "po_number": intake_row["po_number"],
            "receive_date": intake_row["receive_date"],
            "inspect_date": request.form.get("inspect_date"),
            "inspector": g.user["display_name"] or g.user["username"],
            "quantity": intake_row["quantity"],
        }

        items_with_results = []
        overall_ok = True
        fi_units_to_save = None   # 전수 모드일 때만 채움

        if full_inspect_config:
            # ── 전수검사 모드 ── 유닛별 값에서 항목별 min/max/판정을 자동 산출
            qty = int(intake_row["quantity"] or 0)
            units_data = []
            for u in range(1, qty + 1):
                serial = request.form.get(f"unit_{u}_serial", "").strip()
                remark = request.form.get(f"unit_{u}_remark", "").strip()
                vals = {}
                for spec in specs:
                    v = request.form.get(f"unit_{u}_{spec['id']}", "").strip()
                    vals[spec["item_name"]] = v
                units_data.append({"unit_no": u, "serial_no": serial,
                                   "remark": remark, "values": vals})

            _bad_pf = {"NG", "X", "×", "△", "불합격", "FAIL"}

            def _unit_result(unit_vals):
                """유닛 단위 판정: 모든 항목 값 통과면 OK, 하나라도 이탈이면 NG."""
                has_any = False
                for sp in specs:
                    v = unit_vals.get(sp["item_name"], "").strip()
                    if not v:
                        continue
                    has_any = True
                    if sp["judge_type"] == "numeric":
                        try:
                            n = float(v.replace(",", ""))
                        except ValueError:
                            return "NG"
                        lo, hi = sp["lower_limit"], sp["upper_limit"]
                        if lo is not None and n < lo: return "NG"
                        if hi is not None and n > hi: return "NG"
                    else:
                        if v.upper() in _bad_pf: return "NG"
                return "OK" if has_any else ""

            for spec in specs:
                cell_vals = [u["values"].get(spec["item_name"], "").strip() for u in units_data]
                non_empty = [v for v in cell_vals if v]

                if spec["judge_type"] == "numeric":
                    nums = []
                    for v in non_empty:
                        try: nums.append(float(v.replace(",", "")))
                        except ValueError: pass
                    lo, hi = spec["lower_limit"], spec["upper_limit"]
                    if not non_empty:
                        min_v = max_v = None; result = "미측정"
                    elif lo is None and hi is None:
                        min_v = min(nums) if nums else None
                        max_v = max(nums) if nums else None
                        result = "규격미입력"
                    else:
                        min_v = min(nums) if nums else None
                        max_v = max(nums) if nums else None
                        result = "합격"
                        for n in nums:
                            if lo is not None and n < lo: result = "불합격"; break
                            if hi is not None and n > hi: result = "불합격"; break
                        if len(nums) != len(non_empty):   # 숫자 파싱 실패한 값 있음
                            result = "불합격"
                else:
                    min_v = max_v = None
                    has_bad = any(v.upper() in _bad_pf for v in non_empty)
                    result = "불합격" if has_bad else ("합격" if non_empty else "미측정")

                if result != "합격":
                    overall_ok = False

                if len(non_empty) <= 6:
                    raw_value = ",".join(non_empty)
                elif min_v is not None and max_v is not None:
                    raw_value = f"전수 {len(non_empty)}개 (범위 {min_v}~{max_v})"
                else:
                    raw_value = f"전수 {len(non_empty)}개 측정"

                gauge_expiry = request.form.get(f"item_{spec['id']}_gauge_expiry", "").strip() or None
                gauge_name = request.form.get(f"item_{spec['id']}_gauge_name", "").strip() or None

                items_with_results.append({
                    "item_name": spec["item_name"],
                    "measured_value": raw_value,
                    "max_value": max_v, "min_value": min_v, "result": result,
                    "gauge_expiry": gauge_expiry, "gauge_name": gauge_name,
                    "part_material_no": spec["material_no"],
                })

            fi_units_to_save = [{
                "unit_no": u["unit_no"], "serial_no": u["serial_no"],
                "values": u["values"], "result": _unit_result(u["values"]),
                "remark": u["remark"],
            } for u in units_data]

        else:
            # ── 일반 검사 모드 ── (기존 로직)
            for spec in specs_with_sample:
                allowed = aql_ac_allowance(spec["aql"], spec["sample_qty"])
                if spec["judge_type"] == "numeric":
                    vals = []
                    for i in range(1, 7):
                        v = request.form.get(f"item_{spec['id']}_{i}", "").strip()
                        if v:
                            vals.append(v)
                    raw_value = ",".join(vals)
                    result, max_v, min_v = judge_numeric(raw_value, spec["lower_limit"], spec["upper_limit"], allowed)
                else:
                    # judge_type == 'ok_ng'/'visual'/'numeric_pair' — 화면(inspect_form.html)이
                    # numeric_pair 전용 입력을 아직 지원하지 않아서(2채널 UI 미구현) 지금은
                    # O/X 입력으로 받는다. numeric_pair UI를 따로 만들기 전까진 이 경로 유지.
                    vals = [request.form.get(f"item_{spec['id']}_{i}", "").strip() for i in range(1, 7)]
                    raw_value = ",".join(v for v in vals if v)
                    result, max_v, min_v = judge_visual(raw_value, allowed)

                if result != "합격":
                    overall_ok = False

                gauge_expiry = request.form.get(f"item_{spec['id']}_gauge_expiry", "").strip() or None
                gauge_name = request.form.get(f"item_{spec['id']}_gauge_name", "").strip() or None

                items_with_results.append({
                    "item_name": spec["item_name"],
                    "measured_value": raw_value,
                    "max_value": max_v, "min_value": min_v, "result": result,
                    "gauge_expiry": gauge_expiry, "gauge_name": gauge_name,
                    "part_material_no": spec["material_no"],
                })

        overall_result = "합격" if overall_ok else "검토필요"
        actual_time_sec = request.form.get("actual_time_sec", "").strip()
        actual_time_sec = int(actual_time_sec) if actual_time_sec.isdigit() else 0
        est_time_label = format_duration(actual_time_sec)
        total_time_sec = compute_total_time_sec(specs_with_sample, actual_time_sec) if actual_time_sec else 0
        inspection_id = db.create_inspection(header, items_with_results, overall_result,
                                              intake_id=intake_id, est_time_label=est_time_label,
                                              actual_time_sec=actual_time_sec,
                                              total_time_sec=total_time_sec,
                                              created_by_user_id=g.user["id"])
        record_change("성적서 등록", "inspection", inspection_id,
                      f"자재 {material_no}, 업체 {intake_row['supplier']}, 판정 {overall_result}")

        remark_inspector = (request.form.get("remark_inspector") or "").strip()
        if remark_inspector:
            db.update_inspection_remark(inspection_id, "inspector", remark_inspector)

        # 전수 데이터도 같이 저장 (전수 모드일 때)
        if fi_units_to_save is not None:
            db.get_or_create_full_inspection(inspection_id)
            db.update_full_inspection(inspection_id,
                                      inspect_date=request.form.get("inspect_date") or None)
            db.save_full_inspection_units(inspection_id, fi_units_to_save)

        db.clear_inspection_progress(intake_id)
        db.delete_inspection_draft(intake_id)
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    prior_defect_count = db.get_defect_count_for(intake_row["supplier"], material_no)
    inspector_name = g.user["display_name"] or g.user["username"]
    db.register_inspector(intake_id, inspector_name)
    gauges = db.list_gauges()

    # 서버에 저장된 임시저장(다른 기기에서 입력하던 것 포함)
    draft_row = db.get_inspection_draft(intake_id)
    server_draft = draft_row["payload"] if draft_row else None
    draft_info = None
    if draft_row:
        draft_info = {"username": draft_row["username"], "updated_at": draft_row["updated_at"]}

    # 이 업체·자재에 최근 4M 변경이 있었으면 검사 전에 알려준다
    change_points = db.recent_change_points_for(intake_row["supplier"], material_no)

    full_inspect_config = db.get_full_inspect_config(material_no)
    return render_template("inspect_form.html", intake=intake_row, specs=specs_with_sample,
                           is_group=is_group, group_name=group_name,
                           prior_defect_count=prior_defect_count, gauges=gauges,
                           sel_gauge=build_sel_gauge(specs_with_sample, gauges),
                           server_draft=server_draft, draft_info=draft_info,
                           change_points=change_points,
                           full_inspect_config=full_inspect_config,
                           gauge_master_empty=(len(gauges) == 0))


@app.route("/inspect/draft/<int:intake_id>", methods=["POST"])
@perm_required("inspect_input")
def inspect_draft_save(intake_id):
    """검사 입력 중간값을 서버에 저장 (검사 화면에서 입력할 때마다 자동 호출).
    브라우저 localStorage만 쓰면 태블릿이 꺼지거나 기기를 바꿀 때 날아가서 서버에도 남긴다."""
    import json as _json
    payload = request.get_data(as_text=True) or ""
    if len(payload) > 200_000:          # 비정상적으로 큰 요청 차단
        return {"ok": False, "error": "too_large"}, 413
    try:
        _json.loads(payload)            # JSON 형식인지만 확인 (내용은 화면이 만든 그대로 보관)
    except ValueError:
        return {"ok": False, "error": "bad_json"}, 400

    db.save_inspection_draft(intake_id, payload,
                             user_id=g.user["id"],
                             username=g.user["display_name"] or g.user["username"])
    return {"ok": True}


@app.route("/inspect/draft/<int:intake_id>/delete", methods=["POST"])
@perm_required("inspect_input")
def inspect_draft_delete(intake_id):
    db.delete_inspection_draft(intake_id)
    return {"ok": True}


@app.route("/inspect/draft/<int:intake_id>/status")
@perm_required("inspect_input")
def inspect_draft_status(intake_id):
    """다른 계정이 같은 검사 건 임시저장을 갱신했는지 가볍게 확인하기 위한 폴링용 엔드포인트.
    실시간 동기화(자동 반영)는 아니고, 화면에 '새로고침 권유' 배너를 띄우는 용도라 payload는 안 돌려준다."""
    draft_row = db.get_inspection_draft(intake_id)
    if not draft_row:
        return {"exists": False}
    return {"exists": True, "user_id": draft_row["user_id"], "username": draft_row["username"],
            "updated_at": draft_row["updated_at"]}


NO_SPEC_RESULT = "규격미입력"   # 판정 기준 자체가 없어서 합격/불합격을 말할 수 없는 상태


def judge_numeric(raw_value, lower, upper, allowed_defects=0):
    """
    raw_value: 콤마로 구분된 여러 측정값 가능 (예: "12.1,12.3,12.0")
    lower/upper 둘 다 있으면 범위 판정, 하나만 있으면("OO 이상"/"OO 이하") 그쪽만 판정.
    allowed_defects: 규격을 벗어난 값이 몇 개까지는 있어도 합격으로 보는지(Ac, aql_ac_allowance()
        참고). 기본값 0 = 예전처럼 무결점.

    둘 다 없으면(규격 미확정) '규격미입력'을 돌려준다 — 불합격이 아니다.
    규격이 안 채워진 건 제품 문제가 아니라 우리 데이터 문제인데, 불합격으로 처리하면
    부적합 통보서가 협력사로 나가버린다. 합격도 아니므로 성적서 전체는 '검토필요'가 되고
    승인 단계에서 사람이 먼저 규격을 채우도록 막는다.

    반환: (판정, 최대값, 최소값)
    """
    if not raw_value:
        return "미측정", None, None
    try:
        values = [float(v.strip()) for v in raw_value.split(",") if v.strip() != ""]
    except ValueError:
        return "입력오류", None, None

    if not values:
        return "미측정", None, None

    max_v, min_v = max(values), min(values)

    if lower is None and upper is None:
        return NO_SPEC_RESULT, max_v, min_v

    def _within(v):
        # 하한·상한 중 하나만 있는 단측 표기("5 이상")도 그쪽만 본다
        if lower is not None and v < lower:
            return False
        if upper is not None and v > upper:
            return False
        return True

    defect_count = sum(1 for v in values if not _within(v))
    ok = defect_count <= allowed_defects
    return ("합격" if ok else "불합격"), max_v, min_v


def judge_numeric_pair(raw_value, lower, upper, allowed_defects=0):
    """2채널 숫자 항목(예: 하우징 고저항 CT/ROD).

    저장 형식: "a/b,c/d,e/f" — 각 샘플이 채널1/채널2 두 값.
    양쪽 채널 모든 값이 lower~upper 범위 안이어야 합격 — allowed_defects(Ac)개까지는
    벗어나도 합격으로 본다(기본값 0 = 무결점). max/min은 두 채널 통틀어 계산해서
    성적서 max/min 칸에 그대로 쓴다.
    """
    if not raw_value:
        return "미측정", None, None
    try:
        values = []
        for chunk in raw_value.split(","):
            chunk = chunk.strip()
            if not chunk:
                continue
            parts = [p.strip() for p in chunk.split("/") if p.strip() != ""]
            if not parts:
                continue
            for p in parts:
                values.append(float(p))
    except ValueError:
        return "입력오류", None, None

    if not values:
        return "미측정", None, None

    max_v, min_v = max(values), min(values)

    if lower is None and upper is None:
        return NO_SPEC_RESULT, max_v, min_v

    def _within(v):
        if lower is not None and v < lower:
            return False
        if upper is not None and v > upper:
            return False
        return True

    defect_count = sum(1 for v in values if not _within(v))
    ok = defect_count <= allowed_defects
    return ("합격" if ok else "불합격"), max_v, min_v


def judge_visual(raw_value, allowed_defects=0):
    """육안 항목: O(합격)/X(불합격) 여러 칸 입력, 콤마로 구분.
    allowed_defects(Ac)개까지는 X가 있어도 합격으로 본다(기본값 0 = 무결점)."""
    if not raw_value:
        return "미측정", None, None
    vals = [v.strip().upper() for v in raw_value.split(",") if v.strip()]
    if not vals:
        return "미측정", None, None
    defect_count = sum(1 for v in vals if v != "O")
    ok = defect_count <= allowed_defects
    return ("합격" if ok else "불합격"), None, None


_BAD_MARKERS = {"X", "NG", "×", "불합격", "FAIL"}


def _out_of_spec_flags(measured_value, judge_type, lower, upper):
    """측정값 문자열(콤마 구분)을 낱개로 쪼개서, 각 값이 실제로 규격을 벗어났는지 배열로
    돌려준다. 항목 전체가 불합격이어도 샘플 6개 중 실제로 벗어난 건 1개뿐일 수 있는데,
    예전엔 항목이 불합격이면 값 전부를 빨갛게 칠해서 어느 값이 문제인지 구분이 안 됐다
    (2026-08-31 사용자 피드백). numeric은 하한/상한과 직접 비교, 그 외(visual/ok_ng)는
    X/NG 같은 불량 표기 자체를 벗어난 값으로 본다."""
    if not measured_value:
        return []
    vals = [v.strip() for v in measured_value.split(",")]
    if judge_type == "numeric":
        flags = []
        for v in vals:
            try:
                n = float(v)
            except ValueError:
                flags.append(True)   # 숫자로 못 읽는 값도 눈에 띄게 강조
                continue
            bad = (lower is not None and n < lower) or (upper is not None and n > upper)
            flags.append(bad)
        return flags
    return [v.upper() in _BAD_MARKERS for v in vals]


# ---------- 성적서 상세 ----------

@app.route("/inspection/<int:inspection_id>")
@login_required
def inspection_detail(inspection_id):
    header, items = db.get_inspection(inspection_id)
    if header is None:
        flash("존재하지 않는 성적서야.")
        return redirect(url_for("home"))

    # 개당 측정시간 = 스톱워치로 잰 "샘플 1개(1사이클)" 시간 그대로
    per_cycle_sec = header["actual_time_sec"] or 0
    per_cycle_label = format_duration(per_cycle_sec)

    total_time_label = "-"
    specs, _, _ = _get_specs_for_material(header["material_no"])
    if specs and per_cycle_sec:
        specs_with_sample = build_specs_with_sample(specs, header["quantity"])
        total_time_label = format_duration(compute_total_time_sec(specs_with_sample, per_cycle_sec))

    # 계측기 유효기간이 한 달(30일) 이내로 남았거나 이미 지난 항목들을 항목별로 나열
    gauge_alerts = []
    today = _dt.now().date()
    for it in items:
        expiry_str = it["gauge_expiry"]
        if not expiry_str:
            continue
        try:
            expiry_date = _dt.strptime(expiry_str, "%Y-%m-%d").date()
        except ValueError:
            continue
        days_left = (expiry_date - today).days
        if days_left <= 30:
            gauge_alerts.append({
                "item_name": it["item_name"],
                "gauge_name": it["gauge_name"],
                "days_left": days_left,
            })

    # AQL·샘플수량·Ac 매핑 (specs → item_name 기준)
    specs_all = build_specs_with_sample(specs, header["quantity"]) if specs else []
    specs_map = {s["item_name"]: s for s in specs_all}

    # 항목이 불합격이어도 샘플 중 실제로 규격을 벗어난 값만 강조하기 위한 개별 값 플래그
    # (item id → [bool, ...], measured_value를 콤마로 쪼갠 순서와 동일)
    value_flags = {
        it["id"]: _out_of_spec_flags(it["measured_value"], it["judge_type"], it["lower_limit"], it["upper_limit"])
        for it in items
    }

    # AQL 그룹별 샘플수 요약 (대시보드·설명용)
    aql_groups = _unique_aql_sample_qty(specs_all) if specs_all else {}

    # 판정 대기(pending) 중에 자재 규격이 수정되면, 화면의 규격 표기/하한상한은
    # specs 테이블과 실시간 조인이라 바로 바뀌어 보이는데, 이미 저장된 항목별
    # 판정(result/max/min)은 측정 당시 값 그대로 얼어붙어 있어서 안 바뀐다.
    # "규격 수정했는데 반영이 안 된다"는 혼동의 원인 — 여기서 미리 감지해서 안내한다.
    # (승인 이후엔 감사기록 보존을 위해 절대 건드리지 않고, 이 체크도 안 한다)
    stale_spec_items = []
    if header["status"] == "pending":
        for it in items:
            spec_row = specs_map.get(it["item_name"])
            if not spec_row or spec_row["judge_type"] != "numeric":
                continue
            allowed = aql_ac_allowance(spec_row.get("aql"), spec_row.get("sample_qty"))
            fresh_result, _mx, _mn = judge_numeric(
                it["measured_value"], it["lower_limit"], it["upper_limit"], allowed)
            if fresh_result != it["result"]:
                stale_spec_items.append(it["item_name"])

    existing_ncr_list = db.list_ncr(inspection_id=inspection_id)
    return_requests = db.get_return_requests_by_inspection(inspection_id)
    prior_defect_count = db.get_defect_count_for(header["supplier"], header["material_no"])
    is_failed = header["overall_result"] not in ("합격", "", None)
    # 전수검사 — 자재에 설정이 있으면 현황 같이 넘긴다
    full_inspect_config = db.get_full_inspect_config(header["material_no"])
    full_inspect = db.get_full_inspection(inspection_id) if full_inspect_config else None
    if full_inspect:
        fi_units = db.list_full_inspection_units(inspection_id)
        full_inspect["ok_cnt"] = sum(1 for u in fi_units if u.get("result") == "OK")
        full_inspect["ng_cnt"] = sum(1 for u in fi_units if u.get("result") == "NG")
    return render_template("inspection_detail.html", header=header, items=items,
                           total_time_label=total_time_label, per_cycle_label=per_cycle_label,
                           aql_groups=aql_groups,
                           gauge_alerts=gauge_alerts, existing_ncr_list=existing_ncr_list,
                           return_requests=return_requests,
                           prior_defect_count=prior_defect_count,
                           specs_map=specs_map, is_failed=is_failed,
                           full_inspect_config=full_inspect_config,
                           full_inspect=full_inspect,
                           stale_spec_items=stale_spec_items,
                           value_flags=value_flags)


@app.route("/inspection/<int:inspection_id>/remark", methods=["POST"])
@perm_required("inspect_input", "inspect_edit_all", "approve")
def inspection_remark(inspection_id):
    """비고란 저장 — 화면에서 어느 칸(검사자/중간관리자/최종결정권자)인지 넘겨받고, 해당 권한이 있는지 확인."""
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        flash("존재하지 않는 성적서야.")
        return redirect(url_for("home"))

    # 검사자란=inspect_input, 중간관리자란=inspect_edit_all, 최종결정권자란=approve 권한 보유자만 그 칸에 쓸 수 있음
    REMARK_PERM_MAP = {"inspector": "inspect_input", "manager": "inspect_edit_all", "approver": "approve"}
    role_key = request.form.get("target_role", "")
    needed_perm = REMARK_PERM_MAP.get(role_key)
    user_perms = _user_perms(g.user)
    if role_key not in db.REMARK_FIELDS or needed_perm not in user_perms:
        flash("이 비고란에 쓸 권한이 없어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    text = request.form.get("remark_text", "").strip()
    name = g.user["display_name"] or g.user["username"]
    combined = f"{name}: {text}" if text else ""
    db.update_inspection_remark(inspection_id, role_key, combined)
    flash("비고가 저장됐어.")
    record_change("비고 작성", "inspection", inspection_id, f"{role_key} 비고: {text[:50]}")
    return_to = request.form.get("return_to", "")
    if return_to == "approve":
        return redirect(url_for("approve_view", inspection_id=inspection_id))
    return redirect(url_for("inspection_detail", inspection_id=inspection_id))


# ---------- 측정값 수정 (pending 상태만) ----------

def _can_edit_inspection(header):
    """inspect_edit_all 권한이 있으면 전체 수정 가능, 없으면(inspect_input만) 본인이 만든 성적서만."""
    user_perms = _user_perms(g.user)
    if "inspect_edit_all" in user_perms:
        return True
    if "inspect_input" in user_perms:
        return header["created_by_user_id"] == g.user["id"]
    return False


@app.route("/inspection/<int:inspection_id>/edit")
@perm_required("inspect_input")
def edit_inspection(inspection_id):
    header, items = db.get_inspection(inspection_id)
    if header is None or header["status"] != "pending":
        flash("수정할 수 없는 상태야.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))
    if not _can_edit_inspection(header):
        flash("본인이 입력한 성적서만 수정할 수 있어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    intake_row = db.get_intake(header["intake_id"]) if header["intake_id"] else None
    material_no = header["material_no"]
    specs, is_group, group_name = _get_specs_for_material(material_no)
    specs_with_sample = build_specs_with_sample(specs, header["quantity"])

    # 기존 측정값 prefill — 그룹 검사는 item_name만으로는 부품 간 중복될 수 있어 spec.id(전역 고유)로 맞춤
    prefill = {
        "inspect_date": header["inspect_date"] or "",
        "inspector": header["inspector"] or "",
        "remark_inspector": header["remark_inspector"] or "",
        "actual_time_sec": header["actual_time_sec"] or 0,
    }
    items_by_key = {(it["part_material_no"], it["item_name"]): it for it in items}
    for s in specs:
        it = items_by_key.get((s["material_no"], s["item_name"]))
        if it is None:
            continue
        vals = (it["measured_value"] or "").split(",")
        for i, v in enumerate(vals, start=1):
            prefill[f"item_{s['id']}_{i}"] = v.strip()
        prefill[f"item_{s['id']}_gauge_expiry"] = it["gauge_expiry"] or ""
        prefill[f"item_{s['id']}_gauge_name"] = it["gauge_name"] or ""

    # intake_row가 없으면 header 정보로 임시 구성
    if intake_row is None:
        class FakeIntake:
            pass
        fake = FakeIntake()
        fake.__class__.__getitem__ = lambda s, k: getattr(s, k, None)
        intake_row = {
            "id": header["intake_id"],
            "material_no": header["material_no"],
            "supplier": header["supplier"],
            "po_number": header["po_number"],
            "receive_date": header["receive_date"],
            "quantity": header["quantity"],
        }

    prior_defect_count = db.get_defect_count_for(intake_row["supplier"], header["material_no"])
    gauges = db.list_gauges()
    return render_template("inspect_form.html",
                           intake=intake_row, specs=specs_with_sample,
                           prefill=prefill, edit_mode=inspection_id,
                           is_group=is_group, group_name=group_name,
                           gauges=gauges,
                           sel_gauge=build_sel_gauge(specs_with_sample, gauges, prefill),
                           gauge_master_empty=(len(gauges) == 0),
                           prior_defect_count=prior_defect_count)


@app.route("/inspection/<int:inspection_id>/edit", methods=["POST"])
@perm_required("inspect_input")
def edit_inspection_submit(inspection_id):
    header, _ = db.get_inspection(inspection_id)
    if header is None or header["status"] != "pending":
        flash("수정할 수 없는 상태야.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))
    if not _can_edit_inspection(header):
        flash("본인이 입력한 성적서만 수정할 수 있어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    specs, _, _ = _get_specs_for_material(header["material_no"])
    specs_with_sample = build_specs_with_sample(specs, header["quantity"])
    items_with_results = []
    overall_ok = True

    for spec in specs_with_sample:
        allowed = aql_ac_allowance(spec["aql"], spec["sample_qty"])
        if spec["judge_type"] == "numeric":
            vals = [request.form.get(f"item_{spec['id']}_{i}", "").strip() for i in range(1, 7)]
            raw_value = ",".join(v for v in vals if v)
            result, max_v, min_v = judge_numeric(raw_value, spec["lower_limit"], spec["upper_limit"], allowed)
        else:
            vals = [request.form.get(f"item_{spec['id']}_{i}", "").strip() for i in range(1, 7)]
            raw_value = ",".join(v for v in vals if v)
            result, max_v, min_v = judge_visual(raw_value, allowed)

        if result != "합격":
            overall_ok = False
        gauge_expiry = request.form.get(f"item_{spec['id']}_gauge_expiry", "").strip() or None
        gauge_name = request.form.get(f"item_{spec['id']}_gauge_name", "").strip() or None
        items_with_results.append({
            "item_name": spec["item_name"],
            "measured_value": raw_value,
            "max_value": max_v,
            "min_value": min_v,
            "result": result,
            "gauge_expiry": gauge_expiry,
            "gauge_name": gauge_name,
            "part_material_no": spec["material_no"],
        })

    overall_result = "합격" if overall_ok else "검토필요"
    actual_time_sec = request.form.get("actual_time_sec", "").strip()
    actual_time_sec = int(actual_time_sec) if actual_time_sec.isdigit() else None
    est_time_label = format_duration(actual_time_sec) if actual_time_sec is not None else None
    qty = header["quantity"]
    specs_with_sample = build_specs_with_sample(specs, qty)
    total_time_sec = compute_total_time_sec(specs_with_sample, actual_time_sec) if actual_time_sec else None
    db.update_inspection_items(
        inspection_id,
        request.form.get("inspect_date"),
        g.user["display_name"] or g.user["username"],
        items_with_results,
        overall_result,
        est_time_label=est_time_label,
        actual_time_sec=actual_time_sec,
        total_time_sec=total_time_sec,
    )
    remark_inspector = (request.form.get("remark_inspector") or "").strip()
    db.update_inspection_remark(inspection_id, "inspector", remark_inspector)

    flash("측정값이 수정됐어.")
    record_change("성적서 수정", "inspection", inspection_id,
                  f"자재 {header['material_no']}, 판정 {overall_result}")
    return redirect(url_for("inspection_detail", inspection_id=inspection_id))


# ---------- 승인 / 반려 ----------

def _save_signature(inspection_id, signature_data_url):
    """캔버스 서명(base64 dataURL)을 PNG 파일로 저장. 반환: (경로, 에러메시지)"""
    if not signature_data_url or "," not in signature_data_url:
        return None, "서명 데이터가 비어 있어."
    try:
        header_part, b64_part = signature_data_url.split(",", 1)
        png_bytes = base64.b64decode(b64_part)
    except Exception as e:
        return None, f"서명 데이터 디코딩 실패: {e}"

    try:
        os.makedirs(SIGNATURE_DIR, exist_ok=True)
        path = os.path.join(SIGNATURE_DIR, f"{inspection_id}.png")
        with open(path, "wb") as f:
            f.write(png_bytes)
    except Exception as e:
        return None, f"서명 파일 저장 실패({SIGNATURE_DIR}): {e}"

    return path, None


def _save_signature_upload(inspection_id, file_storage):
    """업로드된 서명 이미지 파일(PNG/JPG)을 저장. 반환: (경로, 에러메시지)"""
    if file_storage is None or not file_storage.filename:
        return None, "업로드할 서명 이미지 파일을 선택해줘."

    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in (".png", ".jpg", ".jpeg"):
        return None, "PNG 또는 JPG 이미지 파일만 업로드할 수 있어."

    try:
        os.makedirs(SIGNATURE_DIR, exist_ok=True)
        path = os.path.join(SIGNATURE_DIR, f"{inspection_id}{ext}")
        file_storage.save(path)
    except Exception as e:
        return None, f"서명 파일 저장 실패({SIGNATURE_DIR}): {e}"

    return path, None


def _generate_report_files(inspection_id, approver, signature_path):
    """DB에 저장된 성적서 데이터를 report_builder 입력 형식으로 변환해 xlsx/PDF 생성.
    조립품 그룹 검사면 부품별로 시트를 나눈 통합 워크북을 만든다.
    자재에 커스텀(자유양식) 성적서가 지정돼 있으면 reportlab로 PDF만 직접 그린다.
    반환: (xlsx_path, pdf_path, error_message, is_custom) — error_message는 성공 시 None,
          is_custom=True면 기준서 페이지가 없는 커스텀 PDF이므로 호출부에서 병합 방식을 달리해야 함"""
    header, items = db.get_inspection(inspection_id)
    if header is None:
        return None, None, "성적서 정보를 찾을 수 없어.", False

    # ── 커스텀(자유양식) 성적서가 지정된 자재면 reportlab 경로로 분기 ──
    template_id = db.get_material_template_id(header["material_no"])
    if template_id:
        tmpl = db.get_custom_template(template_id)
        if tmpl is None:
            return None, None, "지정된 커스텀 양식을 찾을 수 없어(양식이 삭제됐을 수 있어).", False
        import custom_report
        data = {
            "fields": _custom_fields_from_header(header),
            "items": _custom_items_from_inspection(items),
            "signature_path": signature_path or header["signature_path"],
            "logo_path": report_builder.LOGO_PATH,
        }
        fname = report_builder.build_report_filename(
            header["supplier"], header["material_no"], header["material_name"] or "",
            inspect_date=header["inspect_date"])
        out_pdf = report_builder._dedupe_path(
            os.path.join(report_builder.report_output_dir(), fname + ".pdf"))
        try:
            path, err = custom_report.build_custom_report(tmpl, data, out_pdf)
        except Exception:
            import traceback
            return None, None, f"커스텀 성적서 생성 중 오류:\n{traceback.format_exc(limit=3)}", True
        if err or not path:
            return None, None, err or "커스텀 성적서 생성 실패", True
        return None, path, None, True   # xlsx 없음 · 커스텀 표시

    rb_header = {
        "vendor": header["supplier"],
        "po_no": header["po_number"],
        "lot": header["receive_date"],
        "inspect_date": header["inspect_date"],
        "inspector": header["inspector"],
        "qty": header["quantity"],
    }

    combined_specs = db.get_specs_by_material(header["material_no"])

    # 개당 측정시간 = 스톱워치로 잰 "샘플 1개(1사이클)" 시간 그대로
    per_cycle_sec = header["actual_time_sec"] or 0
    per_cycle_label = format_duration(per_cycle_sec)

    # 총 측정시간 = 기준(AQL4) 샘플수×개당시간 + 다른 AQL 그룹의 초과분×10초
    specs_with_sample = build_specs_with_sample(combined_specs, header["quantity"])
    total_time_label = format_duration(compute_total_time_sec(specs_with_sample, per_cycle_sec))

    def _build_result(it):
        raw = (it["measured_value"] or "")
        raw_tokens = [v.strip() for v in raw.split(",") if v.strip() != ""]
        values = []
        for v in raw_tokens:
            try:
                values.append(float(v))
            except ValueError:
                pass  # 숫자로 안 바뀌면(O/X 등) 아래에서 별도 처리
        # 육안/O,X 항목은 OK/NG 문자열로 변환해서 그대로 기입
        if not values and raw_tokens and all(t.upper() in ("O", "X") for t in raw_tokens):
            values = ["OK" if t.upper() == "O" else "NG" for t in raw_tokens]
        sample_qty = sample_size(it["aql"], header["quantity"]) if it["aql"] is not None else None
        ac_allowance = aql_ac_allowance(it["aql"], sample_qty) if it["aql"] is not None else None
        return {
            "item_name": it["item_name"],
            "values": values,
            "max": it["max_value"],
            "min": it["min_value"],
            "verdict": it["result"] or "",
            "gauge_expiry": it["gauge_expiry"],
            "spec_display": it["spec_display"],
            "aql": it["aql"],
            "sample_qty": sample_qty,
            "ac_allowance": ac_allowance,
            "inspect_method": it["inspect_method"],
            "lower_limit": it["lower_limit"],
            "upper_limit": it["upper_limit"],
        }

    remarks = {
        "inspector": header["remark_inspector"],
        "manager": header["remark_manager"],
        "approver": header["remark_approver"],
    }

    def _standard_info_for(material_no):
        m = db.get_material(material_no)
        if m is None:
            return None
        return {
            "drawing_version": m["drawing_version"],
            "revision_date": m["revision_date"],
            "edition": m["edition"],
            "unit": m["unit"],
        }

    try:
        results = [_build_result(it) for it in items]
        xlsx_path, pdf_path, pdf_error, signature_error = report_builder.build_report(
            header["material_no"], header["material_name"] or "",
            rb_header, results, header["overall_result"],
            approver=approver, signature_path=signature_path,
            per_cycle_label=per_cycle_label, total_time_label=total_time_label,
            remarks=remarks, approval_type=header["approval_type"],
            standard_info=_standard_info_for(header["material_no"]),
        )
    except Exception:
        import traceback
        return None, None, f"성적서 파일 생성 중 오류:\n{traceback.format_exc(limit=3)}", False

    error_msg = None
    if signature_error and pdf_error:
        error_msg = f"{signature_error} / {pdf_error}"
    elif signature_error:
        error_msg = signature_error
    elif pdf_error:
        error_msg = pdf_error

    return xlsx_path, pdf_path, error_msg, False


def compute_content_hash(inspection_id):
    """성적서 내용(헤더 + 항목별 측정값·판정)을 SHA-256으로 굳힌다.

    승인 시점에 저장해두면 나중에 DB 값이 바뀌었는지 검증할 수 있다.
    - 항목은 이름순으로 정렬해서 담는다 → 조회 순서가 달라져도 해시가 흔들리지 않게
    - PDF가 아니라 '판정 내용' 자체를 해시하므로, 성적서를 다시 출력해도 값은 그대로다
    """
    import hashlib, json as _json
    header, items = db.get_inspection(inspection_id)
    if header is None:
        return None
    payload = {
        "material_no":  header["material_no"],
        "material_name": header["material_name"],
        "supplier":     header["supplier"],
        "po_number":    header["po_number"],
        "receive_date": header["receive_date"],
        "inspect_date": header["inspect_date"],
        "inspector":    header["inspector"],
        "quantity":     header["quantity"],
        "overall_result": header["overall_result"],
        "items": sorted(
            [{"item": it["item_name"], "value": it["measured_value"], "result": it["result"]}
             for it in (items or [])],
            key=lambda d: str(d["item"]),
        ),
    }
    blob = _json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def compute_file_hash(path):
    """파일 내용의 SHA-256. 파일이 없으면 None."""
    import hashlib
    if not path or not os.path.exists(path):
        return None
    h = hashlib.sha256()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
    except OSError:
        return None
    return h.hexdigest()


def verify_inspection_integrity(inspection_id):
    """성적서가 승인 당시 그대로인지 검증.
    반환: {"checked": bool, "content_ok": bool|None, "pdf_ok": bool|None, "message": str}
    """
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        return {"checked": False, "content_ok": None, "pdf_ok": None, "message": "성적서를 찾을 수 없어."}

    stored_content = header["content_hash"] if "content_hash" in header.keys() else None
    stored_pdf = header["pdf_hash"] if "pdf_hash" in header.keys() else None

    if not stored_content:
        return {"checked": False, "content_ok": None, "pdf_ok": None,
                "message": "위변조 검증 기능이 생기기 전에 승인된 성적서라 검증 기준값이 없어."}

    content_ok = (compute_content_hash(inspection_id) == stored_content)
    pdf_ok = None
    if stored_pdf:
        actual = compute_file_hash(header["pdf_path"])
        pdf_ok = (actual == stored_pdf) if actual else False

    if content_ok and pdf_ok is not False:
        msg = "승인 당시 내용과 일치해."
    elif not content_ok:
        msg = "⚠️ 승인 이후 성적서 내용이 바뀌었어. 원인을 확인해줘."
    else:
        msg = "⚠️ 발행된 PDF 파일이 승인 당시와 달라. 파일이 교체됐을 수 있어."
    return {"checked": True, "content_ok": content_ok, "pdf_ok": pdf_ok, "message": msg}


def _final_decision_block_reason(inspection_id, action, reject_reason=None):
    """승인·특채·불합격 확정을 막아야 하는 이유가 있으면 그 문구를, 없으면 None을 돌려준다.

    단건 승인과 일괄 승인이 같은 규칙을 쓰도록 한 곳에 모아둔 검증.
    """
    if action not in ("approve", "special", "failed"):
        return None

    header, items = db.get_inspection(inspection_id)
    if header is None:
        return "성적서를 찾을 수 없어."
    items = items or []

    # 1) 판정 근거가 없는 항목이 있으면 어떤 결정도 내릴 수 없다.
    #    승인뿐 아니라 불합격도 막는다 — 규격이 비어서 못 잰 걸 불합격으로 확정하면
    #    우리 데이터 누락이 협력사 부적합 통보서로 나가버린다.
    no_spec = [it["item_name"] for it in items if it["result"] == NO_SPEC_RESULT]
    if no_spec:
        return (f"규격(하한·상한)이 비어 있는 항목이 있어서 결정할 수 없어: {', '.join(no_spec)} — "
                f"자재 규격을 먼저 채우고 재검사해줘.")

    # 2) 아직 측정 안 한 항목이 남아 있으면 결정 불가
    not_measured = [it["item_name"] for it in items if it["result"] in ("미측정", "입력오류")]
    if not_measured:
        return f"측정값이 비어 있거나 잘못된 항목이 있어: {', '.join(not_measured)} — 검사를 먼저 마무리해줘."

    # (예전엔 여기서 "측정 이후 규격이 바뀐 항목"을 감지해서 승인 자체를 막았는데,
    # 규격을 정정하는 일(오탈자 수정, 단위 보정 등)이 실제로는 흔히 있고 그때마다
    # 이미 대기 중인 다른 성적서들의 결정이 통째로 막혀버리는 부작용이 커서
    # 2026-09-01 제거했다 — 안내는 inspection_detail.html의 배너로만 하고, 최종 결정은
    # 사람이 판단해서 내리도록 둔다. 감지 로직 자체는 inspection_detail()의
    # stale_spec_items 계산에 남아있음.)

    # 3) 특채는 '규격을 벗어났지만 예외적으로 쓴다'는 결정이다.
    #    전 항목 합격인 성적서를 특채로 올리는 건 성립하지 않는다(그냥 합격 승인해야 함).
    if action == "special" and header["overall_result"] == "합격":
        return "전 항목 합격인 성적서는 특채 대상이 아니야. 그냥 '합격 승인'으로 처리해줘."

    # 3-1) 반대 방향도 막아야 한다 — 규격을 벗어난 항목이 실제로 있는데 '합격 승인'을
    #      누르면 불량이 정상 합격으로 둔갑해서 NCR·특채 절차 없이 그냥 통과해버린다.
    #      (2번 검사에서 발견 — 서버가 이 방향은 막지 않고 있었음)
    if action == "approve" and header["overall_result"] != "합격":
        return ("규격을 벗어난 항목이 있어서 '합격 승인'으로 처리할 수 없어. "
                "'특채' 또는 '불합격 확정'을 선택해줘.")

    # 4) 불합격 확정은 협력사로 나가는 결정이라 사유가 반드시 남아야 한다.
    #    (반려는 사유가 필수인데 더 중대한 불합격이 선택이던 건 앞뒤가 안 맞았다)
    if action == "failed" and not (reject_reason or "").strip():
        return "불합격 사유를 입력해줘. 부적합 통보서와 협력사 통보의 근거가 돼."

    return None


def _can_make_final_decision(user, what="승인·특채·불합격 확정"):
    """최종 결정을 내릴 수 있는 사람인지.

    최종결정권자가 아직 아무도 지정 안 됐으면(초기 상태) '승인' 권한만으로 통과시킨다 —
    지정하기 전에 시스템이 잠겨버리면 안 되기 때문. 한 명이라도 지정되는 순간부터는
    최종결정권자만 결정할 수 있다.

    what: 안내 문구에 넣을 행위 이름 (성적서 승인 / 부적합 통보서 확인 / 성적표 승인 …)
    반환: (가능여부, 안 되는 이유)
    """
    approvers = db.list_final_approvers()
    if not approvers or (user and user["is_final_approver"]):
        return True, None
    names = ", ".join((a["display_name"] or a["username"]) for a in approvers)
    return False, f"{what}은(는) 최종결정권자만 할 수 있어. (현재 최종결정권자: {names})"


@app.route("/inspection/<int:inspection_id>/approve", methods=["POST"])
@perm_required("approve")
def approve(inspection_id):
    action = request.form.get("action")          # 'approve' / 'special' / 'reject'
    approver = g.user["display_name"] or g.user["username"]
    reject_reason = request.form.get("reject_reason", "").strip()
    from_approve = request.form.get("from_approve") == "1"

    def _back(on_error=False):
        if from_approve and on_error:
            return redirect(url_for("approve_view", inspection_id=inspection_id))
        if from_approve:
            return redirect(url_for("approve_list"))
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    if not approver:
        flash("결정권자 이름을 입력해줘.")
        return _back(on_error=True)

    # 최종 결정(승인·특채·불합격 확정)은 최종결정권자만 할 수 있다
    if action in ("approve", "special", "failed"):
        allowed, why = _can_make_final_decision(g.user)
        if not allowed:
            flash(why)
            return _back(on_error=True)

    # 승인·특채·불합격 확정은 전부 최종 결정이므로 서명을 똑같이 요구한다
    blocked = _final_decision_block_reason(inspection_id, action, reject_reason)
    if blocked:
        flash(blocked)
        return _back(on_error=True)

    signature_path = None
    if action in ("approve", "special", "failed"):
        signature_method = request.form.get("signature_method", "draw")
        stamp_type = (request.form.get("signature_stamp_type") or "").strip()
        if signature_method == "upload":
            signature_file = request.files.get("signature_file")
            signature_path, sig_save_error = _save_signature_upload(inspection_id, signature_file)
            # 파일 업로드일 때는 도장/사인 종류 필수 — 안 고르면 진행 차단
            if not sig_save_error and stamp_type not in ("stamp", "sign"):
                flash("업로드한 이미지가 '도장'인지 '사인'인지 골라줘.")
                return _back(on_error=True)
        else:
            signature_data = request.form.get("signature_data", "").strip()
            if not signature_data:
                flash("서명을 입력해줘.")
                return _back(on_error=True)
            signature_path, sig_save_error = _save_signature(inspection_id, signature_data)
            # 직접 그린 서명은 사인 취급 (기본 크기)
            stamp_type = "sign"

        if sig_save_error:
            flash(f"서명 저장 실패: {sig_save_error}")
            return _back(on_error=True)

        # 도장/사인 종류를 사이드카 파일로 저장 → 출력 시 report_builder가 읽어서 크기 결정
        # 재승인 시 stale sidecar를 남기지 않도록 항상 지운 다음 필요할 때만 새로 씀
        if signature_path:
            sidecar = signature_path + ".stamp"
            try:
                if os.path.exists(sidecar):
                    os.remove(sidecar)
            except Exception:
                pass
            if stamp_type == "stamp":
                try:
                    with open(sidecar, "w", encoding="utf-8") as f:
                        f.write("stamp")
                except Exception:
                    pass  # 사이드카 실패해도 서명 자체는 살아있음

    if action in ("approve", "special"):
        approval_type = "special" if action == "special" else "normal"
        db.update_inspection_status(inspection_id, "approved", approver=approver, approval_type=approval_type)
        # PDF는 여기서 바로 만들지 않음 — "출력 대기" 화면에서 선택/전체 출력할 때 생성됨
        db.set_report_files(inspection_id, signature_path=signature_path, pdf_path=None, xlsx_path=None)
        # 승인 시점의 판정 내용을 해시로 굳혀둔다 (이후 변조 여부 검증용)
        db.set_inspection_hashes(inspection_id, content_hash=compute_content_hash(inspection_id))

        if action == "special":
            flash(f"특채 승인 완료 ({approver}) — 성적서 파일은 '출력 대기' 화면에서 선택 출력하면 돼.")
            record_change("성적서 특채 승인", "inspection", inspection_id, f"승인자 {approver}")
        else:
            flash(f"승인 완료 ({approver}) — 성적서 파일은 '출력 대기' 화면에서 선택 출력하면 돼.")
            record_change("성적서 승인", "inspection", inspection_id, f"승인자 {approver}")

    elif action == "reject":
        if not reject_reason:
            flash("반려 사유를 입력해줘.")
            return _back(on_error=True)
        db.update_inspection_status(inspection_id, "rejected",
                                    approver=approver, reject_reason=reject_reason)
        flash("반려 처리됐어. 검사자가 재입력할 수 있어.")
        record_change("성적서 반려", "inspection", inspection_id, f"반려자 {approver}, 사유: {reject_reason}")
    elif action == "failed":
        db.update_inspection_status(inspection_id, "approved", approver=approver,
                                    approval_type="failed",
                                    reject_reason=reject_reason)
        # 불합격도 서명이 남아야 한다 — 부적합 통보서로 협력사에 나가는 결정이라 근거가 필요함
        db.set_report_files(inspection_id, signature_path=signature_path, pdf_path=None, xlsx_path=None)
        db.set_inspection_hashes(inspection_id, content_hash=compute_content_hash(inspection_id))
        flash(f"불합격 확정됐어 ({approver}). 검사자가 부적합 통보서를 작성할 수 있어.")
        record_change("성적서 불합격 확정", "inspection", inspection_id, f"확정자 {approver}")
    else:
        flash("알 수 없는 액션이야.")

    return _back()


# ---------- 사용자 기본 서명 저장/확인 ----------

@app.route("/signature/default/save", methods=["POST"])
@perm_required("approve")
def signature_default_save():
    """현재 사용자의 기본(default) 서명 PNG를 저장."""
    sig_data = request.form.get("signature_data", "").strip()
    if not sig_data or "," not in sig_data:
        return {"error": "서명 데이터가 없어."}, 400
    try:
        _, b64 = sig_data.split(",", 1)
        png_bytes = base64.b64decode(b64)
    except Exception as e:
        return {"error": f"서명 디코딩 실패: {e}"}, 400
    user_id = g.user["id"]
    os.makedirs(SIGNATURE_DIR, exist_ok=True)
    path = os.path.join(SIGNATURE_DIR, f"user_{user_id}_default.png")
    try:
        with open(path, "wb") as f:
            f.write(png_bytes)
    except Exception as e:
        return {"error": f"저장 실패: {e}"}, 500
    return {"ok": True, "url": f"/static/signatures/user_{user_id}_default.png"}


@app.route("/signature/default/check")
@perm_required("approve")
def signature_default_check():
    """현재 사용자의 기본 서명 존재 여부 반환."""
    user_id = g.user["id"]
    path = os.path.join(SIGNATURE_DIR, f"user_{user_id}_default.png")
    if os.path.exists(path):
        return {"exists": True, "url": f"/static/signatures/user_{user_id}_default.png"}
    return {"exists": False}


# ---------- 일괄 합격 승인 ----------

@app.route("/approve/batch", methods=["POST"])
@perm_required("approve")
def approve_batch():
    """선택된 성적서들을 일괄 합격 승인 (서명 하나로 전부 처리)."""
    id_strs = request.form.getlist("inspection_ids")
    if not id_strs:
        flash("선택된 성적서가 없어.")
        return redirect(url_for("approve_list", tab="pending"))

    allowed, why = _can_make_final_decision(g.user)
    if not allowed:
        flash(why)
        return redirect(url_for("approve_list", tab="pending"))

    signature_data = request.form.get("signature_data", "").strip()
    if not signature_data:
        flash("서명을 먼저 해줘.")
        return redirect(url_for("approve_list", tab="pending"))

    approver = g.user["display_name"] or g.user["username"]
    success_ids = []
    errors = []

    for id_str in id_strs:
        try:
            inspection_id = int(id_str)
        except ValueError:
            continue
        header, _items = db.get_inspection(inspection_id)
        if header is None or header["status"] != "pending":
            errors.append(f"#{inspection_id}")
            continue
        # 단건 승인과 똑같은 규칙을 적용한다 (규격 미입력·미측정 등)
        if _final_decision_block_reason(inspection_id, "approve"):
            errors.append(f"#{inspection_id}(판정 근거 미비)")
            continue

        sig_path, sig_err = _save_signature(inspection_id, signature_data)
        if sig_err:
            errors.append(f"#{inspection_id}(서명오류)")
            continue

        db.update_inspection_status(inspection_id, "approved", approver=approver, approval_type="normal")
        db.set_report_files(inspection_id, signature_path=sig_path, pdf_path=None, xlsx_path=None)
        db.set_inspection_hashes(inspection_id, content_hash=compute_content_hash(inspection_id))
        record_change("성적서 승인(일괄)", "inspection", inspection_id, f"승인자 {approver}")
        success_ids.append(inspection_id)

    if errors:
        flash(f"{len(success_ids)}건 합격 승인 완료, {len(errors)}건 건너뜀 ({', '.join(errors)}) — 이미 처리됐거나 대기 상태가 아닌 항목이야.")
    else:
        flash(f"{len(success_ids)}건 합격 승인 완료 — 성적서 파일은 '출력 대기'에서 출력해줘.")
    return redirect(url_for("approve_list", tab="pass"))


# ---------- 검사 이력 삭제 ----------

# 관리자 삭제가 돌아올 수 있는 리스트 페이지 화이트리스트.
# return_to로 임의 라우트를 못 부르게 여기서 통제한다.
_ADMIN_DELETE_RETURNS = {
    "history": "history",
    "approve_list": "approve_list",
    "approval_history": "approval_history",
    "output_list": "output_list",
    "output_history": "output_history",
    "ncr_list": "ncr_list",
    "supplier_report_list": "supplier_report_list",
    "defect_history": "defect_history",
}

def _resolve_return_to(default="history"):
    key = (request.form.get("return_to") or "").strip()
    return _ADMIN_DELETE_RETURNS.get(key, default)


@app.route("/history/delete-selected", methods=["POST"])
@perm_required("history_delete")
def history_delete_selected():
    """선택된 성적서 일괄 삭제. return_to로 원래 페이지로 돌아간다."""
    ids_raw = request.form.getlist("inspection_ids")
    ids = []
    for x in ids_raw:
        try:
            ids.append(int(x))
        except ValueError:
            pass
    back = _resolve_return_to("history")
    if not ids:
        flash("삭제할 성적서를 선택해줘.")
        return redirect(url_for(back))
    for iid in ids:
        record_change("성적서 삭제", "inspection", iid, f"삭제자: {g.user['display_name'] or g.user['username']}")
    db.delete_inspections(ids)
    flash(f"{len(ids)}건 삭제됐어. 입고 항목 상태도 확인해봐.")
    return redirect(url_for(back))


@app.route("/ncr/delete-selected", methods=["POST"])
def ncr_delete_selected():
    """admin 전용 부적합 통보서 일괄 삭제."""
    guard = _admin_only()
    if guard: return guard
    ids_raw = request.form.getlist("ncr_ids")
    ids = []
    for x in ids_raw:
        try:
            ids.append(int(x))
        except ValueError:
            pass
    if not ids:
        flash("삭제할 통보서를 선택해줘.")
        return redirect(url_for("ncr_list"))
    for nid in ids:
        record_change("부적합 통보서 삭제(admin)", "ncr", nid,
                      f"삭제자: {g.user['display_name'] or g.user['username']}")
    db.delete_ncrs(ids)
    flash(f"통보서 {len(ids)}건 삭제됐어.")
    return redirect(url_for("ncr_list"))


@app.route("/supplier-reports/delete-selected", methods=["POST"])
def supplier_report_delete_selected():
    """admin 전용 업체 성적표 일괄 삭제 (상태 불문)."""
    guard = _admin_only()
    if guard: return guard
    ids_raw = request.form.getlist("report_ids")
    ids = []
    for x in ids_raw:
        try:
            ids.append(int(x))
        except ValueError:
            pass
    if not ids:
        flash("삭제할 성적표를 선택해줘.")
        return redirect(url_for("supplier_report_list"))
    for rid in ids:
        record_change("업체 성적표 삭제(admin)", "supplier_report", rid,
                      f"삭제자: {g.user['display_name'] or g.user['username']}")
    db.delete_supplier_reports_admin(ids)
    flash(f"성적표 {len(ids)}건 삭제됐어.")
    return redirect(url_for("supplier_report_list"))


# ---------- 승인 목록 / 승인 전용 화면 ----------

@app.route("/approve")
@perm_required("approve")
def approve_list():
    """승인 목록 — 필터/검색 지원."""
    tab = request.args.get("tab", "pending")
    q   = request.args.get("q", "").strip()

    if tab == "pending":
        rows = db.list_inspections(status="pending")
    elif tab == "pass":
        rows = [r for r in db.list_inspections(status="approved") if r["approval_type"] == "normal"]
    elif tab == "failed":
        rows = [r for r in db.list_inspections(status="approved") if r["approval_type"] == "failed"]
    elif tab == "special":
        rows = [r for r in db.list_inspections(status="approved") if r["approval_type"] == "special"]
    elif tab == "rejected":
        rows = db.list_inspections(status="rejected")
    else:
        rows = db.list_inspections(status="pending")

    if q:
        ql = q.lower()
        rows = [r for r in rows if ql in (r["material_no"] or "").lower()
                or ql in (r["material_name"] or "").lower()
                or ql in (r["supplier"] or "").lower()
                or ql in (r["inspector"] or "").lower()]

    counts = {
        "pending":  len(db.list_inspections(status="pending")),
        "pass":     sum(1 for r in db.list_inspections(status="approved") if r["approval_type"] == "normal"),
        "failed":   sum(1 for r in db.list_inspections(status="approved") if r["approval_type"] == "failed"),
        "special":  sum(1 for r in db.list_inspections(status="approved") if r["approval_type"] == "special"),
        "rejected": len(db.list_inspections(status="rejected")),
    }
    time_labels = {r["id"]: total_time_label_for(r) for r in rows}
    return render_template("approve_list.html", rows=rows, tab=tab, q=q, counts=counts,
                           time_labels=time_labels)


@app.route("/approve/<int:inspection_id>/revoke", methods=["POST"])
@perm_required("approve_revoke")
def approve_revoke(inspection_id):
    """승인/반려/불합격 결정을 취소하고 대기 상태로 되돌림."""
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        flash("존재하지 않는 성적서야.")
        return redirect(url_for("approve_list"))
    if header["status"] not in ("approved", "rejected"):
        flash("대기 중인 성적서는 회수할 수 없어.")
        return redirect(url_for("approve_list"))
    # 연결된 부적합 통보서/반품 기록이 있으면 회수 금지 (고아 레코드 방지)
    linked_ncr = db.list_ncr(inspection_id=inspection_id)
    linked_return = db.get_return_requests_by_inspection(inspection_id)
    if linked_ncr or linked_return:
        parts = []
        if linked_ncr:
            parts.append(f"부적합 통보서 {len(linked_ncr)}건")
        if linked_return:
            parts.append(f"반품 {len(linked_return)}건")
        flash(f"이 성적서에 연결된 {', '.join(parts)}이(가) 있어서 회수할 수 없어. 먼저 그 기록들을 처리·삭제해줘.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))
    conn = db.get_conn()
    # 서명·무결성 해시도 같이 지운다.
    #   - 승인이 취소됐는데 승인자 서명이 남아 있으면 안 된다
    #   - 해시는 '승인 시점의 내용'을 굳힌 값이라, 회수 후 값을 고치면 변조로 오인된다
    #     (재승인하면 그 시점 기준으로 다시 굳혀진다)
    conn.execute("""UPDATE inspections
                       SET status='pending', approval_type='normal', approver=NULL,
                           approved_at=NULL, reject_reason=NULL,
                           signature_path=NULL, content_hash=NULL, pdf_hash=NULL
                     WHERE id=?""", (inspection_id,))
    conn.commit()
    conn.close()
    actor = g.user["display_name"] or g.user["username"]
    record_change("결정 회수 (대기로 복원)", "inspection", inspection_id, f"회수자: {actor}")
    flash(f"성적서 #{inspection_id} 결정이 회수됐어. 다시 승인 대기 상태야.")
    return redirect(url_for("approve_list"))


@app.route("/approve/<int:inspection_id>")
@perm_required("approve")
def approve_view(inspection_id):
    """항목·측정값·검사방법이 모두 보이는 승인 전용 화면."""
    header, items = db.get_inspection(inspection_id)
    if header is None:
        flash("존재하지 않는 성적서야.")
        return redirect(url_for("approve_list"))
    if header["status"] != "pending":
        flash("이미 처리된 성적서야.")
        return redirect(url_for("approve_list"))

    per_cycle_sec = header["actual_time_sec"] or 0
    per_cycle_label = format_duration(per_cycle_sec)

    # 총 측정시간 + AQL 그룹
    specs_raw, _, _ = _get_specs_for_material(header["material_no"])
    total_time_label = "-"
    specs_with_sample_a = build_specs_with_sample(specs_raw, header["quantity"])
    if specs_with_sample_a and per_cycle_sec:
        total_time_label = format_duration(compute_total_time_sec(specs_with_sample_a, per_cycle_sec))
    specs_map_a = {s["item_name"]: s for s in specs_with_sample_a}
    aql_groups_a = _unique_aql_sample_qty(specs_with_sample_a) if specs_with_sample_a else {}

    today = _dt.now().date()
    gauge_alerts = []

    # 항목별 측정값 파싱 + 문제/합격 분류
    def _parse_item(it):
        raw = (it["measured_value"] or "").strip()
        tokens = [v.strip() for v in raw.split(",") if v.strip()]
        lo, hi = it["lower_limit"], it["upper_limit"]
        parsed_vals = []
        for t in tokens:
            try:
                v = float(t)
                out = False
                if lo is not None and v < lo:
                    out = True
                if hi is not None and v > hi:
                    out = True
                parsed_vals.append({"display": t, "out": out, "is_num": True})
            except ValueError:
                # O/X 육안 값
                parsed_vals.append({"display": t, "out": t.upper() == "X", "is_num": False})

        expiry_str = it["gauge_expiry"]
        days_left = None
        if expiry_str:
            try:
                expiry_date = _dt.strptime(expiry_str, "%Y-%m-%d").date()
                days_left = (expiry_date - today).days
                if days_left <= 30:
                    gauge_alerts.append({
                        "item_name": it["item_name"],
                        "spec_display": it["spec_display"],
                        "days_left": days_left,
                    })
            except ValueError:
                pass

        return dict(it) | {
            "parsed_vals": parsed_vals,
            "gauge_days_left": days_left,
        }

    NOT_READY_RESULTS = ("미측정", "입력오류", NO_SPEC_RESULT)

    all_items = [_parse_item(it) for it in items]
    problem_items  = [it for it in all_items if it["result"] == "불합격"]
    pending_items  = [it for it in all_items if it["result"] in NOT_READY_RESULTS]
    pass_items     = [it for it in all_items
                       if it["result"] != "불합격" and it["result"] not in NOT_READY_RESULTS]

    problem_count = len(problem_items)
    pending_count = len(pending_items)
    pass_count    = len(pass_items)
    return render_template("approve_form.html",
                           header=header,
                           problem_items=problem_items,
                           pending_items=pending_items,
                           pass_items=pass_items,
                           all_items=all_items,
                           problem_count=problem_count,
                           pending_count=pending_count,
                           pass_count=pass_count,
                           per_cycle_label=per_cycle_label,
                           total_time_label=total_time_label,
                           aql_groups=aql_groups_a,
                           specs_map=specs_map_a,
                           gauge_alerts=gauge_alerts)


# ---------- 검사 이력 ----------

@app.route("/history")
@perm_required("inspect_history")
def history():
    """전체 성적서를 검사일 기준으로 날짜별 그룹화해서 보여줌.

    검사일이 '260821'처럼 다른 형식으로 들어온 값이 섞여 있어서, 원본 문자열을 그대로
    그룹 키로 쓰면 같은 날인데 그룹이 갈라진다. 날짜로 해석해서 YYYY-MM-DD 로 묶는다.

    입고일이 아니라 검사일 기준으로 묶는다 — 입고 처리를 며칠 전에 해두고 실제 검사는
    나중에 하는 경우가 많아서, 입고일로 묶으면 "오늘 검사한 것"이 어제·그제 그룹에
    흩어져 보이는 문제가 있었다(2026-08-27 사용자 피드백으로 변경).
    """
    from collections import defaultdict
    NO_DATE = "날짜 없음"
    inspections = db.list_inspections()

    f = _list_search_params()
    inspections = [
        insp for insp in inspections
        if _row_passes_search(
            f,
            inspector=insp["inspector"] or "", supplier=insp["supplier"] or "",
            product=insp["material_name"] or "", material=insp["material_no"] or "",
            result=insp["overall_result"] or "",
            status=_approval_status_label(insp["status"], insp["overall_result"], insp["approval_type"]),
            insp_date=insp["inspect_date"], recv_date=insp["receive_date"],
        )
    ]

    # 검사일 내림차순으로 정렬해두고 페이지네이션(50개/페이지) — 페이지 안에서
    # 같은 날짜끼리 다시 그룹핑해서 헤더를 붙인다(원래 화면 형태 유지).
    def _date_key(i):
        d = _parse_any_date(i["inspect_date"])
        return d.isoformat() if d else "0000"
    inspections.sort(key=lambda i: (_date_key(i), i["id"]), reverse=True)
    pager = _paginate(inspections)

    by_date = defaultdict(list)
    for insp in pager["items"]:
        d = _parse_any_date(insp["inspect_date"])
        by_date[d.isoformat() if d else NO_DATE].append(insp)

    sorted_dates = sorted(by_date.keys(),
                          key=lambda d: d if d != NO_DATE else "0000",
                          reverse=True)
    # 헤더에 그대로 쓸 수 있게 '2026-08-21 (금)' 형태의 표시용 라벨도 같이 넘긴다
    date_labels = {d: (format_date_korean(d) if d != NO_DATE else NO_DATE) for d in sorted_dates}
    time_labels = {i["id"]: total_time_label_for(i) for i in pager["items"]}
    drawing_materials = materials_with_drawings(i["material_no"] for i in pager["items"])
    return render_template("history.html", by_date=by_date,
                           sorted_dates=sorted_dates, date_labels=date_labels,
                           time_labels=time_labels, drawing_materials=drawing_materials,
                           pager=pager,
                           f=f, result_options=OVERALL_RESULT_OPTIONS, status_options=APPROVAL_STATUS_LABELS)


# ---------- 승인 이력 (필터+엑셀 내보내기) ----------
#
# 검사 이력(/history)은 상태 무관 전체를 검사일 그룹으로 보여주는 화면이라
# "판정이 확정된 것만 뽑아서 업체·발주번호로 필터링 + 엑셀로 내보내기"에는 안 맞았음.
# 아래는 승인 완료(합격·특채·불합격 확정)된 것만 리스트로 정리해서 필터·엑셀 내보내기 전용
# 인터페이스로 따로 뺀 것.

def _multi_arg(name):
    """?supplier=A&supplier=B  또는  ?supplier=A,B  둘 다 받는다.
       대시보드 필터와 같은 방식."""
    out = []
    for v in request.args.getlist(name):
        for part in (v or "").split(","):
            p = part.strip()
            if p:
                out.append(p)
    return out


APPROVAL_HISTORY_STATES = ("합격", "특채", "불합격")

# 검사자 비고의 표준 불량 문구 "검사 수량 N개 중 M개 불량 P%" 를 파싱해서
# 승인 이력에 불량 카운트 컬럼을 자동으로 채운다. 여러 줄이면 합산.
_DEFECT_RE = re.compile(r"검사\s*수량\s*(\d+)\s*개\s*중\s*(\d+)\s*개\s*불량")


def parse_defect_counts(remark_text):
    """검사자 비고에서 불량 수량 합계를 뽑아낸다.
       return (bad_total, matched_line_count) — 매칭 없으면 (0, 0)."""
    if not remark_text:
        return 0, 0
    total_bad = 0
    matches = 0
    for m in _DEFECT_RE.finditer(remark_text):
        try:
            total_bad += int(m.group(2))
            matches += 1
        except ValueError:
            pass
    return total_bad, matches


def _collect_approval_history():
    """승인 이력 화면·엑셀 내보내기가 공유하는 필터·조회 로직.
       한 곳에 몰아둬야 화면과 파일이 어긋나지 않는다."""
    start_s   = (request.args.get("start") or "").strip()
    end_s     = (request.args.get("end") or "").strip()
    suppliers = _multi_arg("supplier")
    po_nums   = _multi_arg("po_number")
    states    = [s for s in _multi_arg("state") if s in APPROVAL_HISTORY_STATES]

    start_d = _parse_any_date(start_s) if start_s else None
    end_d   = _parse_any_date(end_s)   if end_s   else None

    all_rows = db.list_inspections()
    approved_rows = []
    all_suppliers = set()
    all_pos = set()

    for r in all_rows:
        state = db._lot_state(r["status"], r["approval_type"])
        if state not in APPROVAL_HISTORY_STATES:
            continue
        # 승인 이력 후보 전체에서 필터 선택지(업체·발주번호) 뽑아낸다 —
        # 현재 필터 결과가 아니라 전체 목록에서 골라야 조건을 바꿔가며 조회할 수 있음.
        if r["supplier"]:  all_suppliers.add(r["supplier"])
        if r["po_number"]: all_pos.add(r["po_number"])

        d = _parse_any_date(r["receive_date"])
        if start_d and (not d or d < start_d): continue
        if end_d   and (not d or d > end_d):   continue
        if suppliers and (r["supplier"] or "") not in suppliers: continue
        if po_nums   and (r["po_number"] or "") not in po_nums:  continue
        if states    and state not in states:                    continue

        # 총 검사시간 — 저장된 total_time_sec 우선, 없으면 spec+수량으로 재계산.
        # 옛날 성적서는 total_time_sec 컬럼이 추가되기 전에 만들어져서 NULL임.
        # 사이클당 시간(actual_time_sec)에 샘플수를 곱해야 진짜 "총" 시간이 나온다.
        total_sec = r["total_time_sec"] or 0
        max_sample = 0
        try:
            specs, _grp, _gid = _get_specs_for_material(r["material_no"] or "")
            sws = build_specs_with_sample(specs, r["quantity"] or 0)
            max_sample = max((s.get("sample_qty", 0) or 0) for s in sws) if sws else 0
            if not total_sec and r["actual_time_sec"]:
                total_sec = compute_total_time_sec(sws, r["actual_time_sec"])
        except Exception:
            if not total_sec:
                total_sec = r["actual_time_sec"] or 0

        # 비고에 적힌 표준 불량 문구 → 불량수량 합계 (5번 표준 포맷)
        bad_count, _bad_lines = parse_defect_counts(r["remark_inspector"] or "")
        # 합격수량 = 최대 샘플 수 - 불량수량 (샘플기준으로 표시)
        pass_count = max(0, (max_sample or 0) - bad_count) if max_sample else max(0, -bad_count)

        approved_rows.append({
            "id": r["id"],
            "receive_date": r["receive_date"] or "",
            "receive_date_label": format_date_korean(d.isoformat()) if d else (r["receive_date"] or ""),
            "supplier": r["supplier"] or "",
            "po_number": r["po_number"] or "",
            "material_name": r["material_name"] or "",
            "material_no": r["material_no"] or "",
            "quantity": r["quantity"] or 0,
            "inspector": r["inspector"] or "",
            "total_time_sec": total_sec,
            "total_time_label": format_duration(total_sec),
            "state": state,
            "max_sample": max_sample,
            "bad_count": bad_count,
            "pass_count": pass_count,
        })

    # 최신 입고일부터, 같은 날은 id 큰 것부터
    approved_rows.sort(key=lambda x: (x["receive_date"], x["id"]), reverse=True)

    filt = {
        "start": start_s, "end": end_s,
        "suppliers": suppliers, "po_nums": po_nums, "states": states,
        "all_suppliers": sorted(all_suppliers),
        "all_pos": sorted(all_pos, reverse=True),
        "all_states": list(APPROVAL_HISTORY_STATES),
    }
    return approved_rows, filt


@app.route("/history/approved")
@perm_required("inspect_history")
def approval_history():
    rows, filt = _collect_approval_history()
    pager = _paginate(rows)
    return render_template("approval_history.html", rows=pager["items"], filt=filt, pager=pager)


@app.route("/history/approved/export")
@perm_required("inspect_history")
def approval_history_export():
    """첨부 양식 그대로: B1 제목, 2행 헤더(B~J), 3행부터 데이터."""
    import io
    from datetime import date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    rows, _ = _collect_approval_history()

    # 체크박스로 고른 건만 뽑을 때 사용 — 전체를 한 번에 뽑으면 Render 서버
    # 리소스가 부족해서 타임아웃 나는 문제가 있어(2026-09-04 사용자 리포트),
    # 화면에서 선택한 것만 내보낼 수 있게 함. ids가 없으면 예전처럼 전체.
    selected_ids = _multi_arg("ids")
    if selected_ids:
        wanted = {int(x) for x in selected_ids if x.isdigit()}
        rows = [r for r in rows if r["id"] in wanted]

    wb = Workbook()
    ws = wb.active
    ws.title = "승인이력"

    # 제목
    ws["B1"] = "검사 결과 리스트"
    ws["B1"].font = Font(name="맑은 고딕", size=16, bold=True)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("B1:M1")
    ws.row_dimensions[1].height = 28

    headers = ["입고일", "업체명", "로트번호", "자재명", "자재번호",
               "입고수량", "검사자", "검사시간", "판정여부",
               "합격 수량", "불량 수량", "AQL 최대 샘플"]
    thin = Side(style="thin", color="B7BEC9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E7EAF0")

    for i, h in enumerate(headers):
        c = ws.cell(row=2, column=2 + i, value=h)
        c.font = Font(name="맑은 고딕", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = header_fill
        c.border = border
    ws.row_dimensions[2].height = 22

    for row_i, r in enumerate(rows, start=3):
        qty_disp = f"{r['quantity']}개" if r["quantity"] else ""
        values = [
            r["receive_date_label"],
            r["supplier"],
            r["po_number"],
            r["material_name"],
            r["material_no"],
            qty_disp,
            r["inspector"],
            r["total_time_label"],
            r["state"],
            f"{r['pass_count']}개" if r["max_sample"] else "",
            f"{r['bad_count']}개" if r["bad_count"] else ("0개" if r["max_sample"] else ""),
            f"{r['max_sample']}개" if r["max_sample"] else "",
        ]
        for j, v in enumerate(values):
            c = ws.cell(row=row_i, column=2 + j, value=v)
            c.font = Font(name="맑은 고딕")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

    widths = [15, 14, 14, 34, 14, 10, 10, 12, 10, 11, 11, 13]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord('B') + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"승인이력_{_date.today().isoformat()}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


# ---------- 불량 이력 ----------

@app.route("/defects")
@perm_required("defect_history")
def defect_history():
    from datetime import date, timedelta
    # 완료 archive 기간 필터 (active 항목은 전체 표시)
    preset = request.args.get("preset", "3m")
    start  = request.args.get("start", "")
    end    = request.args.get("end", "")
    today  = date.today()

    if preset == "1m":
        start = today.replace(day=1).isoformat()
        end   = today.isoformat()
    elif preset == "3m":
        start = (today - timedelta(days=90)).isoformat()
        end   = today.isoformat()
    elif preset == "6m":
        start = (today - timedelta(days=180)).isoformat()
        end   = today.isoformat()

    data = db.get_defect_followup(
        completed_start=start or None,
        completed_end=end or None,
    )

    f = _list_search_params()

    def _filt(rows):
        return [
            r for r in rows
            if _row_passes_search(
                f,
                inspector=r.get("inspector") or "", supplier=r.get("supplier") or "",
                product=r.get("material_name") or "", material=r.get("material_no") or "",
                insp_date=r.get("inspect_date"), recv_date=r.get("receive_date"),
            )
        ]
    data = {k: _filt(v) for k, v in data.items()}

    # 완료(archive) 버킷만 페이지네이션 — 시간이 갈수록 계속 쌓이는 유일한 항목이라서.
    # 나머지 4개 버킷(재검사/작성/확인/발송)은 처리하면 사라지는 대기열이라 페이지 나눌 필요 없음.
    completed_pager = _paginate(data.get("completed", []), page_arg="page")
    data["completed"] = completed_pager["items"]

    return render_template("defect_history.html",
                           data=data,
                           preset=preset, start=start, end=end,
                           pager=completed_pager,
                           f=f, show_result=False, show_status=False)


# ---------- 전수검사 기록지 ----------

def _fi_columns_from_specs(material_no):
    """specs 테이블 → 전수검사 열 정의 (key/header/type/lo/hi) 변환."""
    specs = db.get_specs_by_material(material_no)
    cols = []
    for sp in specs:
        jtype = sp["judge_type"] if sp["judge_type"] else "ok_ng"
        if jtype == "numeric":
            t = "num"
        elif jtype == "numeric_pair":
            t = "num_pair"
        else:
            t = "pf"
        cols.append({
            "key": sp["item_name"],
            "header": sp["spec_display"] or sp["item_name"],
            "type": t,
            "lo": sp["lower_limit"] if t in ("num", "num_pair") else None,
            "hi": sp["upper_limit"] if t in ("num", "num_pair") else None,
        })
    return cols


@app.route("/inspection/<int:inspection_id>/full-inspect", methods=["GET"])
@perm_required("inspect_input")
def full_inspect_form(inspection_id):
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        flash("존재하지 않는 성적서야.")
        return redirect(url_for("home"))
    config = db.get_full_inspect_config(header["material_no"])
    if not config:
        flash("이 자재는 전수검사 설정이 없어. 자재 상세에서 먼저 활성화해줘.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))
    columns = _fi_columns_from_specs(header["material_no"])
    fi = db.get_or_create_full_inspection(inspection_id)
    units = db.list_full_inspection_units(inspection_id)
    qty = int(header["quantity"] or 0)
    unit_map = {u["unit_no"]: u for u in units}
    rows = []
    for i in range(1, qty + 1):
        rows.append(unit_map.get(i, {
            "unit_no": i, "serial_no": "", "values": {}, "result": "", "remark": "", "gauge_name": ""}))
    tmpl = "full_inspect_housing.html" if config.get("template") == "housing" \
        else "full_inspect_form.html"
    return render_template(tmpl,
                           header=dict(header), fi=fi, rows=rows,
                           columns=columns, config=config, qty=qty)


@app.route("/inspection/<int:inspection_id>/full-inspect/save", methods=["POST"])
@perm_required("inspect_input")
def full_inspect_save(inspection_id):
    """자동저장 + 수동저장 공용. JSON으로 units 배열을 받는다."""
    from flask import jsonify
    if g.user is None:
        return jsonify({"ok": False, "expired": True}), 401
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        return jsonify({"ok": False, "error": "성적서 없음"}), 404
    config = db.get_full_inspect_config(header["material_no"])
    if not config:
        return jsonify({"ok": False, "error": "전수검사 설정 없음"}), 400

    columns = _fi_columns_from_specs(header["material_no"])
    col_keys = [c["key"] for c in columns]

    payload = request.get_json(silent=True) or {}
    inspect_date = payload.get("inspect_date", "")
    units_raw = payload.get("units", [])

    def _auto_result(unit_vals):
        has_any = any(str(v).strip() for v in unit_vals.values())
        if not has_any:
            return ""
        for col in columns:
            key = col["key"]
            val = str(unit_vals.get(key, "")).strip()
            if not val:
                continue
            if col["type"] == "pf":
                if val.upper() in ("NG", "X", "×", "△", "불합격", "FAIL"):
                    return "NG"
            elif col["type"] == "num":
                try:
                    v = float(val.replace(",", ""))
                    if col.get("lo") is not None and v < float(col["lo"]):
                        return "NG"
                    if col.get("hi") is not None and v > float(col["hi"]):
                        return "NG"
                except (ValueError, TypeError):
                    pass
            elif col["type"] == "num_pair":
                # "ct/rod" 또는 "ct/rod,..." — 두 채널 모두 범위 안이어야 OK
                for part in val.replace(",", "/").split("/"):
                    p = part.strip()
                    if not p:
                        continue
                    try:
                        v = float(p)
                        if col.get("lo") is not None and v < float(col["lo"]):
                            return "NG"
                        if col.get("hi") is not None and v > float(col["hi"]):
                            return "NG"
                    except (ValueError, TypeError):
                        pass
        return "OK"

    units_to_save = []
    for u in units_raw:
        vals = {k: str(u.get(k, "")).strip() for k in col_keys}
        result = _auto_result(vals)
        units_to_save.append({
            "unit_no": int(u.get("unit_no", 0)),
            "serial_no": str(u.get("serial_no", "")).strip(),
            "values": vals,
            "result": result,
            "remark": str(u.get("remark", "")).strip(),
        })

    db.get_or_create_full_inspection(inspection_id)
    db.update_full_inspection(inspection_id, inspect_date=inspect_date or None)
    db.save_full_inspection_units(inspection_id, units_to_save)
    ok_cnt = sum(1 for u in units_to_save if u["result"] == "OK")
    ng_cnt = sum(1 for u in units_to_save if u["result"] == "NG")
    return jsonify({"ok": True, "ok_cnt": ok_cnt, "ng_cnt": ng_cnt})


@app.route("/inspection/<int:inspection_id>/full-inspect/complete", methods=["POST"])
@perm_required("inspect_input")
def full_inspect_complete(inspection_id):
    """완료 버튼 — 현재 시각을 완료날짜로 저장."""
    from datetime import datetime as _dt2
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        flash("존재하지 않는 성적서야.")
        return redirect(url_for("home"))
    db.get_or_create_full_inspection(inspection_id)
    db.update_full_inspection(inspection_id,
                              complete_date=_dt2.now().strftime("%Y-%m-%d"),
                              status="complete")
    record_change("전수검사 완료", "full_inspection", inspection_id, header["material_no"])
    flash("전수검사 기록이 완료 처리됐어.")
    return redirect(url_for("full_inspect_form", inspection_id=inspection_id))


@app.route("/inspection/<int:inspection_id>/full-inspect/pdf")
@perm_required("output")
def full_inspect_pdf(inspection_id):
    """전수검사 기록지 PDF 미리보기/다운로드."""
    import tempfile
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        flash("존재하지 않는 성적서야.")
        return redirect(url_for("home"))
    config = db.get_full_inspect_config(header["material_no"])
    if not config:
        flash("전수검사 설정이 없어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))
    fi = db.get_full_inspection(inspection_id)
    if fi is None:
        flash("전수검사 기록이 없어. 먼저 입력해줘.")
        return redirect(url_for("full_inspect_form", inspection_id=inspection_id))
    columns = _fi_columns_from_specs(header["material_no"])
    units = db.list_full_inspection_units(inspection_id)
    fi_header_data = dict(fi)
    hd = dict(header)
    insp_hdr = {
        "material_name": hd.get("material_name", ""),
        "quantity": hd.get("quantity", 0),
        "intake_date": hd.get("intake_date", ""),
        "inspector": hd.get("inspector", ""),
    }
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp.close()
    from custom_report import build_full_inspection_sheet
    build_full_inspection_sheet(fi_header_data, units, columns, config, insp_hdr, tmp.name)
    fname = f"전수검사_{hd.get('material_no','')}_{hd.get('intake_date','')}.pdf"
    return send_file(tmp.name, as_attachment=False,
                     download_name=fname, mimetype="application/pdf")


# ---------- 성적서 출력 (승인된 것 중 선택/전체) ----------

@app.route("/output")
@perm_required("output")
def output_list():
    pending = db.list_pending_output_inspections()
    pager = _paginate(list(pending))
    return render_template("output_list.html", pending=pager["items"], pager=pager)


@app.route("/output/generate", methods=["POST"])
@perm_required("output")
def output_generate():
    if request.form.get("select_all") == "1":
        target_ids = [i["id"] for i in db.list_pending_output_inspections()]
    else:
        target_ids = [int(v) for v in request.form.getlist("inspection_ids")]

    if not target_ids:
        flash("출력할 성적서를 선택해줘.")
        return redirect(url_for("output_list"))

    incl_drawing  = request.form.get("incl_drawing")  == "1"
    incl_standard = request.form.get("incl_standard") == "1"
    incl_report   = request.form.get("incl_report")   == "1"
    # 전체 출력 버튼은 form에 체크박스 값을 포함하지 않으므로 기본값 True 처리
    if request.form.get("select_all") == "1":
        incl_drawing = incl_standard = incl_report = True

    results = []  # {"id","material_no","pdf_path","error"}
    for insp_id in target_ids:
        header, _ = db.get_inspection(insp_id)
        if header is None or header["status"] != "approved":
            results.append({"id": insp_id, "material_no": "-", "pdf_path": None,
                            "error": "승인된 성적서가 아니야(다시 확인해줘)."})
            continue

        xlsx_path, pdf_path, error_msg, is_custom = _generate_report_files(
            insp_id, header["approver"], header["signature_path"]
        )

        # 도면 포함 여부에 따라 PDF 병합
        if pdf_path and not error_msg:
            drawing_pdf = find_drawing_pdf(header["material_no"])

            def _do_merge(src, drw, incl_d, incl_s, incl_r):
                """병합 후 결과를 src(원본 PDF)에 덮어씌워 잔재 파일 없애기. 반환: (최종 경로, 에러)"""
                tmp = os.path.splitext(src)[0] + "_merged.pdf"
                result, err = report_builder.merge_report_with_drawing(
                    src, drw, tmp,
                    incl_drawing=incl_d, incl_standard=incl_s, incl_report=incl_r,
                )
                if err:
                    return src, err
                try:
                    os.replace(result, src)   # tmp → src 덮어쓰기, tmp 자동 삭제
                except Exception:
                    return result, None       # 덮어쓰기 실패 시 _merged.pdf 그대로
                return src, None

            if is_custom:
                # 커스텀 성적서는 기준서 페이지가 없음 → 성적서/기준서 토글은 무시하고
                # 도면만 옵션으로 앞에 붙인다.
                if incl_drawing and drawing_pdf:
                    pdf_path, merge_err = _do_merge(
                        pdf_path, drawing_pdf, True, False, True)
                    if merge_err:
                        error_msg = merge_err
            else:
                needs_merge = (incl_drawing and drawing_pdf) or not incl_standard or not incl_report
                all_included = incl_drawing and incl_standard and incl_report
                if needs_merge and not all_included:
                    pdf_path, merge_err = _do_merge(
                        pdf_path, drawing_pdf, incl_drawing, incl_standard, incl_report)
                    if merge_err:
                        error_msg = merge_err
                elif incl_drawing and drawing_pdf:
                    pdf_path, merge_err = _do_merge(
                        pdf_path, drawing_pdf, True, True, True)
                    if merge_err:
                        error_msg = merge_err

        # 전수검사 기록지 — 완료된 전수검사가 있으면 성적서 뒤에 붙인다
        if pdf_path and not error_msg:
            fi_config = db.get_full_inspect_config(header["material_no"])
            fi = db.get_full_inspection(insp_id) if fi_config else None
            if fi and fi.get("status") == "complete":
                try:
                    import tempfile as _tf
                    from custom_report import build_full_inspection_sheet
                    fi_units = db.list_full_inspection_units(insp_id)
                    hd2 = dict(header)
                    insp_hdr_data = {
                        "material_name": hd2.get("material_name", ""),
                        "quantity": hd2.get("quantity", 0),
                        "intake_date": hd2.get("intake_date", ""),
                        "inspector": hd2.get("inspector", ""),
                    }
                    fi_cols = _fi_columns_from_specs(header["material_no"])
                    fi_tmp = _tf.NamedTemporaryFile(delete=False, suffix="_fi.pdf")
                    fi_tmp.close()
                    build_full_inspection_sheet(dict(fi), fi_units, fi_cols,
                                                fi_config, insp_hdr_data, fi_tmp.name)
                    # 성적서 PDF + 전수검사 PDF 병합
                    merged_tmp = os.path.splitext(pdf_path)[0] + "_with_fi.pdf"
                    _, merge_fi_err = report_builder.append_pdf(
                        pdf_path, fi_tmp.name, merged_tmp)
                    if not merge_fi_err and os.path.exists(merged_tmp):
                        os.replace(merged_tmp, pdf_path)
                    try:
                        os.unlink(fi_tmp.name)
                    except Exception:
                        pass
                except Exception as fi_ex:
                    error_msg = (error_msg or "") + f" (전수검사 병합 실패: {fi_ex})"

        db.set_report_files(insp_id, signature_path=header["signature_path"],
                            pdf_path=pdf_path, xlsx_path=xlsx_path)
        # 발행된 PDF 파일 자체의 해시 — 나중에 파일이 덮어써졌는지 확인용
        db.set_inspection_hashes(insp_id, pdf_hash=compute_file_hash(pdf_path))
        results.append({
            "id": insp_id, "material_no": header["material_no"],
            "pdf_path": pdf_path, "error": error_msg,
        })

    success_count = sum(1 for r in results if r["pdf_path"] and not r["error"])
    record_change("성적서 출력", "inspection", None,
                  f"{len(target_ids)}건 중 {success_count}건 성공")

    return render_template("output_result.html", results=results)


@app.route("/output/download/<int:inspection_id>")
@perm_required("output")
def output_download(inspection_id):
    """생성된 성적서 PDF를 태블릿/PC 기기로 다운로드."""
    header, _ = db.get_inspection(inspection_id)
    if header is None or not header["pdf_path"] or not os.path.exists(header["pdf_path"]):
        flash("다운로드할 PDF 파일이 없어. 먼저 출력해줘.")
        return redirect(url_for("output_list"))
    fname = os.path.basename(header["pdf_path"])
    return send_file(header["pdf_path"], as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


@app.route("/output/download-zip")
@perm_required("output")
def output_download_zip():
    """여러 성적서 PDF를 zip 하나로 묶어서 기기로 다운로드."""
    ids = [int(v) for v in request.args.getlist("id") if v.isdigit()]
    buf = io.BytesIO()
    used_names = set()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for insp_id in ids:
            header, _ = db.get_inspection(insp_id)
            if header is None or not header["pdf_path"] or not os.path.exists(header["pdf_path"]):
                continue
            name = os.path.basename(header["pdf_path"])
            if name in used_names:
                base, ext = os.path.splitext(name)
                name = f"{base}_{insp_id}{ext}"
            used_names.add(name)
            zf.write(header["pdf_path"], arcname=name)

    if not used_names:
        flash("다운로드할 PDF 파일이 없어. 먼저 출력해줘.")
        return redirect(url_for("output_list"))

    buf.seek(0)
    today = _dt.now().strftime("%Y%m%d_%H%M%S")
    return send_file(buf, as_attachment=True, download_name=f"성적서_{today}.zip",
                     mimetype="application/zip")


# ---------- 커스텀(자유양식) 성적서 템플릿 ----------

# 디자이너에서 끌어다 쓰는 데이터 필드 → 성적서 헤더에서 값을 뽑는 규칙
CUSTOM_FIELD_KEYS = ["자재번호", "제품명", "업체명", "검사일", "검사자",
                     "로트번호", "도면번호", "입고수량", "종합판정"]


def _custom_fields_from_header(header):
    """성적서 헤더(dict/Row)에서 디자이너 데이터 필드 값들을 뽑는다."""
    def g(k):
        try:
            return header[k]
        except (KeyError, IndexError, TypeError):
            return None
    qty = g("quantity")
    overall = db._lot_state(g("status"), g("approval_type"))
    if overall in ("미결", "대체됨"):
        overall = ""    # 발행 시점엔 판정이 확정돼 있어야 정상
    return {
        "자재번호": g("material_no") or "",
        "제품명":   g("material_name") or "",
        "업체명":   g("supplier") or "",
        "검사일":   g("inspect_date") or "",
        "검사자":   g("inspector") or "",
        "로트번호": g("po_number") or "",
        "도면번호": report_builder.compute_drawing_no(g("material_no") or ""),
        "입고수량": f"{qty:,}" if isinstance(qty, int) else (str(qty) if qty is not None else ""),
        "종합판정": overall,
    }


def _custom_items_from_inspection(items):
    """검사 항목들을 커스텀 표에 넣을 행 형식으로 변환."""
    rows = []
    for it in (items or []):
        rows.append({
            "label":   report_builder.item_label(it["item_name"], it["aql"]),
            "spec":    it["spec_display"] or "",
            "method":  it["inspect_method"] or "",
            "value":   it["measured_value"] or "",
            "verdict": it["result"] or "",
        })
    return rows


@app.route("/custom-templates")
@perm_required("custom_template")
def custom_template_list():
    templates = db.list_custom_templates()
    return render_template("custom_template_list.html", templates=templates)


@app.route("/custom-templates/create", methods=["POST"])
@perm_required("custom_template")
def custom_template_create():
    name = (request.form.get("name") or "새 성적서 양식").strip()
    tid = db.create_custom_template(name=name, created_by=g.user["username"])
    record_change("커스텀 양식 생성", "custom_template", tid, name)
    return redirect(url_for("custom_template_edit", template_id=tid))


@app.route("/custom-templates/<int:template_id>")
@perm_required("custom_template")
def custom_template_edit(template_id):
    tmpl = db.get_custom_template(template_id)
    if tmpl is None:
        flash("양식을 찾을 수 없어.")
        return redirect(url_for("custom_template_list"))
    return render_template("custom_template_edit.html",
                           tmpl=tmpl, field_keys=CUSTOM_FIELD_KEYS)


@app.route("/custom-templates/<int:template_id>/save", methods=["POST"])
@perm_required("custom_template")
def custom_template_save(template_id):
    from flask import jsonify
    tmpl = db.get_custom_template(template_id)
    if tmpl is None:
        return jsonify({"ok": False, "error": "양식을 찾을 수 없어."}), 404
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or tmpl["name"]).strip() or "제목 없음"
    layout = payload.get("layout")
    import json as _json
    layout_json = _json.dumps(layout, ensure_ascii=False) if layout is not None else None
    db.update_custom_template(
        template_id, name=name, layout_json=layout_json,
        canvas_w=payload.get("canvas_w"), canvas_h=payload.get("canvas_h"),
        orientation=payload.get("orientation"),
    )
    record_change("커스텀 양식 저장", "custom_template", template_id, name)
    return jsonify({"ok": True})


@app.route("/custom-templates/<int:template_id>/delete", methods=["POST"])
@perm_required("custom_template")
def custom_template_delete(template_id):
    tmpl = db.get_custom_template(template_id)
    if tmpl:
        db.delete_custom_template(template_id)
        record_change("커스텀 양식 삭제", "custom_template", template_id, tmpl["name"])
        flash(f"'{tmpl['name']}' 양식을 삭제했어.")
    return redirect(url_for("custom_template_list"))


@app.route("/custom-templates/<int:template_id>/assign", methods=["GET", "POST"])
@perm_required("custom_template")
def custom_template_assign(template_id):
    tmpl = db.get_custom_template(template_id)
    if tmpl is None:
        flash("양식을 찾을 수 없어.")
        return redirect(url_for("custom_template_list"))

    if request.method == "POST":
        chosen = set(request.form.getlist("material_no"))
        # 화면에 실제로 표시됐던 자재만 대상으로 처리(검색 필터로 안 보인 자재는 건드리지 않음)
        shown = set(request.form.getlist("shown"))
        current = {m["material_no"] for m in db.materials_for_template(template_id)}
        # 이 화면에서 체크 해제된(=현재 이 양식이지만 이번에 안 고른) 자재 → 기본 양식
        for mno in (current & shown) - chosen:
            db.set_material_template(mno, None)
        # 새로 고른 자재 → 이 양식(다른 양식에 물려 있었어도 이쪽으로 이동)
        for mno in chosen - current:
            db.set_material_template(mno, template_id)
        record_change("커스텀 양식 지정", "custom_template", template_id,
                      f"{len(chosen)}개 자재")
        flash(f"'{tmpl['name']}' 양식 지정을 저장했어(선택 {len(chosen)}개).")
        return redirect(url_for("custom_template_assign", template_id=template_id))

    q = (request.args.get("q") or "").strip()
    materials = db.get_materials()
    if q:
        materials = [m for m in materials
                     if q.lower() in (m["material_no"] or "").lower()
                     or q.lower() in (m["material_name"] or "").lower()]
    # 다른 템플릿에 이미 물린 자재는 표시(중복 지정 방지 안내용)
    other_names = {t["id"]: t["name"] for t in db.list_custom_templates()}
    return render_template("custom_template_assign.html",
                           tmpl=tmpl, materials=materials, q=q,
                           other_names=other_names, this_id=template_id)


@app.route("/custom-templates/<int:template_id>/preview.pdf")
@perm_required("custom_template")
def custom_template_preview(template_id):
    import custom_report
    tmpl = db.get_custom_template(template_id)
    if tmpl is None:
        flash("양식을 찾을 수 없어.")
        return redirect(url_for("custom_template_list"))

    # 미리보기 데이터: 이 양식이 지정된 자재의 최신 승인 성적서가 있으면 그걸로, 없으면 샘플값
    data = _custom_preview_data(template_id)
    # 미리보기는 실제 발행폴더가 아니라 임시폴더에 만든다(성적서 발행 폴더 오염 방지)
    import tempfile
    tmp_pdf = os.path.join(tempfile.gettempdir(), f"iqc_preview_custom_{template_id}.pdf")
    path, err = custom_report.build_custom_report(tmpl, data, tmp_pdf)
    if err or not path:
        flash(f"미리보기 생성 실패: {err or '알 수 없는 오류'}")
        return redirect(url_for("custom_template_edit", template_id=template_id))
    return send_file(path, mimetype="application/pdf")


def _custom_preview_data(template_id):
    """미리보기용 데이터 — 지정 자재의 최신 승인건이 있으면 실제값, 없으면 샘플."""
    from datetime import date
    for m in db.materials_for_template(template_id):
        insp = db.latest_inspection_for_material(m["material_no"]) \
            if hasattr(db, "latest_inspection_for_material") else None
        if insp:
            header, items = db.get_inspection(insp["id"])
            if header:
                return {
                    "fields": _custom_fields_from_header(header),
                    "items": _custom_items_from_inspection(items),
                    "signature_path": header["signature_path"],
                    "logo_path": report_builder.LOGO_PATH,
                }
    # 샘플 데이터
    return {
        "fields": {
            "자재번호": "600005P086", "제품명": "둥근머리 볼트(M416L,STS304)",
            "업체명": "ACE", "검사일": date.today().isoformat(), "검사자": "홍길동",
            "로트번호": "CK260823-01", "도면번호": "A600005-086", "입고수량": "2,000",
        },
        "items": [
            {"label": "*A", "spec": "Ø9.0 ±0.1", "method": "캘리퍼", "value": "9.02", "verdict": "합격"},
            {"label": "B", "spec": "16 +0.2", "method": "캘리퍼", "value": "16.1", "verdict": "합격"},
            {"label": "C", "spec": "백색 아연도금 5㎛ 이상", "method": "육안", "value": "양호", "verdict": "합격"},
        ],
        "signature_path": None,
        "logo_path": report_builder.LOGO_PATH,
    }


# ---------- 출력 기록 ----------

@app.route("/output/history")
@perm_required("output")
def output_history():
    q = request.args.get("q", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    rows = db.list_output_history(q, date_from, date_to)
    return render_template("output_history.html", rows=rows, q=q,
                           date_from=date_from, date_to=date_to)


@app.route("/output/history/export.xlsx")
@perm_required("output")
def output_history_export():
    """출력기록 엑셀 내보내기 — ids= 파라미터로 선택 항목만, 없으면 전체(현재 검색 조건)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font as XFont, Alignment as XAlign, PatternFill, Border, Side

    q = request.args.get("q", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    rows = [dict(r) for r in db.list_output_history(q, date_from, date_to)]

    selected_ids = _multi_arg("ids")
    if selected_ids:
        wanted = {int(x) for x in selected_ids if x.isdigit()}
        rows = [r for r in rows if r["id"] in wanted]

    wb = Workbook()
    ws = wb.active
    ws.title = "출력기록"

    ws["B1"] = "출력 기록"
    ws["B1"].font = XFont(name="맑은 고딕", size=16, bold=True)
    ws["B1"].alignment = XAlign(horizontal="center", vertical="center")
    ws.merge_cells("B1:I1")
    ws.row_dimensions[1].height = 28

    headers = ["번호", "자재번호", "자재명", "업체", "로트번호", "결정", "승인자", "승인일시"]
    thin = Side(style="thin", color="B7BEC9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="E7EAF0")
    for i, h in enumerate(headers):
        c = ws.cell(row=2, column=2 + i, value=h)
        c.font = XFont(name="맑은 고딕", bold=True)
        c.alignment = XAlign(horizontal="center", vertical="center")
        c.fill = header_fill
        c.border = border
    ws.row_dimensions[2].height = 22

    state_label = {"normal": "합격", "special": "특채", "failed": "불합격"}
    for row_i, r in enumerate(rows, start=3):
        values = [
            r["id"], r["material_no"], r["material_name"] or "",
            r["supplier"] or "", r["po_number"] or "",
            state_label.get(r["approval_type"] or "", "-"),
            r["approver"] or "",
            r["approved_at"] or "",
        ]
        for j, v in enumerate(values):
            c = ws.cell(row=row_i, column=2 + j, value=v)
            c.font = XFont(name="맑은 고딕")
            c.border = border

    from openpyxl.utils import get_column_letter
    col_widths = [8, 18, 30, 18, 18, 8, 12, 20]
    for i, w in enumerate(col_widths, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date as _date
    fname = f"출력기록_{_date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/output/history/<int:inspection_id>")
@perm_required("output")
def output_history_detail(inspection_id):
    header, items = db.get_inspection(inspection_id)
    if header is None or header["pdf_path"] is None:
        flash("출력된 적 없는 성적서야.")
        return redirect(url_for("output_history"))
    drawing_pdf = find_drawing_pdf(header["material_no"])
    return render_template("output_history_detail.html",
                           header=header, items=items,
                           has_drawing=bool(drawing_pdf))


@app.route("/output/regenerate/<int:inspection_id>", methods=["POST"])
@perm_required("output")
def output_regenerate(inspection_id):
    """출력 기록에서 특정 성적서를 다시 생성."""
    header, _ = db.get_inspection(inspection_id)
    if header is None or header["status"] != "approved":
        flash("승인된 성적서가 아니야.")
        return redirect(url_for("output_history"))

    incl_drawing  = request.form.get("incl_drawing")  == "1"
    incl_standard = request.form.get("incl_standard") == "1"
    incl_report   = request.form.get("incl_report")   == "1"

    xlsx_path, pdf_path, error_msg, is_custom = _generate_report_files(
        inspection_id, header["approver"], header["signature_path"]
    )

    if pdf_path and not error_msg:
        drawing_pdf = find_drawing_pdf(header["material_no"])

        def _do_merge_r(src, drw, incl_d, incl_s, incl_r):
            tmp = os.path.splitext(src)[0] + "_merged.pdf"
            result, err = report_builder.merge_report_with_drawing(
                src, drw, tmp, incl_drawing=incl_d, incl_standard=incl_s, incl_report=incl_r)
            if err:
                return src, err
            try:
                os.replace(result, src)
            except Exception:
                return result, None
            return src, None

        if is_custom:
            # 커스텀 성적서는 기준서 페이지가 없음 → 도면만 옵션으로 앞에 붙인다.
            if incl_drawing and drawing_pdf:
                pdf_path, merge_err = _do_merge_r(pdf_path, drawing_pdf, True, False, True)
                if merge_err:
                    error_msg = merge_err
        else:
            needs_merge = (incl_drawing and drawing_pdf) or not incl_standard or not incl_report
            all_included = incl_drawing and incl_standard and incl_report
            if needs_merge and not all_included:
                pdf_path, merge_err = _do_merge_r(
                    pdf_path, drawing_pdf, incl_drawing, incl_standard, incl_report)
                if merge_err:
                    error_msg = merge_err
            elif incl_drawing and drawing_pdf:
                pdf_path, merge_err = _do_merge_r(pdf_path, drawing_pdf, True, True, True)
                if merge_err:
                    error_msg = merge_err

    db.set_report_files(inspection_id, signature_path=header["signature_path"],
                        pdf_path=pdf_path, xlsx_path=None)
    # 재출력해도 판정 내용은 안 바뀌므로 content_hash는 그대로 두고 pdf_hash만 갱신
    db.set_inspection_hashes(inspection_id, pdf_hash=compute_file_hash(pdf_path))

    if error_msg:
        flash(f"재출력 중 오류: {error_msg}")
    else:
        flash(f"재출력 완료 — {pdf_path}")
    record_change("성적서 재출력", "inspection", inspection_id,
                  f"승인자 {header['approver'] or '-'}")
    return redirect(url_for("output_history_detail", inspection_id=inspection_id))


# ---------- 반려 후 재검사 입력 ----------

@app.route("/inspection/<int:inspection_id>/reinspect")
@perm_required("inspect_input")
def reinspect(inspection_id):
    """반려된 성적서를 기반으로 재검사 폼을 열어줌 — 이전 측정값 prefill"""
    header, items = db.get_inspection(inspection_id)
    if header is None or header["status"] != "rejected":
        flash("재검사 대상이 아니야.")
        return redirect(url_for("home"))
    if not _can_edit_inspection(header):
        flash("본인이 입력한 성적서만 재검사할 수 있어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    intake_row = db.get_intake(header["intake_id"]) if header["intake_id"] else None
    if intake_row is None:
        flash("입고 정보를 찾을 수 없어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    material_no = header["material_no"]
    specs, is_group, group_name = _get_specs_for_material(material_no)
    specs_with_sample = build_specs_with_sample(specs, intake_row["quantity"])

    # 이전 측정값을 prefill dict에 담기 — 그룹 검사는 (부품자재, 항목명)으로 정확히 매칭
    prefill = {
        "inspect_date": header["inspect_date"] or "",
        "inspector": header["inspector"] or "",
        "reject_reason": header["reject_reason"] or "",
    }
    items_by_key = {(it["part_material_no"], it["item_name"]): it for it in items}
    for s in specs:
        it = items_by_key.get((s["material_no"], s["item_name"]))
        if it is None:
            continue
        vals = (it["measured_value"] or "").split(",")
        for i, v in enumerate(vals, start=1):
            prefill[f"item_{s['id']}_{i}"] = v.strip()
        prefill[f"item_{s['id']}_gauge_expiry"] = it["gauge_expiry"] or ""
        prefill[f"item_{s['id']}_gauge_name"] = it["gauge_name"] or ""

    prior_defect_count = db.get_defect_count_for(intake_row["supplier"], header["material_no"])
    gauges = db.list_gauges()
    return render_template("inspect_form.html",
                           intake=intake_row, specs=specs_with_sample,
                           prefill=prefill, reinspect_from=inspection_id,
                           is_group=is_group, group_name=group_name,
                           gauges=gauges,
                           sel_gauge=build_sel_gauge(specs_with_sample, gauges, prefill),
                           gauge_master_empty=(len(gauges) == 0),
                           prior_defect_count=prior_defect_count)


@app.route("/inspection/<int:inspection_id>/reinspect", methods=["POST"])
@perm_required("inspect_input")
def reinspect_submit(inspection_id):
    """재검사 결과 저장 — 기존 반려 건은 superseded 처리"""
    old_header, _ = db.get_inspection(inspection_id)
    if old_header is None:
        flash("성적서를 찾을 수 없어.")
        return redirect(url_for("home"))
    if not _can_edit_inspection(old_header):
        flash("본인이 입력한 성적서만 재검사할 수 있어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    intake_id = old_header["intake_id"]
    intake_row = db.get_intake(intake_id) if intake_id else None
    if intake_row is None:
        flash("입고 정보를 찾을 수 없어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    material_no = old_header["material_no"]
    specs, is_group, group_name = _get_specs_for_material(material_no)

    header = {
        "material_no": material_no,
        "material_name": group_name if is_group else _resolve_material_name(material_no, specs),
        "supplier": intake_row["supplier"],
        "po_number": intake_row["po_number"],
        "receive_date": intake_row["receive_date"],
        "inspect_date": request.form.get("inspect_date"),
        "inspector": g.user["display_name"] or g.user["username"],
        "quantity": intake_row["quantity"],
    }

    specs_with_sample = build_specs_with_sample(specs, intake_row["quantity"])
    items_with_results = []
    overall_ok = True
    for spec in specs_with_sample:
        allowed = aql_ac_allowance(spec["aql"], spec["sample_qty"])
        if spec["judge_type"] == "numeric":
            vals = [request.form.get(f"item_{spec['id']}_{i}", "").strip()
                    for i in range(1, 7)]
            raw_value = ",".join(v for v in vals if v)
            result, max_v, min_v = judge_numeric(raw_value, spec["lower_limit"], spec["upper_limit"], allowed)
        else:
            vals = [request.form.get(f"item_{spec['id']}_{i}", "").strip()
                    for i in range(1, 7)]
            raw_value = ",".join(v for v in vals if v)
            result, max_v, min_v = judge_visual(raw_value, allowed)

        if result != "합격":
            overall_ok = False
        gauge_expiry = request.form.get(f"item_{spec['id']}_gauge_expiry", "").strip() or None
        gauge_name = request.form.get(f"item_{spec['id']}_gauge_name", "").strip() or None
        items_with_results.append({
            "item_name": spec["item_name"],
            "measured_value": raw_value,
            "max_value": max_v,
            "min_value": min_v,
            "result": result,
            "gauge_expiry": gauge_expiry,
            "gauge_name": gauge_name,
            "part_material_no": spec["material_no"],
        })

    overall_result = "합격" if overall_ok else "검토필요"
    actual_time_sec = request.form.get("actual_time_sec", "").strip()
    actual_time_sec = int(actual_time_sec) if actual_time_sec.isdigit() else 0
    est_time_label = format_duration(actual_time_sec)
    specs_with_sample_r = build_specs_with_sample(specs, intake_row["quantity"])
    total_time_sec = compute_total_time_sec(specs_with_sample_r, actual_time_sec) if actual_time_sec else 0
    # 반려된 이전 건을 superseded로 표시하고 intake를 다시 대기로 돌림
    db.update_inspection_status(inspection_id, "superseded")
    db.set_intake_status(intake_id, "대기")
    new_id = db.create_inspection(header, items_with_results, overall_result,
                                   intake_id=intake_id, est_time_label=est_time_label,
                                   actual_time_sec=actual_time_sec,
                                   total_time_sec=total_time_sec,
                                   created_by_user_id=g.user["id"])
    flash("재검사 성적서가 생성됐어. 다시 승인을 요청해줘.")
    record_change("재검사 성적서 등록", "inspection", new_id,
                  f"자재 {material_no}, 이전 성적서 #{inspection_id}, 판정 {overall_result}")
    return redirect(url_for("inspection_detail", inspection_id=new_id))


# =========================================================================
# 자재별 검사 이력 추적
# =========================================================================

@app.route("/material/<material_no>/history")
@perm_required("inspect_history")
def material_history(material_no):
    inspections, items_by_inspection = db.get_material_inspection_history(material_no)
    if not inspections:
        flash(f"{material_no}의 검사 이력이 없어.")
        return redirect(url_for("spec_detail", material_no=material_no))

    # 항목명 목록 (순서 유지, 중복 제거)
    item_names = []
    seen = set()
    for insp in inspections:
        for it in items_by_inspection.get(insp["id"], []):
            if it["item_name"] not in seen:
                item_names.append(it["item_name"])
                seen.add(it["item_name"])

    # 인스펙션별 항목 딕셔너리로 변환 {item_name: row}
    items_dict = {}
    for insp_id, rows in items_by_inspection.items():
        items_dict[insp_id] = {r["item_name"]: r for r in rows}

    material = db.get_material(material_no)
    has_drawing = find_drawing_pdf(material_no) is not None
    return render_template("material_history.html",
                           material_no=material_no,
                           material=material,
                           inspections=inspections,
                           item_names=item_names,
                           items_dict=items_dict,
                           has_drawing=has_drawing)


# =========================================================================
# 계측기 마스터 관리
# =========================================================================

@app.route("/gauges")
@perm_required("gauge")
def gauge_list():
    from datetime import date, timedelta
    import math
    today_dt = date.today()
    today = today_dt.isoformat()
    d15 = (today_dt + timedelta(days=15)).isoformat()
    d30 = (today_dt + timedelta(days=30)).isoformat()
    raw = db.list_gauges()
    gauges = []
    for row in raw:
        g = dict(row)
        if g.get("expiry_date"):
            try:
                exp = date.fromisoformat(g["expiry_date"])
                g["days_left"] = (exp - today_dt).days
            except ValueError:
                g["days_left"] = None
        else:
            g["days_left"] = None
        gauges.append(g)
    return render_template("gauge_master.html", gauges=gauges, today=today, d15=d15, d30=d30)

@app.route("/gauges/save", methods=["POST"])
@perm_required("gauge")
def gauge_save():
    gauge_id = request.form.get("gauge_id") or None
    if gauge_id:
        gauge_id = int(gauge_id)
    name = request.form.get("name", "").strip()
    if not name:
        flash("계측기명을 입력해줘.")
        return redirect(url_for("gauge_list"))
    db.upsert_gauge(
        gauge_id=gauge_id,
        gauge_no=request.form.get("gauge_no", "").strip(),
        name=name,
        model=request.form.get("model", "").strip(),
        location=request.form.get("location", "").strip(),
        last_calibrated=request.form.get("last_calibrated", "").strip() or None,
        expiry_date=request.form.get("expiry_date", "").strip() or None,
        notes=request.form.get("notes", "").strip(),
    )
    action = "수정" if gauge_id else "등록"
    record_change(f"계측기 {action}", "gauge", gauge_id, name)
    flash(f"계측기 '{name}' {action}됐어.")
    return redirect(url_for("gauge_list"))

@app.route("/gauges/import-methods", methods=["POST"])
@perm_required("gauge")
def gauge_import_methods():
    """자재 규격에 적힌 '측정 방식'들을 계측기 종류로 한 번에 등록한다.
    - 육안/외관/전수 같은 비(非)계측기 값은 뺀다
    - 이미 등록된 이름(공백무시)은 건너뛴다
    - 유효기간은 비워두고 넣는다 → 관리자가 각 종류의 교정 유효기간을 채우면
      검사폼에서 그 종류의 항목들이 유효기간을 자동으로 받아온다"""
    existing = {_norm_method(g["name"]) for g in db.list_gauges()}
    added = []
    for method, _cnt in db.distinct_inspect_methods():
        if method in db.NON_GAUGE_METHODS:
            continue
        key = _norm_method(method)
        if not key or key in existing:
            continue
        # gauge_no는 UNIQUE라 빈 문자열이 여러 개면 충돌 → NULL로 두고 관리번호는 나중에 채움
        db.upsert_gauge(gauge_id=None, gauge_no=None, name=method, model="",
                        location="", last_calibrated=None, expiry_date=None, notes="측정방식 자동등록")
        existing.add(key)
        added.append(method)
    if added:
        record_change("계측기 종류 자동등록", "gauge", None, f"{len(added)}종: " + ", ".join(added[:15]))
        flash(f"측정 방식에서 계측기 종류 {len(added)}개를 등록했어. 각 종류의 교정 유효기간을 채워줘.")
    else:
        flash("새로 등록할 계측기 종류가 없어 (이미 다 등록돼 있어).")
    return redirect(url_for("gauge_list"))


@app.route("/gauges/<int:gauge_id>/delete", methods=["POST"])
@perm_required("gauge")
def gauge_delete(gauge_id):
    g_row = db.get_gauge(gauge_id)
    db.delete_gauge(gauge_id)
    record_change("계측기 삭제", "gauge", gauge_id, g_row["name"] if g_row else "")
    flash("계측기가 삭제됐어.")
    return redirect(url_for("gauge_list"))


# =========================================================================
# 업체 정보 관리
# =========================================================================

@app.route("/suppliers", methods=["GET", "POST"])
@perm_required("supplier")
def supplier_list():
    if request.method == "POST":
        action = request.form.get("action")
        name = request.form.get("name", "").strip()
        if action == "delete" and name:
            db.delete_supplier(name)
            record_change("업체 삭제", "supplier", name, name)
            flash(f"업체 '{name}' 삭제됐어.")
        elif name:
            db.upsert_supplier(name,
                request.form.get("email", "").strip(),
                request.form.get("contact", "").strip(),
                request.form.get("notes", "").strip(),
                request.form.get("address", "").strip(),
                request.form.get("biz_no", "").strip(),
                request.form.get("contact_name", "").strip(),
                request.form.get("contact2", "").strip(),
                request.form.get("items", "").strip())
            record_change("업체 등록/수정", "supplier", name, name)
            flash(f"업체 '{name}' 저장됐어.")
        return redirect(url_for("supplier_list"))
    suppliers = db.list_suppliers()
    return render_template("suppliers.html", suppliers=suppliers)


# =========================================================================
# 부적합 통보서 (NCR)
# =========================================================================

NCR_PHOTO_DIR = os.path.join(db.DATA_DIR, "ncr_photos")
os.makedirs(NCR_PHOTO_DIR, exist_ok=True)


@app.route("/ncr/new", methods=["GET", "POST"])
@perm_required("ncr", "approve")
def ncr_new_manual():
    """자재번호·업체 등을 직접 입력해서 부적합 통보서를 발행하는 화면.

    이 시스템 도입 전(회사 로컬 PC에만 있던) 과거 불량 기록을 지금 통보서로
    남기고 싶을 때 쓴다 — 성적서(inspection_id) 연결 없이도 발행 가능
    (2026-08-31 사용자 요청)."""
    if request.method == "POST":
        material_no = request.form.get("material_no", "").strip()
        material_name = request.form.get("material_name", "").strip()
        supplier = request.form.get("supplier", "").strip()
        if not material_no or not supplier:
            flash("자재번호와 업체는 필수야.")
            return redirect(url_for("ncr_new_manual"))

        ncr_id, ncr_no = db.create_ncr(
            inspection_id=None,
            material_no=material_no,
            material_name=material_name,
            supplier=supplier,
            defect_description=request.form.get("defect_description", "").strip(),
            action_required=request.form.get("action_required", "").strip(),
            due_date=request.form.get("due_date", "").strip(),
            issued_by=g.user["display_name"] or g.user["username"],
            issued_date=request.form.get("issued_date", "").strip(),
            lot_number=request.form.get("lot_number", "").strip() or None,
            receive_date=request.form.get("receive_date", "").strip() or None,
        )
        record_change("부적합 통보서 발행(수기입력)", "ncr", ncr_id,
                      f"{ncr_no} — {material_no} / {supplier}")

        import uuid
        saved_photos = 0
        for file in request.files.getlist("photos"):
            if not file or not file.filename:
                continue
            ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
            if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
                continue
            fname = f"{ncr_no}_{uuid.uuid4().hex[:8]}{ext}"
            try:
                file.save(os.path.join(NCR_PHOTO_DIR, fname))
                db.add_ncr_photo(ncr_id, fname)
                saved_photos += 1
            except Exception:
                pass
        if saved_photos:
            record_change("NCR 사진 첨부", "ncr", ncr_id, f"{saved_photos}장 (작성 시 일괄)")

        flash(f"부적합 통보서 {ncr_no} 발행됐어." + (f" 사진 {saved_photos}장 첨부." if saved_photos else ""))
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    from datetime import date
    return render_template("ncr_manual_form.html", today=date.today().isoformat())


@app.route("/ncr/new/<int:inspection_id>", methods=["GET", "POST"])
@perm_required("ncr", "approve")
def ncr_new(inspection_id):
    header, items = db.get_inspection(inspection_id)
    if header is None:
        flash("성적서를 찾을 수 없어.")
        return redirect(url_for("home"))

    # 불합격 확정 건은 물론, 합격·특채로 이미 승인된 건도 나중에 문제가 발견되면
    # 사후에 부적합 통보서를 쓸 수 있어야 한다(2026-08-30 사용자 요청) — 승인 자체가
    # 안 된 건(대기중·반려)은 아직 판정이 확정 안 된 상태라 통보서 대상이 아니다.
    if header["status"] != "approved":
        flash("승인이 확정된 성적서만 부적합 통보서를 작성할 수 있어.")
        return redirect(url_for("inspection_detail", inspection_id=inspection_id))

    if request.method == "POST":
        ncr_id, ncr_no = db.create_ncr(
            inspection_id=inspection_id,
            material_no=header["material_no"],
            material_name=header["material_name"],
            supplier=header["supplier"],
            defect_description=request.form.get("defect_description", "").strip(),
            action_required=request.form.get("action_required", "").strip(),
            due_date=request.form.get("due_date", "").strip(),
            issued_by=g.user["display_name"] or g.user["username"],
            issued_date=request.form.get("issued_date", "").strip(),
        )
        record_change("부적합 통보서 발행", "ncr", ncr_id,
                      f"{ncr_no} — {header['material_no']} / {header['supplier']}")

        # 사진 첨부(작성 시점에 여러 장 가능) — 실패해도 통보서 발행 자체는 성공 처리
        import uuid
        saved_photos = 0
        for file in request.files.getlist("photos"):
            if not file or not file.filename:
                continue
            ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
            if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
                continue
            fname = f"{ncr_no}_{uuid.uuid4().hex[:8]}{ext}"
            try:
                file.save(os.path.join(NCR_PHOTO_DIR, fname))
                db.add_ncr_photo(ncr_id, fname)
                saved_photos += 1
            except Exception:
                pass
        if saved_photos:
            record_change("NCR 사진 첨부", "ncr", ncr_id, f"{saved_photos}장 (작성 시 일괄)")

        flash(f"부적합 통보서 {ncr_no} 발행됐어." + (f" 사진 {saved_photos}장 첨부." if saved_photos else ""))
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    from datetime import date
    # 부적합 통보서에는 '협력사 귀책'인 항목만 올린다.
    # 규격미입력은 우리 쪽 데이터 누락이므로 업체에 보내는 통보서에 넣으면 안 됨.
    defect_items = [it for it in items
                    if it["result"] not in ("합격", "미측정", "", NO_SPEC_RESULT)]
    supplier_info = db.get_supplier(header["supplier"] or "")
    return render_template("ncr_form.html", header=header, items=defect_items,
                           today=date.today().isoformat(), supplier_info=supplier_info)


@app.route("/ncr")
@perm_required("ncr", "ncr_confirm")
def ncr_list():
    from datetime import date
    status_filter = request.args.get("status", "")
    ncrs = db.list_ncr()
    if status_filter:
        ncrs = [n for n in ncrs if n["status"] == status_filter]

    f = _list_search_params()
    ncrs = [
        n for n in ncrs
        if _row_passes_search(
            f,
            inspector=n["insp_inspector"] or "", supplier=n["supplier"] or "",
            product=n["material_name"] or "", material=n["material_no"] or "",
            result=n["insp_overall_result"] or "",
            status=_approval_status_label(n["insp_status"], n["insp_overall_result"], n["insp_approval_type"]),
            insp_date=n["insp_inspect_date"], recv_date=n["insp_receive_date"],
        )
    ]

    pager = _paginate(ncrs)
    return render_template("ncr_list.html", ncrs=pager["items"], status_filter=status_filter,
                           today=date.today().isoformat(), pager=pager,
                           f=f, result_options=OVERALL_RESULT_OPTIONS, status_options=APPROVAL_STATUS_LABELS)


@app.route("/ncr/<int:ncr_id>")
@perm_required("ncr", "ncr_confirm")
def ncr_detail(ncr_id):
    ncr = db.get_ncr(ncr_id)
    if ncr is None:
        flash("통보서를 찾을 수 없어.")
        return redirect(url_for("home"))
    import json
    photos = json.loads(ncr["photos"] or "[]")
    # 사진은 최대 6장까지만 표시(그 이상은 무시)
    photos = photos[:6]
    supplier_info = db.get_supplier(ncr["supplier"] or "")

    # 연결된 성적서에서 로트번호(po_number)·입고날짜를 가져온다 — 성적서 연결이 없는
    # 수기입력 통보서는 ncr 테이블에 직접 저장된 lot_number/receive_date로 대신한다.
    po_number = ncr["lot_number"] or "" if "lot_number" in ncr.keys() else ""
    receive_date = ncr["receive_date"] or "" if "receive_date" in ncr.keys() else ""
    try:
        insp_header, _ = db.get_inspection(ncr["inspection_id"])
        if insp_header:
            po_number = insp_header["po_number"] or po_number
            receive_date = insp_header["receive_date"] or receive_date
    except Exception:
        pass

    # 도장 여부(사이드카 .stamp 파일 유무) — 상세 화면에서 서명 크기를 다르게 표시
    is_stamp_sig = False
    sig_path_check = ncr["confirm_signature"] if "confirm_signature" in ncr.keys() else None
    if sig_path_check and os.path.exists(sig_path_check + ".stamp"):
        is_stamp_sig = True

    # 확인·발송은 최종결정권자만 — 화면에서도 미리 알려준다
    can_confirm, block_reason = _can_make_final_decision(g.user, "부적합 통보서 확인")

    # 저장된 승인 서명을 문서에 띄우기 위한 URL (static/ 하위 상대경로로 변환)
    confirm_signature_url = None
    sig_path = ncr["confirm_signature"] if "confirm_signature" in ncr.keys() else None
    if sig_path and os.path.exists(sig_path):
        confirm_signature_url = "/static/signatures/" + os.path.basename(sig_path)

    return render_template("ncr_detail.html", ncr=ncr, photos=photos,
                           supplier_info=supplier_info,
                           can_confirm=can_confirm,
                           confirm_block_reason=block_reason or "",
                           confirm_signature_url=confirm_signature_url,
                           po_number=po_number,
                           ncr_receive_date=receive_date,
                           is_stamp_sig=is_stamp_sig,
                           logo_url=url_for("static", filename="logo.png"))


@app.route("/ncr/<int:ncr_id>/confirm", methods=["POST"])
@perm_required("ncr_confirm")
def ncr_confirm(ncr_id):
    ncr = db.get_ncr(ncr_id)
    if ncr is None:
        flash("통보서를 찾을 수 없어.")
        return redirect(url_for("ncr_list"))
    if ncr["status"] != "draft":
        flash("이미 확인 완료된 통보서야.")
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    # 협력사로 나가는 문서이므로 최종결정권자의 승인 서명이 반드시 있어야 한다
    allowed, why = _can_make_final_decision(g.user, "부적합 통보서 확인")
    if not allowed:
        flash(why)
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    # 캔버스 드로잉 우선, 없으면 업로드 파일 → 어느쪽도 없으면 에러
    sig_source = request.form.get("signature_source", "draw")
    stamp_type = request.form.get("signature_stamp_type", "sign")
    signature_path = None
    if sig_source == "upload":
        upload = request.files.get("signature_file")
        signature_path, sig_err = _save_signature_upload(f"ncr{ncr_id}", upload)
    else:
        signature_path, sig_err = _save_signature(f"ncr{ncr_id}",
                                                  request.form.get("signature_data", "").strip())
    if sig_err:
        flash(f"승인 서명이 필요해: {sig_err}")
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    # 도장/사인 사이드카 처리 — 도장이면 .stamp 파일을 남긴다(성적서 로직과 동일)
    if signature_path:
        sidecar = signature_path + ".stamp"
        try:
            if os.path.exists(sidecar):
                os.remove(sidecar)
        except Exception:
            pass
        if stamp_type == "stamp":
            try:
                with open(sidecar, "w", encoding="utf-8") as f:
                    f.write("stamp")
            except Exception:
                pass

    confirmed_by = g.user["display_name"] or g.user["username"]
    db.confirm_ncr(ncr_id, confirmed_by, signature_path=signature_path)
    # 성적서는 이미 '불합격 확정' 상태이므로 여기서 상태를 바꾸지 않는다.
    record_change("NCR 확인 완료", "ncr", ncr_id,
                  f"{ncr['ncr_no']} — 확인자: {confirmed_by} (서명 첨부)")
    flash(f"{ncr['ncr_no']} 확인 완료됐어. 이제 이메일 발송이 가능해.")
    return redirect(url_for("ncr_detail", ncr_id=ncr_id))


@app.route("/ncr/<int:ncr_id>/photo", methods=["POST"])
@perm_required("ncr_confirm")
def ncr_add_photo(ncr_id):
    ncr = db.get_ncr(ncr_id)
    if ncr is None:
        return "not found", 404
    file = request.files.get("photo")
    if not file or file.filename == "":
        flash("사진 파일을 선택해줘.")
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))
    import uuid
    ext = os.path.splitext(file.filename)[1].lower() or ".jpg"
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
        flash("이미지 파일만 첨부할 수 있어.")
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))
    fname = f"{ncr['ncr_no']}_{uuid.uuid4().hex[:8]}{ext}"
    file.save(os.path.join(NCR_PHOTO_DIR, fname))
    db.add_ncr_photo(ncr_id, fname)
    record_change("NCR 사진 첨부", "ncr", ncr_id, fname)
    return redirect(url_for("ncr_detail", ncr_id=ncr_id))


@app.route("/ncr/<int:ncr_id>/email", methods=["POST"])
@perm_required("ncr_confirm")
def ncr_send_email(ncr_id):
    import smtplib, json
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage

    ncr = db.get_ncr(ncr_id)
    if ncr is None:
        flash("통보서를 찾을 수 없어.")
        return redirect(url_for("home"))

    if ncr["status"] == "draft":
        flash("확인 완료 후에 발송할 수 있어. 먼저 '확인 완료' 버튼을 눌러줘.")
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    to_email = request.form.get("to_email", "").strip()
    if not to_email:
        flash("받는 사람 이메일 주소를 입력해줘.")
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    smtp_host = db.get_setting("smtp_host", "")
    smtp_port = int(db.get_setting("smtp_port", "587") or 587)
    smtp_user = db.get_setting("smtp_user", "")
    smtp_pass = db.get_setting("smtp_pass", "")
    from_name = db.get_setting("smtp_from_name", "Chardon QMS")

    if not smtp_host or not smtp_user:
        flash("SMTP 설정이 없어. 설정 → 이메일 설정에서 먼저 입력해줘.")
        return redirect(url_for("ncr_detail", ncr_id=ncr_id))

    photos = json.loads(ncr["photos"] or "[]")
    body_html = render_template("ncr_email_body.html", ncr=ncr)

    msg = MIMEMultipart("related")
    msg["Subject"] = f"[부적합 통보서] {ncr['ncr_no']} — {ncr['material_no']} ({ncr['supplier']})"
    msg["From"] = f"{from_name} <{smtp_user}>"
    msg["To"] = to_email

    alt = MIMEMultipart("alternative")
    msg.attach(alt)
    alt.attach(MIMEText(body_html, "html", "utf-8"))

    for fname in photos:
        fpath = os.path.join(NCR_PHOTO_DIR, fname)
        if os.path.exists(fpath):
            with open(fpath, "rb") as f:
                img = MIMEImage(f.read())
                img.add_header("Content-Disposition", "attachment", filename=fname)
                msg.attach(img)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as s:
            s.ehlo()
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, [to_email], msg.as_string())
        db.mark_ncr_email_sent(ncr_id, to_email)
        record_change("NCR 이메일 발송", "ncr", ncr_id, f"→ {to_email}")
        flash(f"{to_email} 으로 발송됐어.")
    except Exception as e:
        flash(f"이메일 발송 실패: {e}")

    return redirect(url_for("ncr_detail", ncr_id=ncr_id))


@app.route("/settings/email", methods=["GET", "POST"])
@perm_required("smtp")
def smtp_settings():
    if request.method == "POST":
        for key in ("smtp_host", "smtp_port", "smtp_user", "smtp_pass", "smtp_from_name"):
            db.set_setting(key, request.form.get(key, "").strip())
        flash("이메일 설정 저장됐어.")
        return redirect(url_for("smtp_settings"))
    settings = {k: db.get_setting(k, "") for k in
                ("smtp_host", "smtp_port", "smtp_user", "smtp_pass", "smtp_from_name")}
    return render_template("smtp_settings.html", s=settings)


# ---------- 검사원 제외 ----------

@app.route("/inspect/withdraw/<int:intake_id>", methods=["POST"])
@perm_required("inspect_input")
def withdraw_inspector(intake_id):
    name = g.user["display_name"] or g.user["username"]
    progress = db.get_progress_by_intake_ids([intake_id])
    inspectors = progress.get(intake_id, [])
    if len(inspectors) <= 1:
        flash("검사원이 최소 1명은 있어야 해서 제외할 수 없어.")
        return redirect(url_for("inspect_select"))
    remaining = db.withdraw_inspector(intake_id, name)
    record_change("검사원 제외", "intake", intake_id, f"{name} 제외, 남은 검사원: {remaining}")
    flash(f"{name}이(가) 검사에서 제외됐어.")
    return redirect(url_for("inspect_select"))


# ---------- 반품 처리 ----------

@app.route("/returns")
@login_required
def return_list():
    from datetime import date
    status_filter = request.args.get("status", "")
    returns = db.list_return_requests(status=status_filter or None)
    statuses = ["반품요청", "반품완료", "재납품대기", "재검사완료"]

    f = _list_search_params()
    returns = [
        r for r in returns
        if _row_passes_search(
            f,
            inspector=r["insp_inspector"] or "", supplier=r["supplier"] or "",
            product=r["material_name"] or "", material=r["material_no"] or "",
            result=r["insp_overall_result"] or "",
            status=_approval_status_label(r["insp_status"], r["insp_overall_result"], r["insp_approval_type"]),
            insp_date=r["insp_inspect_date"], recv_date=r["insp_receive_date"],
        )
    ]

    pager = _paginate(returns)
    return render_template("return_list.html", returns=pager["items"],
                           status_filter=status_filter, statuses=statuses,
                           today=date.today().isoformat(), pager=pager,
                           f=f, result_options=OVERALL_RESULT_OPTIONS, status_options=APPROVAL_STATUS_LABELS)


@app.route("/return/new/<int:inspection_id>", methods=["GET", "POST"])
@perm_required("return")
def return_new(inspection_id):
    from datetime import date
    header, _ = db.get_inspection(inspection_id)
    if header is None:
        flash("성적서를 찾을 수 없어.")
        return redirect(url_for("home"))
    if request.method == "POST":
        rid = db.create_return_request(
            inspection_id=inspection_id,
            material_no=header["material_no"],
            material_name=header["material_name"],
            supplier=header["supplier"],
            return_date=request.form.get("return_date", "").strip(),
            reason=request.form.get("reason", "").strip(),
            quantity=request.form.get("quantity", "").strip() or None,
            created_by=g.user["display_name"] or g.user["username"],
        )
        record_change("반품 처리 등록", "return", rid,
                      f"{header['material_no']} / {header['supplier']}")
        flash("반품 처리가 등록됐어.")
        return redirect(url_for("return_detail", return_id=rid))
    return render_template("return_form.html", header=header,
                           today=date.today().isoformat())


@app.route("/return/<int:return_id>")
@login_required
def return_detail(return_id):
    rr = db.get_return_request(return_id)
    if rr is None:
        flash("반품 건을 찾을 수 없어.")
        return redirect(url_for("return_list"))
    return render_template("return_detail.html", rr=rr)


@app.route("/return/<int:return_id>/status", methods=["POST"])
@perm_required("return")
def return_update_status(return_id):
    rr = db.get_return_request(return_id)
    if rr is None:
        flash("반품 건을 찾을 수 없어.")
        return redirect(url_for("return_list"))
    new_status = request.form.get("status", "").strip()
    allowed = ["반품요청", "반품완료", "재납품대기", "재검사완료"]
    if new_status not in allowed:
        flash("잘못된 상태야.")
        return redirect(url_for("return_detail", return_id=return_id))
    db.update_return_status(return_id, new_status)
    record_change("반품 상태 변경", "return", return_id, f"→ {new_status}")
    flash(f"상태가 '{new_status}'(으)로 변경됐어.")
    return redirect(url_for("return_detail", return_id=return_id))


# ---------- MA 자동출력 데이터 임포트 ----------

@app.route("/admin/import-assembly", methods=["GET", "POST"])
@perm_required("users")
def import_assembly():
    """MA 자동출력.xlsm의 DATABASE 시트에서 조립 제품 파츠 분해를 임포트."""
    if request.method == "GET":
        return render_template("import_assembly.html")

    file = request.files.get("assembly_file")
    if not file:
        flash("엑셀 파일을 선택해줘.")
        return redirect(url_for("import_assembly"))

    try:
        import tempfile
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp:
                tmp_path = tmp.name
                file.save(tmp.name)
                imported, err = db.import_assembly_from_excel(tmp.name)

            if err:
                flash(f"임포트 실패: {err}")
            else:
                flash(f"{imported}개 MA 조립 제품을 임포트했어.")
                record_change("MA 조립 제품 임포트", "assembly", None, f"{imported}건")
        finally:
            # 임시 파일 정리 시도 (실패해도 무시)
            if tmp_path:
                try:
                    os.remove(tmp_path)
                except:
                    pass
    except Exception as e:
        flash(f"임포트 중 오류: {e}")

    return redirect(url_for("import_assembly"))


# ---------- 조립품 관리 (MA 외 다른 조립품도 직접 등록) ----------

@app.route("/assemblies")
@perm_required("material_view")
def assembly_list():
    """조립품 목록. 입고 때 파츠 하나만 넣어도 전체가 펼쳐지는 기준이 되는 표."""
    return render_template("assembly_list.html", assemblies=db.list_all_assemblies())


@app.route("/assemblies/new", methods=["GET", "POST"])
@app.route("/assemblies/<int:assembly_id>", methods=["GET", "POST"])
@perm_required("material_edit")
def assembly_detail(assembly_id=None):
    if request.method == "POST":
        assembly_no = request.form.get("assembly_no", "").strip()
        # 파츠는 줄바꿈 또는 쉼표로 구분해서 한 번에 붙여넣을 수 있게 받는다
        raw = request.form.get("component_nos", "")
        component_nos = [p.strip() for p in raw.replace(",", "\n").splitlines()]

        new_id, err = db.save_assembly(assembly_no, component_nos, assembly_id=assembly_id)
        if err:
            flash(err)
            return redirect(url_for("assembly_detail", assembly_id=assembly_id) if assembly_id
                            else url_for("assembly_new"))
        action = "조립품 수정" if assembly_id else "조립품 등록"
        record_change(action, "assembly", new_id, f"{assembly_no} ({len([p for p in component_nos if p])}개 파츠)")
        flash(f"'{assembly_no}' 저장 완료.")
        return redirect(url_for("assembly_detail", assembly_id=new_id))

    if assembly_id is None:
        return render_template("assembly_detail.html", master=None, parts=[])

    master, parts = db.get_assembly(assembly_id)
    if master is None:
        flash("존재하지 않는 조립품이야.")
        return redirect(url_for("assembly_list"))
    return render_template("assembly_detail.html", master=master, parts=parts)


# url_for('assembly_new') 로도 부를 수 있게 별칭 엔드포인트를 하나 둔다
app.add_url_rule("/assemblies/new", endpoint="assembly_new",
                 view_func=assembly_detail, methods=["GET", "POST"])


# ---------- 품질 현황 대시보드 ----------

def _resolve_period(period_type, start, end, preset=None):
    """기간 유형/프리셋을 실제 시작·종료일로 바꾼다. 직접 입력한 값이 있으면 그걸 우선."""
    from datetime import date, timedelta
    today = date.today()
    if start and end:
        return start, end

    if preset == "this_month" or period_type == "monthly":
        s = today.replace(day=1)
    elif preset == "today" or period_type == "daily":
        s = today
    elif period_type == "weekly":
        s = today - timedelta(days=today.weekday())
    elif period_type == "quarterly":
        s = date(today.year, ((today.month - 1) // 3) * 3 + 1, 1)
    elif period_type == "half":
        s = date(today.year, 1 if today.month <= 6 else 7, 1)
    else:                       # yearly
        s = date(today.year, 1, 1)

    # 기간 유형별로 "충분히 뒤가 보이도록" 시작점을 조금 넉넉히 잡는다
    lookback = {"daily": 30, "weekly": 84, "monthly": 365,
                "quarterly": 365 * 2, "half": 365 * 3, "yearly": 365 * 5}
    s = min(s, today - timedelta(days=lookback.get(period_type, 365)))
    return s.isoformat(), today.isoformat()


ROW_LIMIT_CHOICES = [20, 50, 100, 300, 0]   # 0 = 전체


def _dashboard_params():
    """대시보드 필터를 URL 쿼리에서 읽어온다.
    업체·발주번호는 화면에서 여러 개 고를 수 있어서 getlist로 받는다(선택 칩 방식)."""
    period_type = request.args.get("period", "monthly")
    if period_type not in [p for p, _ in db.PERIOD_TYPES]:
        period_type = "monthly"
    start, end = _resolve_period(period_type,
                                 request.args.get("start", "").strip(),
                                 request.args.get("end", "").strip())

    def multi(name):
        # 같은 이름이 여러 번 올 수도 있고, 한 번에 콤마로 올 수도 있게 둘 다 받는다
        out = []
        for raw in request.args.getlist(name):
            out += [v.strip() for v in raw.split(",") if v.strip()]
        return list(dict.fromkeys(out))      # 순서 유지하며 중복 제거

    try:
        row_limit = int(request.args.get("rows", 50))
    except ValueError:
        row_limit = 50
    if row_limit not in ROW_LIMIT_CHOICES:
        row_limit = 50

    return {
        "period_type": period_type,
        "start": start,
        "end": end,
        "suppliers": multi("supplier"),
        "po_numbers": multi("po_number"),
        "material": request.args.get("material", "").strip(),
        "states": [s for s in multi("state") if s in db.LOT_STATES],
        "row_limit": row_limit,
    }


def _build_quality_report(p):
    return db.quality_report(
        start_date=p["start"], end_date=p["end"], period_type=p["period_type"],
        supplier=p["suppliers"], po_number=p["po_numbers"],
        material=p["material"] or None, states=p["states"],
    )


def _dashboard_query(p, **override):
    """현재 필터를 URL 쿼리 문자열로 다시 만든다(내보내기 링크·업체 드릴다운에 사용)."""
    from urllib.parse import urlencode
    data = {"period": p["period_type"], "start": p["start"], "end": p["end"],
            "material": p["material"], "rows": p["row_limit"]}
    data.update({k: v for k, v in override.items() if k not in ("supplier", "po_number", "state")})
    pairs = [(k, v) for k, v in data.items() if v not in ("", None)]
    for key, values in (("supplier", override.get("supplier", p["suppliers"])),
                        ("po_number", override.get("po_number", p["po_numbers"])),
                        ("state", override.get("state", p["states"]))):
        pairs += [(key, v) for v in values]
    return urlencode(pairs)


@app.route("/dashboard")
@perm_required("defect_history", "inspect_history")
def quality_dashboard():
    """품질 현황 대시보드 — 기간·업체·발주번호·자재로 걸러서 보고, 팝업으로 근거 문서까지 본다."""
    p = _dashboard_params()
    report = _build_quality_report(p)
    # 발주번호 후보 — 지금 조회 범위 안에 실제로 있는 것만 골라준다
    po_options = sorted({r["po_number"] for r in report["성적서목록"] if r["po_number"]})
    return render_template("dashboard.html",
                           p=p, report=report,
                           period_types=db.PERIOD_TYPES,
                           suppliers=db.list_suppliers(),
                           po_options=po_options,
                           lot_states=db.LOT_STATES,
                           row_choices=ROW_LIMIT_CHOICES,
                           qs=_dashboard_query(p))


@app.route("/dashboard/export.json")
@perm_required("users")
def dashboard_export_json():
    """발표자료·보고서 작성용 원본 데이터. 화면(HTML)을 긁는 것보다 이게 훨씬 정확하다.
    파일만 봐도 무슨 데이터인지 알 수 있게 기간·필터·불량률 기준을 같이 담는다."""
    import json
    p = _dashboard_params()
    report = _build_quality_report(p)
    report.pop("성적서목록", None)      # 원본 행은 너무 길어서 요약 내보내기에선 뺀다
    body = json.dumps(report, ensure_ascii=False, indent=2, default=str)
    fname = f"품질현황_{p['start']}_{p['end']}.json"
    return send_file(io.BytesIO(body.encode("utf-8")), mimetype="application/json",
                     as_attachment=True, download_name=fname)


@app.route("/dashboard/export.xlsx")
@perm_required("defect_history", "inspect_history")
def dashboard_export_xlsx():
    """엑셀 내보내기 — 시트별로 요약/기간별/업체별/자재별/불량항목/성적서목록."""
    from openpyxl import Workbook
    from openpyxl.styles import Font as XFont

    p = _dashboard_params()
    report = _build_quality_report(p)
    wb = Workbook()
    head_font = XFont(bold=True)

    def sheet(title, columns, rows):
        ws = wb.create_sheet(title[:31])
        ws.append(columns)
        for c in ws[1]:
            c.font = head_font
        for r in rows:
            ws.append([r.get(k) for k in columns])
        for idx, col in enumerate(columns, start=1):
            width = max(len(str(col)), *(len(str(r.get(col, ""))) for r in rows)) if rows else len(str(col))
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(40, width + 3)
        return ws

    ws0 = wb.active
    ws0.title = "요약"
    s = report["요약"]
    ws0.append(["항목", "값"])
    ws0["A1"].font = head_font; ws0["B1"].font = head_font
    for k, v in [("보고기간", f"{p['start']} ~ {p['end']}"),
                 ("기간유형", dict(db.PERIOD_TYPES)[p["period_type"]]),
                 ("업체", report["필터"]["업체"]),
                 ("로트번호", report["필터"]["로트번호"]),
                 ("자재/제품명", report["필터"]["자재/제품명"]),
                 ("불량률기준", report["불량률기준"]),
                 ("검사 로트", s["로트"]), ("입고 수량", s["수량"]),
                 ("검사 표본수(실측)", s["검사표본수"]),
                 ("합격 수량", s["합격수량"]), ("특채 수량", s["특채수량"]),
                 ("불합격 수량", s["불합격수량"]), ("미결 수량", s["미결수량"]),
                 ("불량률(%)", s["불량률"]), ("PPM", s["PPM"]),
                 ("규격이탈률(%)", s["규격이탈률"]),
                 ("표본 불량수", s["표본불량수"]),
                 ("표본 불량률(%)", s["표본불량률"]),
                 ("표본 PPM", s["표본PPM"])]:
        ws0.append([k, v])
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 46

    sheet("기간별", ["구간", "로트", "수량", "검사표본수", "합격수량", "특채수량", "불합격수량", "불량률", "PPM", "표본불량수", "표본불량률", "표본PPM"],
          report["기간별"])
    sheet("업체별", ["업체", "로트", "수량", "검사표본수", "합격수량", "특채수량", "불합격수량", "불량률", "PPM", "표본불량수", "표본불량률", "표본PPM"],
          report["업체별"])
    sheet("자재별", ["자재번호", "자재명", "로트", "수량", "검사표본수", "불합격수량", "불량률", "PPM", "표본불량수", "표본불량률", "표본PPM"],
          report["자재별"])
    sheet("불량항목", ["자재번호", "자재명", "항목", "규격", "발생건수", "업체"],
          report["불량항목순위"])
    sheet("성적서목록", ["id", "inspect_date", "material_no", "material_name", "supplier",
                        "po_number", "quantity", "overall_result", "status", "approval_type"],
          report["성적서목록"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"품질현황_{p['start']}_{p['end']}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/dashboard/capability")
@perm_required("defect_history", "inspect_history")
def dashboard_capability():
    """자재 하나의 공정능력(Cp/Cpk) — 대시보드 팝업에서 호출."""
    material_no = request.args.get("material_no", "").strip()
    if not material_no:
        return {"error": "자재번호가 필요해."}, 400
    return {"material_no": material_no, "items": db.process_capability(material_no)}


# ---------- 업체 월간 품질 성적표 (3단계) ----------

@app.route("/supplier-reports")
@perm_required("defect_history", "supplier")
def supplier_report_list():
    can_approve, block_reason = _can_make_final_decision(g.user, "업체 성적표 승인")
    return render_template("supplier_reports.html",
                           reports=db.list_supplier_reports(),
                           suppliers=db.list_suppliers(),
                           default_period=_dt.now().strftime("%Y-%m"),
                           can_approve=can_approve,
                           approve_block_reason=block_reason or "")


@app.route("/supplier-reports/generate", methods=["POST"])
@perm_required("defect_history", "supplier")
def supplier_report_generate():
    """업체 하나의 월간 성적표를 만든다 — 대시보드 집계를 그대로 재사용."""
    import json, calendar
    supplier = request.form.get("supplier", "").strip()
    period = request.form.get("period", "").strip()          # YYYY-MM
    if not supplier or len(period) != 7:
        flash("업체와 기간(YYYY-MM)을 정확히 골라줘.")
        return redirect(url_for("supplier_report_list"))
    try:
        y, m = int(period[:4]), int(period[5:7])
        start = f"{y:04d}-{m:02d}-01"
        end = f"{y:04d}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"
    except ValueError:
        flash("기간 형식이 잘못됐어 (YYYY-MM).")
        return redirect(url_for("supplier_report_list"))

    report = db.quality_report(start, end, period_type="monthly", supplier=supplier)
    report.pop("성적서목록", None)
    rid, err = db.upsert_supplier_report(
        supplier, period, start, end,
        json.dumps(report, ensure_ascii=False, default=str),
        created_by=g.user["display_name"] or g.user["username"])
    if err:
        flash(err)
    else:
        flash(f"{supplier} {period} 성적표를 만들었어. 최종결정권자 승인 후 발송할 수 있어.")
        record_change("업체 성적표 생성", "supplier_report", rid, f"{supplier} {period}")
    return redirect(url_for("supplier_report_detail", report_id=rid))


@app.route("/supplier-reports/<int:report_id>")
@perm_required("defect_history", "supplier")
def supplier_report_detail(report_id):
    import json
    row = db.get_supplier_report(report_id)
    if row is None:
        flash("성적표를 찾을 수 없어.")
        return redirect(url_for("supplier_report_list"))
    try:
        payload = json.loads(row["payload"])
    except ValueError:
        payload = None
    can_approve, block_reason = _can_make_final_decision(g.user, "업체 성적표 승인")
    sig_url = None
    if row["approve_signature"] and os.path.exists(row["approve_signature"]):
        sig_url = "/static/signatures/" + os.path.basename(row["approve_signature"])
    return render_template("supplier_report_detail.html",
                           r=row, payload=payload,
                           supplier_info=db.get_supplier(row["supplier"] or ""),
                           can_approve=can_approve,
                           approve_block_reason=block_reason or "",
                           signature_url=sig_url)


@app.route("/supplier-reports/<int:report_id>/approve", methods=["POST"])
@perm_required("approve")
def supplier_report_approve(report_id):
    """협력사로 나가는 문서라 최종결정권자 승인 + 서명이 있어야 발송할 수 있다."""
    row = db.get_supplier_report(report_id)
    if row is None or row["status"] != "draft":
        flash("승인할 수 있는 상태가 아니야.")
        return redirect(url_for("supplier_report_list"))

    allowed, why = _can_make_final_decision(g.user, "업체 성적표 승인")
    if not allowed:
        flash(why)
        return redirect(url_for("supplier_report_detail", report_id=report_id))

    sig_path, sig_err = _save_signature(f"sr{report_id}",
                                        request.form.get("signature_data", "").strip())
    if sig_err:
        flash(f"승인 서명이 필요해: {sig_err}")
        return redirect(url_for("supplier_report_detail", report_id=report_id))

    approver = g.user["display_name"] or g.user["username"]
    db.approve_supplier_report(report_id, approver, sig_path)
    record_change("업체 성적표 승인", "supplier_report", report_id,
                  f"{row['supplier']} {row['period']} — 승인자 {approver}")
    flash("승인 완료. 이제 발송할 수 있어.")
    return redirect(url_for("supplier_report_detail", report_id=report_id))


@app.route("/supplier-reports/<int:report_id>/send", methods=["POST"])
@perm_required("approve", "smtp")
def supplier_report_send(report_id):
    row = db.get_supplier_report(report_id)
    if row is None or row["status"] != "approved":
        flash("승인된 성적표만 발송할 수 있어.")
        return redirect(url_for("supplier_report_list"))
    to_email = request.form.get("to_email", "").strip()
    if not to_email:
        flash("받는 사람 이메일을 입력해줘.")
        return redirect(url_for("supplier_report_detail", report_id=report_id))

    db.mark_supplier_report_sent(report_id, to_email)
    record_change("업체 성적표 발송", "supplier_report", report_id,
                  f"{row['supplier']} {row['period']} → {to_email}")
    flash(f"{to_email} 로 발송 처리했어.")
    return redirect(url_for("supplier_report_detail", report_id=report_id))


@app.route("/supplier-reports/<int:report_id>/delete", methods=["POST"])
@perm_required("defect_history", "supplier")
def supplier_report_delete(report_id):
    db.delete_supplier_report(report_id)
    flash("초안 성적표를 삭제했어.")
    return redirect(url_for("supplier_report_list"))


# ---------- 4M 변경점 관리 ----------

@app.route("/change-points", methods=["GET", "POST"])
@perm_required("supplier", "material_edit")
def change_point_list():
    """협력사의 4M(사람·설비·자재·방법) 변경 시점을 기록한다.
    변경 전후 불량률을 비교하려면 '언제 바뀌었는지'가 남아 있어야 한다."""
    if request.method == "POST":
        supplier = request.form.get("supplier", "").strip()
        change_type = request.form.get("change_type", "").strip()
        change_date = request.form.get("change_date", "").strip()
        if not (supplier and change_type and change_date):
            flash("업체 · 변경 구분 · 변경일은 필수야.")
            return redirect(url_for("change_point_list"))
        if change_type not in [c for c, _ in db.CHANGE_TYPES]:
            flash("알 수 없는 변경 구분이야.")
            return redirect(url_for("change_point_list"))

        cp_id = db.add_change_point(
            supplier=supplier,
            material_no=request.form.get("material_no", "").strip(),
            change_type=change_type,
            change_date=change_date,
            description=request.form.get("description", "").strip(),
            reported_by=g.user["display_name"] or g.user["username"],
        )
        record_change("4M 변경점 등록", "change_point", cp_id,
                      f"{supplier} / {change_type} / {change_date}")
        flash("변경점이 등록됐어. 이후 검사에서 이 업체 자재를 열면 경고로 표시돼.")
        return redirect(url_for("change_point_list"))

    supplier_filter = request.args.get("supplier", "").strip()
    rows = db.list_change_points(supplier=supplier_filter or None)
    return render_template("change_points.html",
                           rows=rows,
                           suppliers=db.list_suppliers(),
                           change_types=db.CHANGE_TYPES,
                           supplier_filter=supplier_filter,
                           today=_dt.now().strftime("%Y-%m-%d"))


@app.route("/change-points/<int:cp_id>/delete", methods=["POST"])
@perm_required("supplier", "material_edit")
def change_point_delete(cp_id):
    db.delete_change_point(cp_id)
    record_change("4M 변경점 삭제", "change_point", cp_id, "")
    flash("변경점이 삭제됐어.")
    return redirect(url_for("change_point_list"))


# ---------- 성적서 위변조 검증 ----------

@app.route("/inspection/<int:inspection_id>/verify")
@perm_required("inspect_history", "approve")
def inspection_verify(inspection_id):
    """성적서가 승인 당시 그대로인지 검증한 결과를 JSON으로 돌려준다."""
    return verify_inspection_integrity(inspection_id)


# ---------- 데이터 점검 (관리자 전용) ----------

@app.route("/admin/data-health")
@perm_required("users")
def data_health():
    """자재·규격 데이터에서 검사/성적서를 망가뜨릴 수 있는 것들을 한 화면에 모아서 보여준다.
    관리자('users' 권한)만 접근 가능."""
    return render_template("data_health.html", checks=db.data_health_report())


# ---------- 데이터 이전 (로컬 -> 클라우드 서버, 관리자 전용) ----------

@app.route("/admin/data-migrate", methods=["GET", "POST"])
@perm_required("users")
def data_migrate():
    """로컬 PC에서 만든 iqc.db + 첨부파일 zip을 업로드해서 이 서버의 DATA_DIR에 통째로 덮어쓴다.
    Render 등 클라우드로 운영 데이터를 처음 옮길 때 1회성으로 쓰는 화면."""
    if request.method == "GET":
        return render_template("data_migrate.html", data_dir=db.DATA_DIR)

    zfile = request.files.get("zipfile")
    if not zfile or not zfile.filename:
        flash("zip 파일을 선택해줘.")
        return redirect(url_for("data_migrate"))

    with tempfile.TemporaryDirectory() as tmp:
        zpath = os.path.join(tmp, "upload.zip")
        zfile.save(zpath)

        try:
            zf = zipfile.ZipFile(zpath)
            names = zf.namelist()
        except zipfile.BadZipFile:
            flash("zip 파일이 손상됐거나 zip 형식이 아니야.")
            return redirect(url_for("data_migrate"))

        has_db = "iqc.db" in names
        # 부속 폴더(도면 등)만 추가로 올릴 때도 쓸 수 있게, iqc.db는 있으면 갈아끼우고 없으면 건너뛴다.

        # 안전장치: DB를 덮어쓰기 전에 현재 DB를 먼저 백업해둔다
        if has_db and os.path.exists(db.DB_PATH):
            ts = _dt.now().strftime("%Y%m%d_%H%M%S")
            pre_backup = os.path.join(BACKUP_DIR, f"iqc_전이전백업_{ts}.db")
            try:
                shutil.copy(db.DB_PATH, pre_backup)
            except Exception:
                pass

        # 경로 조작(zip slip) 방지 — 안전한 상대경로만 허용
        safe_names = [n for n in names if not n.startswith("/") and ".." not in n.split("/")]
        for n in safe_names:
            zf.extract(n, tmp)

        moved = []
        if has_db:
            extracted_db = os.path.join(tmp, "iqc.db")
            shutil.copy(extracted_db, db.DB_PATH)
            moved.append("iqc.db")
        for folder, dest in (
            ("signatures", SIGNATURE_DIR),
            ("ncr_photos", NCR_PHOTO_DIR),
            ("backups", BACKUP_DIR),
            ("성적서 발행", report_builder.OUT_DIR),
            ("도면", DRAWING_DIR),
        ):
            src = os.path.join(tmp, folder)
            if os.path.isdir(src):
                for name in os.listdir(src):
                    s, d = os.path.join(src, name), os.path.join(dest, name)
                    if os.path.isdir(s):
                        shutil.copytree(s, d, dirs_exist_ok=True)
                    else:
                        shutil.copy(s, d)
                moved.append(folder)

    if not moved:
        flash("zip 안에 iqc.db나 알아보는 폴더(signatures/ncr_photos/backups/성적서 발행/도면)가 하나도 없어.")
        return redirect(url_for("data_migrate"))

    flash(f"이전 완료: {', '.join(moved)}. 새로고침해서 데이터 확인해봐.")
    return redirect(url_for("home"))


@app.route("/assemblies/<int:assembly_id>/delete", methods=["POST"])
@perm_required("material_edit")
def assembly_delete(assembly_id):
    master, _ = db.get_assembly(assembly_id)
    if master is None:
        flash("존재하지 않는 조립품이야.")
        return redirect(url_for("assembly_list"))
    db.delete_assembly(assembly_id)
    record_change("조립품 삭제", "assembly", assembly_id, master["assembly_no"])
    flash(f"'{master['assembly_no']}' 삭제됐어.")
    return redirect(url_for("assembly_list"))


db.init_db()
ensure_default_admin()
ensure_perm_migration()
ensure_inspect_method_fill_20260825()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
